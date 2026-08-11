#!/usr/bin/env python3
"""Generate a lab shift rota as a formatted Excel workbook.

Two steps:

    python lab_roster.py template --out roster_input.xlsx
    python lab_roster.py generate --input roster_input.xlsx --out rota.xlsx

`template` writes an input workbook you fill in (staff, shifts, leave, benches,
settings).  `generate` reads it, builds a schedule and writes a rota workbook
with a colour-coded calendar grid plus a per-person summary.

Hard rules the scheduler will not break:
  * nobody is scheduled on a day they are on leave
  * one shift per person per day
  * no shift on the day after a night (this covers the night -> early ban)
  * no more than `Max consecutive days` worked in an unbroken run
  * nights are placed in blocks (default 3) followed by rest days (default 2)

Soft targets, in priority order:
  * total shifts spread as evenly as possible across schedulable staff
  * weekend duty spread as evenly as possible
  * at least one Band 6+ senior on every shift flagged "Requires senior"
  * bench/section skill coverage each day

Only third-party dependency is openpyxl.
"""

from __future__ import annotations

import argparse
import random
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.properties import PageSetupProperties
except ImportError:  # pragma: no cover - environment problem, not a code path
    sys.exit("openpyxl is required.  Install it with:  pip install openpyxl")


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------

WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday",
                 "Saturday", "Sunday"]


def norm(text) -> str:
    """Normalise a header or code for tolerant matching."""
    return re.sub(r"[^a-z0-9]", "", str(text or "").lower())


def to_date(value, label="date"):
    """Accept a real date, a datetime, or a dd/mm/yyyy-ish string."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y",
                "%d.%m.%y", "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Could not read {label} from {value!r}. Use dd/mm/yyyy.")


def to_int(value, default=0):
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def to_float(value, default=0.0):
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def is_yes(value, default=False) -> bool:
    text = str(value).strip().lower()
    if text in ("y", "yes", "true", "1", "t"):
        return True
    if text in ("n", "no", "false", "0", "f"):
        return False
    return default


def split_list(value) -> list[str]:
    if value is None:
        return []
    return [part.strip() for part in re.split(r"[,;/|]", str(value)) if part.strip()]


def daterange(start: date, end: date):
    day = start
    while day <= end:
        yield day
        day += timedelta(days=1)


def make_initials(names: list[str]) -> dict[str, str]:
    """Two-letter initials, extended where they would collide (AS -> ASa/ASp)."""
    proposed: dict[str, str] = {}
    for name in names:
        parts = [p for p in re.split(r"\s+", name.strip()) if p]
        if not parts:
            proposed[name] = "??"
        elif len(parts) == 1:
            proposed[name] = parts[0][:2].upper()
        else:
            proposed[name] = (parts[0][0] + parts[-1][0]).upper()

    by_code: dict[str, list[str]] = defaultdict(list)
    for name, code in proposed.items():
        by_code[code].append(name)

    final: dict[str, str] = {}
    for code, owners in by_code.items():
        if len(owners) == 1:
            final[owners[0]] = code
            continue
        for owner in owners:
            parts = [p for p in re.split(r"\s+", owner.strip()) if p]
            surname = parts[-1] if len(parts) > 1 else parts[0]
            extended = code + surname[1:2].lower()
            suffix = 2
            while extended in final.values():
                extended = code + surname[1:1 + suffix].lower()
                suffix += 1
                if suffix > len(surname):
                    extended = f"{code}{len(final) + 1}"
                    break
            final[owner] = extended
    return final


# --------------------------------------------------------------------------
# data model
# --------------------------------------------------------------------------

@dataclass
class Shift:
    code: str
    name: str
    start: str = ""
    end: str = ""
    hours: float = 0.0
    days: str = "All"            # Weekday | Weekend | All
    required: int = 1
    requires_senior: bool = True
    is_night: bool = False
    colour: str = "D9E1F2"
    font_colour: str = "000000"

    def applies_on(self, day: date, weekend_days: set[int]) -> bool:
        weekend = day.weekday() in weekend_days
        scope = norm(self.days)
        if scope in ("all", "everyday", "daily", ""):
            return True
        if scope.startswith("weekend"):
            return weekend
        if scope.startswith("weekday"):
            return not weekend
        return True

    @property
    def label(self) -> str:
        when = f" ({self.start}-{self.end})" if self.start and self.end else ""
        return f"{self.name}{when}"


@dataclass
class Staff:
    name: str
    band: str = "5"
    skills: list[str] = field(default_factory=list)
    nights_ok: bool = True
    group: str = "Main"
    max_shifts: int = 0           # 0 = no personal cap
    notes: str = ""

    @property
    def band_value(self) -> float:
        """Band 8a -> 8.1, 8b -> 8.2 so ordering behaves; '6' -> 6.0."""
        text = str(self.band).strip().lower().replace("band", "").strip()
        match = re.match(r"(\d+)\s*([a-d])?", text)
        if not match:
            return 0.0
        value = float(match.group(1))
        if match.group(2):
            value += (ord(match.group(2)) - ord("a") + 1) / 10.0
        return value

    def has_skill(self, skill: str) -> bool:
        target = norm(skill)
        return any(norm(s) == target for s in self.skills)


@dataclass
class LeaveEntry:
    name: str
    start: date
    end: date
    code: str


@dataclass
class Bench:
    name: str
    skill: str
    days: str = "All"
    min_staff: int = 1
    min_weekend: int | None = None    # None = same as min_staff

    def applies_on(self, day: date, weekend_days: set[int]) -> bool:
        return Shift(code="_", name="_", days=self.days).applies_on(day, weekend_days)

    def required_on(self, day: date, weekend_days: set[int]) -> int:
        """Weekends usually run a skeleton service, so they get their own minimum."""
        if day.weekday() in weekend_days and self.min_weekend is not None:
            return self.min_weekend
        return self.min_staff


@dataclass
class LeaveType:
    code: str
    label: str
    colour: str = "FFFF00"
    font_colour: str = "000000"


@dataclass
class Settings:
    title: str = "Lab Shift Rota"
    start: date = date.today()
    end: date = date.today()
    senior_band: float = 6.0
    max_consecutive: int = 6
    night_block: int = 3
    rest_after_nights: int = 2
    weekend_days: set[int] = field(default_factory=lambda: {5, 6})
    seed: int = 42


@dataclass
class Config:
    settings: Settings
    staff: list[Staff]
    shifts: list[Shift]
    leave: list[LeaveEntry]
    benches: list[Bench]
    leave_types: dict[str, LeaveType]


# --------------------------------------------------------------------------
# styling constants
# --------------------------------------------------------------------------

THIN = Side(style="thin", color="B0B0B0")
MEDIUM = Side(style="medium", color="404040")
GRID_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTRE = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP_CENTRE = Alignment(horizontal="center", vertical="center", wrap_text=True)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
MONTH_FILL = PatternFill("solid", fgColor="2E5496")
DAY_FILL = PatternFill("solid", fgColor="D6DCE4")
WEEKEND_FILL = PatternFill("solid", fgColor="F2E6D9")
GROUP_FILL = PatternFill("solid", fgColor="BFBFBF")
BENCH_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)

FIRST_DATE_COL = 4      # A=Staff, B=Band, C=Initials, D onwards = dates
TITLE_ROW = 1
MONTH_ROW = 2
DAYNUM_ROW = 3
WEEKDAY_ROW = 4
FIRST_STAFF_ROW = 5


# --------------------------------------------------------------------------
# input template
# --------------------------------------------------------------------------

DEFAULT_SHIFTS = [
    # code, name,      start,  end,    hours, days,     req, senior, night, colour,   font
    ("C", "Core",      "09:00", "17:30", 8.0, "Weekday", 6, "Y", "N", "D9E1F2", "000000"),
    ("E", "Early",     "07:00", "15:00", 7.5, "Weekday", 2, "N", "N", "FFF2CC", "000000"),
    ("L", "Late",      "13:00", "21:00", 7.5, "Weekday", 2, "Y", "N", "00B0F0", "000000"),
    ("N", "Night",     "21:00", "07:00", 10.0, "All",    1, "Y", "Y", "7030A0", "FFFFFF"),
    ("W", "Weekend",   "09:00", "17:30", 8.0, "Weekend", 3, "Y", "N", "00B050", "FFFFFF"),
]

DEFAULT_LEAVE_TYPES = [
    ("A/L", "Annual leave", "FFFF00", "000000"),
    ("S/L", "Sick leave", "FF9999", "000000"),
    ("C/L", "Carers or compassionate leave", "FFC000", "000000"),
    ("M/L", "Maternity leave", "D9D9D9", "000000"),
    ("S/D", "Study day", "CC99FF", "000000"),
]

DEFAULT_BENCHES = [
    # bench, skill, days, min staff (weekday), min staff (weekend)
    ("Blood Transfusion", "BT", "All", 2, 1),
    ("Coagulation", "COAG", "All", 1, 1),
    ("Haematology", "HAEM", "All", 2, 1),
    ("Morphology", "MORPH", "Weekday", 1, 0),
]

# Every name here is fictional and deliberately obvious about it.  Never put
# real staff names in the template: it ships as example data and would carry
# personal data into every copy of the tool.  Bands, skills and night flags are
# placeholders too - see the warning written into the Staff sheet.
# "Alex Sample" and "Avery Specimen" both reduce to AS on purpose, so the
# initials de-duplication is exercised by the default data.
DEFAULT_STAFF = [
    # name, band, skills, nights, group
    ("Alex Sample", "7", "HAEM, MORPH, COAG", "N", "Main"),
    ("Jordan Test", "6", "HAEM, BT", "Y", "Main"),
    ("Casey Example", "6", "BT, COAG", "Y", "Main"),
    ("Morgan Demo", "6", "HAEM, BT, COAG", "Y", "Main"),
    ("Riley Placeholder", "6", "HAEM, MORPH", "Y", "Main"),
    ("Jamie Mock", "5", "HAEM, COAG", "Y", "Main"),
    ("Taylor Dummy", "6", "BT, HAEM", "Y", "Main"),
    ("Avery Specimen", "6", "HAEM, COAG, BT", "Y", "Main"),
    ("Quinn Template", "6", "MORPH, HAEM", "Y", "Main"),
    ("Rowan Draft", "5", "BT, COAG", "Y", "Main"),
    ("Skyler Trial", "5", "HAEM, BT", "Y", "Main"),
    ("Parker Proxy", "5", "HAEM, COAG", "Y", "Main"),
    ("Reese Stub", "5", "BT, HAEM", "Y", "Main"),
    ("Devon Filler", "6", "HAEM, COAG, MORPH", "Y", "Main"),
    ("Harper Model", "5", "BT, COAG", "Y", "Main"),
    ("Emerson Pattern", "5", "HAEM, BT", "Y", "Main"),
    ("Finley Sketch", "5", "COAG, HAEM", "Y", "Main"),
    ("Marlow Outline", "5", "BT, MORPH", "Y", "Main"),
    ("Sasha Preview", "5", "HAEM, BT", "Y", "Main"),
    ("Tatum Instance", "5", "COAG, BT", "Y", "Main"),
    ("Blake Notional", "4", "BT", "N", "Main"),
    ("Corey Blank", "4", "HAEM", "N", "Main"),
    ("Drew Dummy", "6", "HAEM, BT, COAG", "Y", "Main"),
    ("Ellis Nominal", "5", "HAEM, COAG, MORPH", "Y", "Main"),
    ("Frankie Generic", "4", "BT, HAEM", "N", "Main"),
    ("Gray Placeholder", "3", "BT", "N", "Extras"),
    ("Hollis Sample", "3", "HAEM", "N", "Extras"),
    ("Indigo Demo", "4", "BT, COAG", "N", "Extras"),
    ("Jesse Example", "4", "HAEM", "N", "Extras"),
    ("Kai Mock", "3", "BT", "N", "Extras"),
    ("Lennox Draft", "5", "HAEM, BT", "N", "Main"),
    ("Marley Test", "5", "COAG", "N", "Main"),
    ("Noor Specimen", "5", "BT", "N", "Main"),
]

INSTRUCTIONS = [
    ("Lab Shift Roster - input workbook", True),
    ("", False),
    ("Fill in the sheets below, save, then run:", False),
    ("    python lab_roster.py generate --input <this file> --out rota.xlsx", False),
    ("", False),
    ("Settings", True),
    ("Date range, rule limits and the random seed. Same seed + same inputs = same rota.", False),
    ("", False),
    ("Staff", True),
    ("One row per person. Band drives seniority (Band 6+ counts as senior by default).", False),
    ("Skills must match the codes used on the Benches sheet (BT, COAG, HAEM, MORPH).", False),
    ("Nights Y/N excludes someone from night duty. Group splits the grid into blocks", False),
    ("(Main, Extras, ...). Max shifts caps one person's total; leave 0 for no cap.", False),
    ("", False),
    ("Shifts", True),
    ("One row per shift type. Required = how many people per day. Days limits a shift to", False),
    ("Weekday, Weekend or All. Requires senior forces a Band 6+ into the first slot.", False),
    ("Is night marks the shift as a night for the block/rest and next-day rules.", False),
    ("Colour is a 6-digit hex fill used in the rota grid and the key.", False),
    ("", False),
    ("Leave", True),
    ("One row per absence. From and To are inclusive. Type must match a Leave types code.", False),
    ("Nobody is ever scheduled inside a leave range.", False),
    ("", False),
    ("Benches", True),
    ("Section coverage shown under the calendar grid. Skill must match a staff skill code.", False),
    ("Min staff weekend lets a bench run a skeleton service at weekends; 0 closes it.", False),
    ("", False),
    ("Leave types", True),
    ("Codes and colours used in the grid and key for absence.", False),
]


def build_template(path: Path) -> None:
    workbook = Workbook()

    # ---- Instructions -------------------------------------------------
    sheet = workbook.active
    sheet.title = "Instructions"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 100
    for index, (text, is_heading) in enumerate(INSTRUCTIONS, start=1):
        cell = sheet.cell(row=index, column=1, value=text)
        if is_heading:
            cell.font = Font(bold=True, size=13 if index == 1 else 11,
                             color="1F3864")

    def header(ws, titles, note=None, widths=None):
        ws.sheet_view.showGridLines = False
        row = 1
        if note:
            cell = ws.cell(row=1, column=1, value=note)
            cell.font = Font(bold=True, color="C00000")
            row = 2
        for col, title in enumerate(titles, start=1):
            cell = ws.cell(row=row, column=col, value=title)
            cell.fill = HEADER_FILL
            cell.font = WHITE_BOLD
            cell.alignment = WRAP_CENTRE
            cell.border = GRID_BORDER
        for col, width in enumerate(widths or [], start=1):
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
        return row

    # ---- Settings -----------------------------------------------------
    settings_sheet = workbook.create_sheet("Settings")
    header(settings_sheet, ["Setting", "Value"], widths=[34, 24])
    today = date.today()
    first = date(today.year + (today.month == 12), (today.month % 12) + 1, 1)
    last_day = (date(first.year + (first.month == 12), (first.month % 12) + 1, 1)
                - timedelta(days=1))
    settings_rows = [
        ("Rota title", "GEH BMS Lab Rota"),
        ("Start date", first),
        ("End date", last_day),
        ("Senior band threshold", 6),
        ("Max consecutive days", 6),
        ("Night block length", 3),
        ("Rest days after nights", 2),
        ("Weekend days", "Saturday, Sunday"),
        ("Random seed", 42),
    ]
    for index, (key, value) in enumerate(settings_rows, start=2):
        settings_sheet.cell(row=index, column=1, value=key).font = BOLD
        cell = settings_sheet.cell(row=index, column=2, value=value)
        cell.border = GRID_BORDER
        if isinstance(value, date):
            cell.number_format = "DD/MM/YYYY"

    # ---- Staff --------------------------------------------------------
    staff_sheet = workbook.create_sheet("Staff")
    head_row = header(
        staff_sheet,
        ["Name", "Band", "Skills", "Nights (Y/N)", "Group", "Max shifts", "Notes"],
        note="EXAMPLE DATA - every name below is fictional, and the bands, "
             "skills and night flags are made up. Replace all of it with your "
             "real staff details before generating a rota you intend to use.",
        widths=[24, 8, 26, 13, 12, 12, 30],
    )
    for offset, (name, band, skills, nights, group) in enumerate(DEFAULT_STAFF):
        row = head_row + 1 + offset
        for col, value in enumerate([name, band, skills, nights, group, 0, ""], start=1):
            cell = staff_sheet.cell(row=row, column=col, value=value)
            cell.border = GRID_BORDER
    yes_no = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    staff_sheet.add_data_validation(yes_no)
    yes_no.add(f"D{head_row + 1}:D400")

    # ---- Shifts -------------------------------------------------------
    shifts_sheet = workbook.create_sheet("Shifts")
    head_row = header(
        shifts_sheet,
        ["Code", "Name", "Start", "End", "Hours", "Days", "Required",
         "Requires senior", "Is night", "Colour", "Font colour"],
        widths=[8, 14, 8, 8, 8, 11, 10, 15, 10, 10, 12],
    )
    for offset, values in enumerate(DEFAULT_SHIFTS):
        row = head_row + 1 + offset
        for col, value in enumerate(values, start=1):
            cell = shifts_sheet.cell(row=row, column=col, value=value)
            cell.border = GRID_BORDER
        swatch = shifts_sheet.cell(row=row, column=10)
        swatch.fill = PatternFill("solid", fgColor=values[9])
        swatch.font = Font(color=values[10], bold=True)
    scope = DataValidation(type="list", formula1='"Weekday,Weekend,All"',
                          allow_blank=True)
    shifts_sheet.add_data_validation(scope)
    scope.add(f"F{head_row + 1}:F100")
    shift_yes_no = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    shifts_sheet.add_data_validation(shift_yes_no)
    shift_yes_no.add(f"H{head_row + 1}:I100")

    # ---- Leave --------------------------------------------------------
    leave_sheet = workbook.create_sheet("Leave")
    head_row = header(leave_sheet, ["Name", "From", "To", "Type"],
                      widths=[24, 14, 14, 10])
    # Names must match the fictional DEFAULT_STAFF entries above.
    example_leave = [
        ("Alex Sample", first + timedelta(days=7), first + timedelta(days=11), "A/L"),
        ("Riley Placeholder", first + timedelta(days=14), first + timedelta(days=18), "A/L"),
        ("Parker Proxy", first + timedelta(days=2), first + timedelta(days=3), "S/L"),
        ("Lennox Draft", first, last_day, "M/L"),
        ("Marley Test", first, last_day, "M/L"),
        ("Noor Specimen", first, last_day, "M/L"),
    ]
    for offset, (name, start, end, code) in enumerate(example_leave):
        row = head_row + 1 + offset
        for col, value in enumerate([name, start, end, code], start=1):
            cell = leave_sheet.cell(row=row, column=col, value=value)
            cell.border = GRID_BORDER
            if isinstance(value, date):
                cell.number_format = "DD/MM/YYYY"

    # ---- Benches ------------------------------------------------------
    bench_sheet = workbook.create_sheet("Benches")
    head_row = header(bench_sheet,
                      ["Bench", "Skill", "Days", "Min staff", "Min staff weekend"],
                      widths=[24, 12, 12, 12, 18])
    for offset, values in enumerate(DEFAULT_BENCHES):
        row = head_row + 1 + offset
        for col, value in enumerate(values, start=1):
            bench_sheet.cell(row=row, column=col, value=value).border = GRID_BORDER

    # ---- Leave types --------------------------------------------------
    types_sheet = workbook.create_sheet("Leave types")
    head_row = header(types_sheet, ["Code", "Label", "Colour", "Font colour"],
                      widths=[10, 34, 12, 12])
    for offset, values in enumerate(DEFAULT_LEAVE_TYPES):
        row = head_row + 1 + offset
        for col, value in enumerate(values, start=1):
            types_sheet.cell(row=row, column=col, value=value).border = GRID_BORDER
        swatch = types_sheet.cell(row=row, column=3)
        swatch.fill = PatternFill("solid", fgColor=values[2])
        swatch.font = Font(color=values[3], bold=True)

    workbook.save(path)


# --------------------------------------------------------------------------
# reading the input workbook
# --------------------------------------------------------------------------

def read_table(sheet, required: list[str]) -> list[dict]:
    """Find the header row (scanning the first 10 rows) and return dict rows."""
    wanted = [norm(name) for name in required]
    header_row = None
    columns: dict[str, int] = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=10), start=1):
        found = {norm(cell.value): cell.column for cell in row if cell.value is not None}
        if all(name in found for name in wanted):
            header_row = row_index
            columns = found
            break
    if header_row is None:
        raise ValueError(
            f"Sheet '{sheet.title}' needs a header row containing: "
            f"{', '.join(required)}"
        )

    records = []
    for row in sheet.iter_rows(min_row=header_row + 1):
        record = {key: sheet.cell(row=row[0].row, column=col).value
                  for key, col in columns.items()}
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def load_config(path: Path) -> Config:
    workbook = load_workbook(path, data_only=True)

    def sheet_for(*candidates):
        for candidate in candidates:
            for title in workbook.sheetnames:
                if norm(title) == norm(candidate):
                    return workbook[title]
        raise ValueError(
            f"Input workbook has no '{candidates[0]}' sheet "
            f"(found: {', '.join(workbook.sheetnames)})"
        )

    # ---- settings -----------------------------------------------------
    raw_settings: dict[str, object] = {}
    for record in read_table(sheet_for("Settings"), ["Setting", "Value"]):
        key = norm(record.get("setting"))
        if key:
            raw_settings[key] = record.get("value")

    def setting(key, default=None):
        value = raw_settings.get(norm(key))
        return default if value in (None, "") else value

    weekend_text = str(setting("Weekend days", "Saturday, Sunday"))
    weekend_days = {WEEKDAY_NAMES.index(name)
                    for token in split_list(weekend_text)
                    for name in WEEKDAY_NAMES
                    if norm(name).startswith(norm(token)) and len(norm(token)) >= 3}
    if not weekend_days:
        weekend_days = {5, 6}

    settings = Settings(
        title=str(setting("Rota title", "Lab Shift Rota")),
        start=to_date(setting("Start date"), "Start date"),
        end=to_date(setting("End date"), "End date"),
        senior_band=to_float(setting("Senior band threshold", 6), 6.0),
        max_consecutive=to_int(setting("Max consecutive days", 6), 6),
        night_block=max(1, to_int(setting("Night block length", 3), 3)),
        rest_after_nights=max(0, to_int(setting("Rest days after nights", 2), 2)),
        weekend_days=weekend_days,
        seed=to_int(setting("Random seed", 42), 42),
    )
    if settings.start is None or settings.end is None:
        raise ValueError("Settings sheet needs both a Start date and an End date.")
    if settings.end < settings.start:
        raise ValueError("End date is before Start date.")

    # ---- staff --------------------------------------------------------
    staff: list[Staff] = []
    seen: set[str] = set()
    for record in read_table(sheet_for("Staff"), ["Name", "Band"]):
        name = str(record.get("name") or "").strip()
        if not name:
            continue
        if name.lower() in seen:
            raise ValueError(f"Staff sheet lists '{name}' more than once.")
        seen.add(name.lower())
        staff.append(Staff(
            name=name,
            band=str(record.get("band") or "").strip(),
            skills=split_list(record.get("skills")),
            nights_ok=is_yes(record.get("nightsyn", record.get("nights")), True),
            group=str(record.get("group") or "Main").strip() or "Main",
            max_shifts=to_int(record.get("maxshifts"), 0),
            notes=str(record.get("notes") or "").strip(),
        ))
    if not staff:
        raise ValueError("Staff sheet is empty.")

    # ---- shifts -------------------------------------------------------
    shifts: list[Shift] = []
    for record in read_table(sheet_for("Shifts"), ["Code", "Name", "Required"]):
        code = str(record.get("code") or "").strip()
        if not code:
            continue
        shifts.append(Shift(
            code=code,
            name=str(record.get("name") or code).strip(),
            start=str(record.get("start") or "").strip(),
            end=str(record.get("end") or "").strip(),
            hours=to_float(record.get("hours"), 0.0),
            days=str(record.get("days") or "All").strip() or "All",
            required=max(0, to_int(record.get("required"), 1)),
            requires_senior=is_yes(record.get("requiressenior"), True),
            is_night=is_yes(record.get("isnight"), False),
            colour=str(record.get("colour") or "D9E1F2").strip().lstrip("#").upper(),
            font_colour=str(record.get("fontcolour") or "000000").strip().lstrip("#").upper(),
        ))
    if not shifts:
        raise ValueError("Shifts sheet is empty.")

    # ---- leave types --------------------------------------------------
    leave_types: dict[str, LeaveType] = {}
    try:
        type_records = read_table(sheet_for("Leave types", "LeaveTypes"),
                                  ["Code", "Label"])
    except ValueError:
        type_records = []
    for record in type_records:
        code = str(record.get("code") or "").strip()
        if not code:
            continue
        leave_types[norm(code)] = LeaveType(
            code=code,
            label=str(record.get("label") or code).strip(),
            colour=str(record.get("colour") or "FFFF00").strip().lstrip("#").upper(),
            font_colour=str(record.get("fontcolour") or "000000").strip().lstrip("#").upper(),
        )
    for code, label, colour, font_colour in DEFAULT_LEAVE_TYPES:
        leave_types.setdefault(norm(code), LeaveType(code, label, colour, font_colour))

    # ---- leave --------------------------------------------------------
    known_names = {name.lower(): name for name in (person.name for person in staff)}
    leave: list[LeaveEntry] = []
    for record in read_table(sheet_for("Leave"), ["Name", "From", "To"]):
        name = str(record.get("name") or "").strip()
        if not name:
            continue
        if name.lower() not in known_names:
            raise ValueError(
                f"Leave sheet names '{name}', who is not on the Staff sheet."
            )
        start = to_date(record.get("from"), f"leave From for {name}")
        end = to_date(record.get("to"), f"leave To for {name}") or start
        if start is None:
            continue
        if end < start:
            raise ValueError(f"Leave for {name} ends before it starts.")
        code = str(record.get("type") or "A/L").strip() or "A/L"
        if norm(code) not in leave_types:
            leave_types[norm(code)] = LeaveType(code, code, "FFFF00", "000000")
        leave.append(LeaveEntry(known_names[name.lower()], start, end,
                                leave_types[norm(code)].code))

    # ---- benches ------------------------------------------------------
    benches: list[Bench] = []
    try:
        bench_records = read_table(sheet_for("Benches"), ["Bench", "Skill"])
    except ValueError:
        bench_records = []
    for record in bench_records:
        name = str(record.get("bench") or "").strip()
        if not name:
            continue
        weekend_min = record.get("minstaffweekend")
        benches.append(Bench(
            name=name,
            skill=str(record.get("skill") or "").strip(),
            days=str(record.get("days") or "All").strip() or "All",
            min_staff=max(0, to_int(record.get("minstaff"), 1)),
            min_weekend=(None if weekend_min in (None, "")
                         else max(0, to_int(weekend_min, 1))),
        ))

    return Config(settings, staff, shifts, leave, benches, leave_types)


# --------------------------------------------------------------------------
# the scheduler
# --------------------------------------------------------------------------

class Scheduler:
    def __init__(self, config: Config):
        self.config = config
        self.settings = config.settings
        self.staff = config.staff
        self.by_name = {person.name: person for person in config.staff}
        self.shifts = config.shifts
        self.shift_by_code = {shift.code: shift for shift in config.shifts}
        self.dates = list(daterange(self.settings.start, self.settings.end))
        self.rng = random.Random(self.settings.seed)

        self.leave: dict[tuple[date, str], str] = {}
        for entry in config.leave:
            for day in daterange(max(entry.start, self.settings.start),
                                 min(entry.end, self.settings.end)):
                self.leave[(day, entry.name)] = entry.code

        self.assign: dict[tuple[date, str], str] = {}
        self.rest_days: dict[str, set[date]] = defaultdict(set)
        self.benches: dict[tuple[date, str], list[str]] = {}
        self.violations: list[dict] = []

        self.night_shifts = [s for s in self.shifts if s.is_night]

    # -- small queries --------------------------------------------------

    def is_senior(self, person: Staff) -> bool:
        return person.band_value >= self.settings.senior_band

    def is_weekend(self, day: date) -> bool:
        return day.weekday() in self.settings.weekend_days

    def on_leave(self, name: str, day: date) -> bool:
        return (day, name) in self.leave

    def shift_on(self, name: str, day: date):
        code = self.assign.get((day, name))
        return self.shift_by_code.get(code) if code else None

    def total_shifts(self, name: str) -> int:
        return sum(1 for (_, person) in self.assign if person == name)

    def count_code(self, name: str, code: str) -> int:
        return sum(1 for (_, person), value in self.assign.items()
                   if person == name and value == code)

    def count_weekends(self, name: str) -> int:
        return sum(1 for (day, person) in self.assign
                   if person == name and self.is_weekend(day))

    def count_nights(self, name: str) -> int:
        return sum(1 for (_, person), code in self.assign.items()
                   if person == name and self.shift_by_code[code].is_night)

    def run_length(self, name: str, day: date) -> int:
        """Length of the unbroken worked run that would contain `day`."""
        length = 1
        cursor = day - timedelta(days=1)
        while (cursor, name) in self.assign:
            length += 1
            cursor -= timedelta(days=1)
        cursor = day + timedelta(days=1)
        while (cursor, name) in self.assign:
            length += 1
            cursor += timedelta(days=1)
        return length

    # -- hard constraints -----------------------------------------------

    def can_assign(self, person: Staff, day: date, shift: Shift) -> bool:
        name = person.name
        if (day, name) in self.assign:
            return False                              # one shift per day
        if self.on_leave(name, day):
            return False                              # leave is absolute
        if day in self.rest_days[name]:
            return False                              # post-nights rest
        if shift.is_night and not person.nights_ok:
            return False
        if person.max_shifts and self.total_shifts(name) >= person.max_shifts:
            return False

        previous = self.shift_on(name, day - timedelta(days=1))
        if previous is not None and previous.is_night and not shift.is_night:
            return False        # no shift the day after a night (bans night->early)
        following = self.shift_on(name, day + timedelta(days=1))
        if shift.is_night and following is not None and not following.is_night:
            return False        # ... and do not create that clash going forwards

        if self.settings.max_consecutive > 0:
            if self.run_length(name, day) > self.settings.max_consecutive:
                return False
        return True

    def candidates(self, day: date, shift: Shift, seniors_only=False,
                   needed_skills: set[str] | None = None) -> list[Staff]:
        pool = [person for person in self.staff
                if (not seniors_only or self.is_senior(person))
                and self.can_assign(person, day, shift)]
        weekend = self.is_weekend(day)
        needed_skills = needed_skills or set()

        def sort_key(person: Staff):
            helps_bench = 0 if any(person.has_skill(s) for s in needed_skills) else 1
            fairness = (self.count_weekends(person.name), self.total_shifts(person.name)) \
                if weekend else (self.total_shifts(person.name),)
            return (
                *fairness,
                helps_bench,
                self.count_code(person.name, shift.code),
                self.run_length(person.name, day),
                self.rng.random(),
            )

        return sorted(pool, key=sort_key)

    # -- building the schedule -------------------------------------------

    def build(self) -> None:
        self.place_night_blocks()
        self.place_day_shifts()
        self.repair_bench_coverage()
        self.allocate_benches()
        self.validate()

    def place_night_blocks(self) -> None:
        """Nights first: they are the most constrained and drive rest days."""
        for shift in self.night_shifts:
            for day in self.dates:
                if not shift.applies_on(day, self.settings.weekend_days):
                    continue
                while self.covered(day, shift) < shift.required:
                    if not self.start_night_block(day, shift):
                        self.violations.append({
                            "rule": "Shift fully staffed",
                            "date": day,
                            "staff": "",
                            "detail": (f"{shift.name}: only {self.covered(day, shift)}"
                                       f" of {shift.required} covered"),
                        })
                        break

    def start_night_block(self, day: date, shift: Shift) -> bool:
        block = [d for d in (day + timedelta(days=offset)
                             for offset in range(self.settings.night_block))
                 if d <= self.settings.end
                 and shift.applies_on(d, self.settings.weekend_days)]
        if not block:
            return False

        need_senior = shift.requires_senior and not self.has_senior(day, shift)
        pools = [self.candidates(day, shift, seniors_only=True), self.candidates(day, shift)] \
            if need_senior else [self.candidates(day, shift)]

        for pool in pools:
            for person in pool:
                placed = []
                for target in block:
                    if self.covered(target, shift) >= shift.required:
                        break
                    if not self.can_assign(person, target, shift):
                        break
                    self.assign[(target, person.name)] = shift.code
                    placed.append(target)
                if not placed:
                    continue
                last = placed[-1]
                for offset in range(1, self.settings.rest_after_nights + 1):
                    self.rest_days[person.name].add(last + timedelta(days=offset))
                return True
        return False

    def place_day_shifts(self) -> None:
        day_shifts = [shift for shift in self.shifts if not shift.is_night]
        for day in self.dates:
            todays = [shift for shift in day_shifts
                      if shift.applies_on(day, self.settings.weekend_days)]
            # Hardest first: senior-critical shifts, then the smallest pool.
            todays.sort(key=lambda s: (not s.requires_senior,
                                       len(self.candidates(day, s))))
            for shift in todays:
                needed = shift.required - self.covered(day, shift)
                if needed <= 0:
                    continue
                if shift.requires_senior and not self.has_senior(day, shift):
                    seniors = self.candidates(day, shift, seniors_only=True,
                                              needed_skills=self.uncovered_skills(day))
                    if seniors:
                        self.assign[(day, seniors[0].name)] = shift.code
                        needed -= 1
                    else:
                        self.violations.append({
                            "rule": "Senior on every shift",
                            "date": day,
                            "staff": "",
                            "detail": f"{shift.name}: no Band "
                                      f"{self.settings.senior_band:g}+ available",
                        })
                for _ in range(needed):
                    pool = self.candidates(day, shift,
                                           needed_skills=self.uncovered_skills(day))
                    if not pool:
                        self.violations.append({
                            "rule": "Shift fully staffed",
                            "date": day,
                            "staff": "",
                            "detail": (f"{shift.name}: only {self.covered(day, shift)}"
                                       f" of {shift.required} covered"),
                        })
                        break
                    self.assign[(day, pool[0].name)] = shift.code

    def covered(self, day: date, shift: Shift) -> int:
        return sum(1 for (d, _), code in self.assign.items()
                   if d == day and code == shift.code)

    def has_senior(self, day: date, shift: Shift) -> bool:
        return any(self.is_senior(self.by_name[name])
                   for (d, name), code in self.assign.items()
                   if d == day and code == shift.code)

    def on_duty(self, day: date, include_nights=False) -> list[str]:
        return [name for (d, name), code in self.assign.items()
                if d == day and (include_nights or not self.shift_by_code[code].is_night)]

    def uncovered_skills(self, day: date) -> set[str]:
        """Bench skills not yet met by the day staff already assigned."""
        needed = set()
        for bench in self.config.benches:
            if not bench.applies_on(day, self.settings.weekend_days):
                continue
            have = sum(1 for name in self.on_duty(day)
                       if self.by_name[name].has_skill(bench.skill))
            if have < bench.required_on(day, self.settings.weekend_days):
                needed.add(bench.skill)
        return needed

    def bench_deficit(self, day: date) -> int:
        """How many bench slots that day have nobody skilled on duty for them."""
        deficit = 0
        for bench in self.config.benches:
            wanted = bench.required_on(day, self.settings.weekend_days)
            if not wanted or not bench.applies_on(day, self.settings.weekend_days):
                continue
            have = sum(1 for name in self.on_duty(day)
                       if self.by_name[name].has_skill(bench.skill))
            deficit += max(0, wanted - have)
        return deficit

    def repair_bench_coverage(self) -> None:
        """Swap in skilled staff where a bench has nobody, without widening the
        shift spread: a swap only happens if the incoming person currently has
        strictly fewer shifts than the person they replace."""
        day_shifts = [shift for shift in self.shifts if not shift.is_night]
        for day in self.dates:
            for _ in range(4):
                if self.bench_deficit(day) == 0:
                    break
                if not self.try_bench_swap(day, day_shifts):
                    break

    def try_bench_swap(self, day: date, day_shifts: list[Shift]) -> bool:
        before = self.bench_deficit(day)
        needed = self.uncovered_skills(day)
        if not needed:
            return False

        for shift in day_shifts:
            if not shift.applies_on(day, self.settings.weekend_days):
                continue
            assigned = sorted(
                (name for (d, name), code in self.assign.items()
                 if d == day and code == shift.code),
                key=lambda n: -self.total_shifts(n),
            )
            for out_name in assigned:
                out_total = self.total_shifts(out_name)
                del self.assign[(day, out_name)]

                pool = [person for person in self.staff
                        if person.name != out_name
                        and any(person.has_skill(skill) for skill in needed)
                        and self.total_shifts(person.name) < out_total]
                pool.sort(key=lambda p: self.total_shifts(p.name))

                for person in pool:
                    if not self.can_assign(person, day, shift):
                        continue
                    self.assign[(day, person.name)] = shift.code
                    improved = self.bench_deficit(day) < before
                    if improved and shift.requires_senior \
                            and not self.has_senior(day, shift):
                        improved = False
                    if improved:
                        return True
                    del self.assign[(day, person.name)]

                self.assign[(day, out_name)] = shift.code
        return False

    def allocate_benches(self) -> None:
        load: Counter[str] = Counter()
        for day in self.dates:
            duty = self.on_duty(day)
            used_today: set[str] = set()
            for bench in self.config.benches:
                wanted = bench.required_on(day, self.settings.weekend_days)
                if not bench.applies_on(day, self.settings.weekend_days) or not wanted:
                    self.benches[(day, bench.name)] = []
                    continue
                pool = [name for name in duty
                        if self.by_name[name].has_skill(bench.skill)]
                pool.sort(key=lambda n: (n in used_today, load[n], n))
                picked = pool[:wanted]
                for name in picked:
                    load[name] += 1
                    used_today.add(name)
                self.benches[(day, bench.name)] = picked
                if len(picked) < wanted:
                    self.violations.append({
                        "rule": "Bench covered",
                        "date": day,
                        "staff": "",
                        "detail": (f"{bench.name}: {len(picked)} of "
                                   f"{wanted} skilled staff on duty"),
                    })

    # -- verification ----------------------------------------------------

    def validate(self) -> None:
        """Re-check every hard rule against the finished schedule."""
        for (day, name), code in sorted(self.assign.items()):
            shift = self.shift_by_code[code]
            if self.on_leave(name, day):
                self.violations.append({
                    "rule": "Leave respected", "date": day, "staff": name,
                    "detail": f"scheduled {shift.name} while on "
                              f"{self.leave[(day, name)]}",
                })
            previous = self.shift_on(name, day - timedelta(days=1))
            if previous is not None and previous.is_night and not shift.is_night:
                self.violations.append({
                    "rule": "No shift after a night", "date": day, "staff": name,
                    "detail": f"{previous.name} then {shift.name} next day",
                })

        for person in self.staff:
            run = 0
            for day in self.dates:
                run = run + 1 if (day, person.name) in self.assign else 0
                if self.settings.max_consecutive and run > self.settings.max_consecutive:
                    self.violations.append({
                        "rule": "Max consecutive days", "date": day,
                        "staff": person.name,
                        "detail": f"{run} days in a row (limit "
                                  f"{self.settings.max_consecutive})",
                    })
                    run = 0     # report once per over-long run

        for day in self.dates:
            for shift in self.shifts:
                if not shift.applies_on(day, self.settings.weekend_days):
                    continue
                if shift.requires_senior and self.covered(day, shift) \
                        and not self.has_senior(day, shift):
                    self.violations.append({
                        "rule": "Senior on every shift", "date": day, "staff": "",
                        "detail": f"{shift.name} has no senior on duty",
                    })

    # -- reporting -------------------------------------------------------

    def summary_rows(self) -> list[dict]:
        rows = []
        for person in self.staff:
            counts = {shift.code: self.count_code(person.name, shift.code)
                      for shift in self.shifts}
            total = sum(counts.values())
            hours = sum(self.shift_by_code[code].hours * count
                        for code, count in counts.items())
            leave_days = sum(1 for (_, name) in self.leave if name == person.name)
            rows.append({
                "name": person.name,
                "band": person.band,
                "senior": "Yes" if self.is_senior(person) else "",
                "counts": counts,
                "total": total,
                "weekends": self.count_weekends(person.name),
                "nights": self.count_nights(person.name),
                "leave": leave_days,
                "hours": hours,
                "off": len(self.dates) - total - leave_days,
            })
        return rows


# --------------------------------------------------------------------------
# writing the rota workbook
# --------------------------------------------------------------------------

def write_rota(scheduler: Scheduler, path: Path) -> None:
    config = scheduler.config
    settings = scheduler.settings
    dates = scheduler.dates
    initials = make_initials([person.name for person in config.staff])

    shift_fill = {shift.code: PatternFill("solid", fgColor=shift.colour)
                  for shift in config.shifts}
    shift_font = {shift.code: Font(bold=True, color=shift.font_colour, size=10)
                  for shift in config.shifts}
    leave_fill = {entry.code: PatternFill("solid", fgColor=entry.colour)
                  for entry in config.leave_types.values()}
    leave_font = {entry.code: Font(bold=True, color=entry.font_colour, size=9)
                  for entry in config.leave_types.values()}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rota"
    sheet.sheet_view.showGridLines = False
    last_col = FIRST_DATE_COL + len(dates) - 1

    # ---- title --------------------------------------------------------
    sheet.merge_cells(start_row=TITLE_ROW, start_column=1,
                      end_row=TITLE_ROW, end_column=last_col)
    title = sheet.cell(row=TITLE_ROW, column=1)
    title.value = (f"{settings.title}   |   "
                   f"{settings.start.strftime('%d %b %Y')} to "
                   f"{settings.end.strftime('%d %b %Y')}")
    title.fill = HEADER_FILL
    title.font = Font(bold=True, size=14, color="FFFFFF")
    title.alignment = LEFT
    sheet.row_dimensions[TITLE_ROW].height = 24

    # ---- month bands --------------------------------------------------
    for label, first_col, span in month_bands(dates):
        sheet.merge_cells(start_row=MONTH_ROW, start_column=first_col,
                          end_row=MONTH_ROW, end_column=first_col + span - 1)
        cell = sheet.cell(row=MONTH_ROW, column=first_col, value=label)
        cell.fill = MONTH_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTRE
        cell.border = GRID_BORDER

    # ---- date headers -------------------------------------------------
    # The three left-hand columns are labelled on the day-number row and left
    # blank (but still filled) on the weekday row below it.
    for column, label in ((1, "Staff"), (2, "Band"), (3, "Init")):
        cell = sheet.cell(row=DAYNUM_ROW, column=column, value=label)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = LEFT if column == 1 else CENTRE
        cell.border = GRID_BORDER

        spacer = sheet.cell(row=WEEKDAY_ROW, column=column)
        spacer.fill = HEADER_FILL
        spacer.border = GRID_BORDER

    for offset, day in enumerate(dates):
        column = FIRST_DATE_COL + offset
        weekend = scheduler.is_weekend(day)

        number = sheet.cell(row=DAYNUM_ROW, column=column, value=day.day)
        number.fill = WEEKEND_FILL if weekend else DAY_FILL
        number.font = Font(bold=True, size=10)
        number.alignment = CENTRE
        number.border = GRID_BORDER

        weekday = sheet.cell(row=WEEKDAY_ROW, column=column,
                             value=WEEKDAY_NAMES[day.weekday()][:3])
        weekday.fill = WEEKEND_FILL if weekend else DAY_FILL
        weekday.font = Font(size=8, bold=weekend)
        weekday.alignment = CENTRE
        weekday.border = GRID_BORDER

        sheet.column_dimensions[get_column_letter(column)].width = 4.6

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 6
    sheet.column_dimensions["C"].width = 6

    # ---- staff rows, grouped ------------------------------------------
    groups: dict[str, list[Staff]] = defaultdict(list)
    for person in config.staff:
        groups[person.group].append(person)
    ordered_groups = ([("Main", groups.pop("Main"))] if "Main" in groups else []) \
        + sorted(groups.items())

    row = FIRST_STAFF_ROW
    show_labels = len(ordered_groups) > 1
    for group_name, members in ordered_groups:
        if show_labels:
            sheet.merge_cells(start_row=row, start_column=1,
                              end_row=row, end_column=last_col)
            cell = sheet.cell(row=row, column=1, value=group_name.upper())
            cell.fill = GROUP_FILL
            cell.font = Font(bold=True, size=10)
            cell.alignment = LEFT
            row += 1

        for person in members:
            name_cell = sheet.cell(row=row, column=1, value=person.name)
            name_cell.alignment = LEFT
            name_cell.font = Font(bold=scheduler.is_senior(person), size=10)
            name_cell.border = GRID_BORDER

            band_cell = sheet.cell(row=row, column=2, value=person.band)
            band_cell.alignment = CENTRE
            band_cell.font = Font(size=9)
            band_cell.border = GRID_BORDER

            init_cell = sheet.cell(row=row, column=3, value=initials[person.name])
            init_cell.alignment = CENTRE
            init_cell.font = Font(size=9, italic=True)
            init_cell.border = GRID_BORDER

            for offset, day in enumerate(dates):
                cell = sheet.cell(row=row, column=FIRST_DATE_COL + offset)
                cell.alignment = CENTRE
                cell.border = GRID_BORDER
                code = scheduler.assign.get((day, person.name))
                leave_code = scheduler.leave.get((day, person.name))
                if code:
                    cell.value = code
                    cell.fill = shift_fill[code]
                    cell.font = shift_font[code]
                elif leave_code:
                    cell.value = leave_code
                    cell.fill = leave_fill.get(
                        leave_code, PatternFill("solid", fgColor="FFFF00"))
                    cell.font = leave_font.get(leave_code, Font(bold=True, size=9))
                elif scheduler.is_weekend(day):
                    cell.fill = WEEKEND_FILL
            row += 1

    # ---- staff-on-duty count row --------------------------------------
    row += 1
    label = sheet.cell(row=row, column=1, value="Staff on duty (day shifts)")
    label.font = BOLD
    label.alignment = LEFT
    label.fill = TOTAL_FILL
    for column in (2, 3):
        sheet.cell(row=row, column=column).fill = TOTAL_FILL
    for offset, day in enumerate(dates):
        cell = sheet.cell(row=row, column=FIRST_DATE_COL + offset,
                          value=len(scheduler.on_duty(day)))
        cell.alignment = CENTRE
        cell.font = Font(bold=True, size=9)
        cell.fill = TOTAL_FILL
        cell.border = GRID_BORDER

    # ---- bench allocation ---------------------------------------------
    if config.benches:
        row += 2
        sheet.merge_cells(start_row=row, start_column=1,
                          end_row=row, end_column=last_col)
        cell = sheet.cell(row=row, column=1, value="BENCH / SECTION ALLOCATION")
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = LEFT
        row += 1

        for bench in config.benches:
            name_cell = sheet.cell(row=row, column=1, value=bench.name)
            name_cell.font = Font(bold=True, size=10)
            name_cell.alignment = LEFT
            name_cell.fill = BENCH_FILL
            name_cell.border = GRID_BORDER
            skill_cell = sheet.cell(row=row, column=2, value=bench.skill)
            skill_cell.alignment = CENTRE
            skill_cell.font = Font(size=8)
            skill_cell.fill = BENCH_FILL
            skill_cell.border = GRID_BORDER
            sheet.cell(row=row, column=3).fill = BENCH_FILL

            for offset, day in enumerate(dates):
                people = scheduler.benches.get((day, bench.name), [])
                cell = sheet.cell(row=row, column=FIRST_DATE_COL + offset)
                cell.value = "/".join(initials[name] for name in people) or None
                cell.alignment = WRAP_CENTRE
                cell.font = Font(size=7)
                cell.border = GRID_BORDER
                if len(people) < bench.required_on(day, settings.weekend_days):
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")
            row += 1

    # ---- key ----------------------------------------------------------
    row += 2
    key_header = sheet.cell(row=row, column=1, value="KEY")
    key_header.font = Font(bold=True, size=12, color="1F3864")
    row += 1
    for shift in config.shifts:
        code_cell = sheet.cell(row=row, column=3, value=shift.code)
        code_cell.fill = shift_fill[shift.code]
        code_cell.font = shift_font[shift.code]
        code_cell.alignment = CENTRE
        code_cell.border = GRID_BORDER
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=12)
        detail = sheet.cell(row=row, column=4)
        detail.value = (f"{shift.label} - {shift.required} per day, {shift.days}"
                        f"{', senior required' if shift.requires_senior else ''}")
        detail.alignment = LEFT
        detail.font = Font(size=10)
        row += 1

    used_leave = {entry.code for entry in config.leave}
    for entry in config.leave_types.values():
        if used_leave and entry.code not in used_leave:
            continue
        code_cell = sheet.cell(row=row, column=3, value=entry.code)
        code_cell.fill = leave_fill[entry.code]
        code_cell.font = leave_font[entry.code]
        code_cell.alignment = CENTRE
        code_cell.border = GRID_BORDER
        sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=12)
        detail = sheet.cell(row=row, column=4, value=entry.label)
        detail.alignment = LEFT
        detail.font = Font(size=10)
        row += 1

    blank = sheet.cell(row=row, column=3)
    blank.fill = WEEKEND_FILL
    blank.border = GRID_BORDER
    sheet.merge_cells(start_row=row, start_column=4, end_row=row, end_column=12)
    sheet.cell(row=row, column=4, value="Weekend / rest day (no shift)").font = Font(size=10)
    row += 2
    sheet.cell(row=row, column=1,
               value=f"Senior = Band {settings.senior_band:g}+ (shown in bold). "
                     f"Nights run in blocks of {settings.night_block} followed by "
                     f"{settings.rest_after_nights} rest day(s); "
                     f"max {settings.max_consecutive} consecutive days.").font = \
        Font(size=9, italic=True)

    sheet.freeze_panes = sheet.cell(row=FIRST_STAFF_ROW, column=FIRST_DATE_COL).coordinate
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    write_summary(scheduler, workbook, initials)
    workbook.save(path)


def month_bands(dates: list[date]) -> list[tuple[str, int, int]]:
    bands = []
    for offset, day in enumerate(dates):
        label = day.strftime("%B %Y").upper()
        column = FIRST_DATE_COL + offset
        if bands and bands[-1][0] == label:
            bands[-1] = (label, bands[-1][1], bands[-1][2] + 1)
        else:
            bands.append((label, column, 1))
    return bands


def write_summary(scheduler: Scheduler, workbook: Workbook, initials: dict[str, str]) -> None:
    config = scheduler.config
    settings = scheduler.settings
    sheet = workbook.create_sheet("Summary")
    sheet.sheet_view.showGridLines = False

    headers = ["Staff", "Band", "Senior", "Init"] \
        + [f"{shift.code} - {shift.name}" for shift in config.shifts] \
        + ["Total shifts", "Weekends", "Nights", "Leave days", "Days off", "Hours"]

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = sheet.cell(row=1, column=1)
    title.value = (f"Shifts per person   |   {settings.start.strftime('%d %b %Y')} to "
                   f"{settings.end.strftime('%d %b %Y')}   |   {len(scheduler.dates)} days")
    title.fill = HEADER_FILL
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.alignment = LEFT
    sheet.row_dimensions[1].height = 22

    for column, header_text in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=column, value=header_text)
        cell.fill = MONTH_FILL
        cell.font = WHITE_BOLD
        cell.alignment = WRAP_CENTRE
        cell.border = GRID_BORDER
    sheet.row_dimensions[2].height = 30

    rows = scheduler.summary_rows()
    schedulable = [row for row in rows if row["total"] or row["leave"] < len(scheduler.dates)]
    totals = [row["total"] for row in schedulable] or [0]
    lowest, highest = min(totals), max(totals)

    row_index = 3
    for record in rows:
        values = [record["name"], record["band"], record["senior"],
                  initials[record["name"]]]
        values += [record["counts"][shift.code] for shift in config.shifts]
        values += [record["total"], record["weekends"], record["nights"],
                   record["leave"], record["off"], round(record["hours"], 2)]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.border = GRID_BORDER
            cell.alignment = LEFT if column == 1 else CENTRE
            cell.font = Font(size=10, bold=(column == 1 and record["senior"] == "Yes"))
        total_cell = sheet.cell(row=row_index, column=4 + len(config.shifts) + 1)
        total_cell.font = Font(bold=True, size=10)
        if record["total"] == highest and highest != lowest:
            total_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif record["total"] == lowest and highest != lowest and record["total"]:
            total_cell.fill = PatternFill("solid", fgColor="FFEB9C")
        row_index += 1

    # totals row
    for column, _ in enumerate(headers, start=1):
        cell = sheet.cell(row=row_index, column=column)
        cell.fill = TOTAL_FILL
        cell.border = GRID_BORDER
        cell.font = BOLD
        cell.alignment = CENTRE
    sheet.cell(row=row_index, column=1, value="TOTAL").alignment = LEFT
    for offset, shift in enumerate(config.shifts):
        sheet.cell(row=row_index, column=5 + offset,
                   value=sum(r["counts"][shift.code] for r in rows))
    base = 4 + len(config.shifts)
    for offset, key in enumerate(["total", "weekends", "nights", "leave", "off"], start=1):
        sheet.cell(row=row_index, column=base + offset, value=sum(r[key] for r in rows))
    sheet.cell(row=row_index, column=base + 6,
               value=round(sum(r["hours"] for r in rows), 2))

    sheet.column_dimensions["A"].width = 24
    for column in range(2, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 11
    sheet.freeze_panes = "B3"

    # ---- fairness ------------------------------------------------------
    row_index += 2
    section(sheet, row_index, "Distribution", len(headers))
    row_index += 1
    mean = sum(totals) / len(totals)
    for label, value in [
        ("Staff available to schedule", len(schedulable)),
        ("Shifts scheduled", sum(r["total"] for r in rows)),
        ("Lowest total", lowest),
        ("Highest total", highest),
        ("Mean total", round(mean, 2)),
        ("Spread (highest - lowest)", highest - lowest),
    ]:
        sheet.cell(row=row_index, column=1, value=label).font = Font(size=10)
        cell = sheet.cell(row=row_index, column=2, value=value)
        cell.font = BOLD
        cell.alignment = LEFT
        row_index += 1

    # ---- rule checks ---------------------------------------------------
    row_index += 1
    section(sheet, row_index, "Rule checks", len(headers))
    row_index += 1
    by_rule: dict[str, list[dict]] = defaultdict(list)
    for violation in scheduler.violations:
        by_rule[violation["rule"]].append(violation)

    rules = [
        "Leave respected",
        "No shift after a night",
        "Max consecutive days",
        "Senior on every shift",
        "Shift fully staffed",
        "Bench covered",
    ]
    for rule in rules:
        found = by_rule.get(rule, [])
        sheet.cell(row=row_index, column=1, value=rule).font = Font(size=10)
        cell = sheet.cell(row=row_index, column=2,
                          value="PASS" if not found else f"{len(found)} issue(s)")
        cell.font = Font(bold=True, color="006100" if not found else "9C0006")
        cell.fill = PatternFill("solid",
                                fgColor="C6EFCE" if not found else "FFC7CE")
        cell.alignment = LEFT
        row_index += 1

    if scheduler.violations:
        row_index += 1
        section(sheet, row_index, "Issue detail", len(headers))
        row_index += 1
        for column, header_text in enumerate(["Rule", "Date", "Staff", "Detail"], start=1):
            cell = sheet.cell(row=row_index, column=column, value=header_text)
            cell.font = BOLD
            cell.border = GRID_BORDER
        row_index += 1
        for violation in scheduler.violations:
            when = violation["date"]
            values = [violation["rule"],
                      when.strftime("%a %d/%m/%Y") if isinstance(when, date) else "",
                      violation["staff"], violation["detail"]]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row=row_index, column=column, value=value)
                cell.font = Font(size=9)
                cell.alignment = LEFT
            row_index += 1

    sheet.column_dimensions["D"].width = 46


def section(sheet, row: int, label: str, width: int) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = sheet.cell(row=row, column=1, value=label.upper())
    cell.fill = MONTH_FILL
    cell.font = WHITE_BOLD
    cell.alignment = LEFT


# --------------------------------------------------------------------------
# command line
# --------------------------------------------------------------------------

def locked_message(path: Path) -> str:
    return (f"Could not write {path}: the file is open in another program, "
            f"most likely Excel.\nClose it and run this again.")


def cmd_template(args) -> int:
    path = Path(args.out)
    if path.exists() and not args.force:
        print(f"{path} already exists. Re-run with --force to overwrite it.")
        return 1
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        build_template(path)
    except PermissionError:
        print(locked_message(path))
        return 1
    print(f"Input template written to {path}")
    print("Fill in Staff / Shifts / Leave / Settings, then run:")
    print(f"  python {Path(__file__).name} generate --input {path} --out rota.xlsx")
    return 0


def cmd_generate(args) -> int:
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Input workbook not found: {input_path}")
        return 1
    try:
        config = load_config(input_path)
    except ValueError as error:
        print(f"Could not read {input_path}: {error}")
        return 1

    try:
        if args.start:
            config.settings.start = to_date(args.start, "--start")
        if args.end:
            config.settings.end = to_date(args.end, "--end")
    except ValueError as error:
        print(error)
        return 1
    if args.seed is not None:
        config.settings.seed = args.seed
    if config.settings.end < config.settings.start:
        print("End date is before start date.")
        return 1

    scheduler = Scheduler(config)
    scheduler.build()

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        write_rota(scheduler, out_path)
    except PermissionError:
        print(locked_message(out_path))
        return 1

    rows = scheduler.summary_rows()
    totals = [row["total"] for row in rows if row["total"]]
    print(f"Rota written to {out_path}")
    print(f"  {len(scheduler.dates)} days, {len(config.staff)} staff, "
          f"{sum(r['total'] for r in rows)} shifts assigned")
    if totals:
        print(f"  shifts per working person: min {min(totals)}, "
              f"max {max(totals)}, spread {max(totals) - min(totals)}")
    if scheduler.violations:
        counts = Counter(v["rule"] for v in scheduler.violations)
        print("  issues (see the Summary sheet):")
        for rule, count in counts.items():
            print(f"    - {rule}: {count}")
    else:
        print("  all rules satisfied")
    return 0


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Generate a lab shift rota as a formatted Excel file.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    template = subparsers.add_parser(
        "template", help="write a blank/example input workbook to fill in")
    template.add_argument("--out", default="roster_input.xlsx")
    template.add_argument("--force", action="store_true",
                          help="overwrite an existing file")
    template.set_defaults(func=cmd_template)

    generate = subparsers.add_parser(
        "generate", help="read an input workbook and write the rota")
    generate.add_argument("--input", default="roster_input.xlsx")
    generate.add_argument("--out", default="rota.xlsx")
    generate.add_argument("--start", help="override the start date (dd/mm/yyyy)")
    generate.add_argument("--end", help="override the end date (dd/mm/yyyy)")
    generate.add_argument("--seed", type=int, help="override the random seed")
    generate.set_defaults(func=cmd_generate)

    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
