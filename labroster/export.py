"""The exported workbook.

The spreadsheet stays one of the strongest things this tool produces, because it
is what gets printed, emailed and pinned up.  Formatting is functional rather
than decorative: freeze panes, filters, sensible widths, a legend, and colour
used alongside words so a status is never carried by colour alone.
"""

from __future__ import annotations

from collections import defaultdict
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from .analysis import CRITICAL, PASSED, REVIEW, Analysis
from .models import WEEKDAY_SHORT
from .scheduler import Scheduler

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTRE = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
BAND_FILL = PatternFill("solid", fgColor="2E5496")
DAY_FILL = PatternFill("solid", fgColor="D6DCE4")
WEEKEND_FILL = PatternFill("solid", fgColor="F2E6D9")
GROUP_FILL = PatternFill("solid", fgColor="D9D9D9")
BENCH_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")

CRITICAL_FILL = PatternFill("solid", fgColor="FBE6E8")
REVIEW_FILL = PatternFill("solid", fgColor="FDF1DC")
PASSED_FILL = PatternFill("solid", fgColor="E3F4E9")
SEVERITY_FILL = {CRITICAL: CRITICAL_FILL, REVIEW: REVIEW_FILL,
                 PASSED: PASSED_FILL}
SEVERITY_FONT = {CRITICAL: Font(bold=True, size=10, color="9C1F2B"),
                 REVIEW: Font(bold=True, size=10, color="8A5300"),
                 PASSED: Font(bold=True, size=10, color="1D6B3F")}

WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True, size=10)
SMALL = Font(size=9)
MUTED = Font(size=9, color="5B6470")


def _title(sheet, text, width, subtitle=""):
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = sheet.cell(row=1, column=1, value=text)
    cell.fill = HEADER_FILL
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.alignment = LEFT
    sheet.row_dimensions[1].height = 24
    if subtitle:
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
        note = sheet.cell(row=2, column=1, value=subtitle)
        note.font = MUTED
        note.alignment = LEFT
        return 3
    return 2


def _headers(sheet, row, headings, widths=None, height=28):
    for column, heading in enumerate(headings, start=1):
        cell = sheet.cell(row=row, column=column, value=heading)
        cell.fill = BAND_FILL
        cell.font = WHITE_BOLD
        cell.alignment = WRAP
        cell.border = BORDER
    sheet.row_dimensions[row].height = height
    for column, width in enumerate(widths or [], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1).coordinate
    sheet.auto_filter.ref = f"A{row}:{get_column_letter(len(headings))}{row}"
    return row + 1


def _row(sheet, row, values, fills=None, fonts=None):
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=row, column=column, value=value)
        cell.border = BORDER
        cell.alignment = LEFT if column == 1 else CENTRE
        cell.font = SMALL
        if isinstance(value, date):
            cell.number_format = "DD/MM/YYYY"
    if fills:
        for column, fill in fills.items():
            sheet.cell(row=row, column=column).fill = fill
    if fonts:
        for column, font in fonts.items():
            sheet.cell(row=row, column=column).font = font
    return row + 1


def _landscape(sheet, fit_width=1):
    sheet.page_setup.orientation = "landscape"
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.page_setup.fitToWidth = fit_width
    sheet.page_setup.fitToHeight = 0


# --------------------------------------------------------------------------

MANAGER = "manager"
STAFF = "staff"


def write_workbook(scheduler: Scheduler, analysis: Analysis, target,
                   audience: str = MANAGER) -> None:
    """Write the exported workbook for a given audience.

    The **manager report** keeps the full analytical output.  The **staff rota** is
    deliberately reduced to what somebody needs in order to read their own rota:
    who is working, when, and in which section.

    Data minimisation is the point of the split, not tidiness.  A rota circulated
    around a department must not disclose colleagues' competency records, private
    working restrictions, individual hours, fairness calculations, workforce
    vulnerabilities, or whether an absence is annual leave or sickness.
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    if audience == STAFF:
        _roster(workbook, scheduler, analysis, audience=STAFF)
        if scheduler.config.benches:
            _staff_sections(workbook, scheduler, analysis)
        _staff_notes(workbook, scheduler)
        workbook.save(target)
        return

    _instructions(workbook, scheduler, analysis)
    _roster(workbook, scheduler, analysis)
    _staff(workbook, scheduler, analysis)
    _competencies(workbook, scheduler, analysis)
    _requirements(workbook, scheduler)
    _bench_allocations(workbook, scheduler, analysis)
    _issues(workbook, analysis)
    _hours(workbook, analysis)
    _fairness(workbook, scheduler, analysis)
    _expiry(workbook, analysis)
    workbook.save(target)


def _instructions(workbook, scheduler, analysis) -> None:
    sheet = workbook.create_sheet("Instructions")
    config = scheduler.config
    details = config.details
    row = _title(sheet, "LabRoster — draft roster", 8,
                 "Competency-aware workforce planning for diagnostic laboratories")
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 64

    row += 1
    sheet.cell(row=row, column=1, value="This roster is a draft").font = Font(
        bold=True, size=12, color="1F3864")
    row += 1
    sheet.cell(row=row, column=1,
               value="Every roster LabRoster produces requires managerial review "
                     "before use. It supports your decisions about staffing; it "
                     "does not make them, and it does not replace HR, payroll or "
                     "professional judgement.")
    sheet.cell(row=row, column=1).alignment = WRAP_LEFT
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    sheet.row_dimensions[row].height = 42
    row += 2

    for label, value in [
        ("Rota name", details.rota_name),
        ("Organisation", details.organisation or "—"),
        ("Department", details.department or "—"),
        ("Site", details.site or "—"),
        ("Prepared by", details.prepared_by or "—"),
        ("Roster period", f"{config.period.start:%d %b %Y} to "
                          f"{config.period.end:%d %b %Y} "
                          f"({config.period.day_count} days)"),
        ("Staff included", len(config.staff)),
        ("Roster status", analysis.metrics["roster_status"]),
        ("Staffing slot coverage",
         f"{analysis.metrics['staffing_slot_coverage_percent']}% "
         f"({analysis.metrics['filled_slots']} of "
         f"{analysis.metrics['required_slots']} positions occupied)"),
        ("Shifts meeting all configured requirements",
         f"{analysis.metrics['shifts_meeting_all_requirements_percent']}% "
         f"({analysis.metrics['shift_instances_met']} of "
         f"{analysis.metrics['shift_instances']} shifts)"),
    ]:
        sheet.cell(row=row, column=1, value=label).font = BOLD
        sheet.cell(row=row, column=2, value=value).font = SMALL
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="What each sheet contains").font = Font(
        bold=True, size=12, color="1F3864")
    row += 1
    for name, description in [
        ("Roster", "The rota itself: staff down the side, dates across the top, "
                   "with bench allocations and staffing counts underneath."),
        ("Staff", "Who was included, their grade, role and working pattern."),
        ("Competencies", "Who is competent in what, and the status of each record."),
        ("Shift Requirements", "The minimum staffing and competencies each shift "
                               "was required to meet."),
        ("Bench Allocations", "Who was allocated to which section, day by day."),
        ("Issues", "Everything needing attention, grouped Critical, Review, Passed."),
        ("Hours Summary", "Contracted target against allocated hours per person."),
        ("Fairness Summary", "How nights, weekends and unpopular shifts are shared."),
        ("Competency Expiry", "Records that have expired or expire soon."),
    ]:
        sheet.cell(row=row, column=1, value=name).font = BOLD
        cell = sheet.cell(row=row, column=2, value=description)
        cell.font = SMALL
        cell.alignment = WRAP_LEFT
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Rules applied").font = Font(
        bold=True, size=12, color="1F3864")
    row += 1
    rules = config.rules
    for label, value in [
        ("Minimum rest between shifts",
         f"{rules.minimum_rest_hours:g} hours (a rule configured for this "
         f"laboratory, not a compliance determination)"),
        ("Maximum consecutive days", rules.max_consecutive_days),
        ("Maximum consecutive nights", rules.max_consecutive_nights),
        ("Nights rostered in blocks of", rules.night_block_length),
        ("Recovery days after nights", rules.recovery_days_after_nights),
        ("Hours tolerance", f"{rules.hours_tolerance_percent:g}%"),
        ("Benches per person at once", rules.max_simultaneous_bench_assignments),
    ]:
        sheet.cell(row=row, column=1, value=label).font = BOLD
        cell = sheet.cell(row=row, column=2, value=value)
        cell.font = SMALL
        cell.alignment = WRAP_LEFT
        row += 1

    row += 1
    cell = sheet.cell(row=row, column=1,
                      value="Your workforce information was processed on your own "
                            "device, in your browser. It was not uploaded to "
                            "Optymum SS.")
    cell.font = Font(size=10, italic=True, color="1D6B3F")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell.alignment = WRAP_LEFT


def _roster(workbook, scheduler, analysis, audience: str = MANAGER) -> None:
    """The calendar grid: staff as rows, dates as columns.

    For a staff audience the grid shows only who is working. Absence is rendered
    the same as a non-working day, so circulating the rota does not disclose who is
    on sick leave, and the role column is dropped.
    """
    staff_facing = audience == STAFF
    sheet = workbook.create_sheet("Roster")
    config = scheduler.config
    days = scheduler.days
    first_col = 4
    last_col = first_col + len(days) - 1

    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    heading = f"{config.details.heading}   |   "               f"{config.period.start:%d %b %Y} to {config.period.end:%d %b %Y}"
    title = sheet.cell(row=1, column=1,
                       value=heading if staff_facing
                       else f"{heading}   |   DRAFT for review")
    title.fill = HEADER_FILL
    title.font = Font(bold=True, size=13, color="FFFFFF")
    title.alignment = LEFT
    sheet.row_dimensions[1].height = 24

    # Month bands.
    bands = []
    for offset, day in enumerate(days):
        label = day.strftime("%B %Y").upper()
        column = first_col + offset
        if bands and bands[-1][0] == label:
            bands[-1] = (label, bands[-1][1], bands[-1][2] + 1)
        else:
            bands.append((label, column, 1))
    for label, start_col, span in bands:
        sheet.merge_cells(start_row=2, start_column=start_col,
                          end_row=2, end_column=start_col + span - 1)
        cell = sheet.cell(row=2, column=start_col, value=label)
        cell.fill = BAND_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTRE
        cell.border = BORDER

    left_columns = ((1, "Staff"), (2, "Band")) if staff_facing         else ((1, "Staff"), (2, "Band"), (3, "Role"))
    for column, label in left_columns:
        cell = sheet.cell(row=3, column=column, value=label)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = LEFT if column == 1 else CENTRE
        cell.border = BORDER
        spacer = sheet.cell(row=4, column=column)
        spacer.fill = HEADER_FILL
        spacer.border = BORDER

    for offset, day in enumerate(days):
        column = first_col + offset
        weekend = scheduler.is_weekend(day)
        number = sheet.cell(row=3, column=column, value=day.day)
        number.fill = WEEKEND_FILL if weekend else DAY_FILL
        number.font = Font(bold=True, size=10)
        number.alignment = CENTRE
        number.border = BORDER
        name = sheet.cell(row=4, column=column, value=WEEKDAY_SHORT[day.weekday()])
        name.fill = WEEKEND_FILL if weekend else DAY_FILL
        name.font = Font(size=8, bold=weekend)
        name.alignment = CENTRE
        name.border = BORDER
        sheet.column_dimensions[get_column_letter(column)].width = 4.8

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 6
    sheet.column_dimensions["C"].width = 18

    shift_fill = {shift.code: PatternFill("solid", fgColor=shift.colour)
                  for shift in config.shifts}
    shift_font = {shift.code: Font(bold=True, size=10, color=shift.font_colour)
                  for shift in config.shifts}

    groups: dict[str, list] = defaultdict(list)
    for person in config.staff:
        groups[person.group].append(person)
    ordered = ([("Main", groups.pop("Main"))] if "Main" in groups else []) \
        + sorted(groups.items())

    row = 5
    for group_name, members in ordered:
        if len(ordered) > 1:
            sheet.merge_cells(start_row=row, start_column=1,
                              end_row=row, end_column=last_col)
            cell = sheet.cell(row=row, column=1, value=group_name.upper())
            cell.fill = GROUP_FILL
            cell.font = BOLD
            cell.alignment = LEFT
            row += 1
        for person in members:
            name_cell = sheet.cell(row=row, column=1, value=person.name)
            name_cell.font = Font(bold=person.is_senior, size=10)
            name_cell.alignment = LEFT
            name_cell.border = BORDER
            left_values = ((2, person.band),) if staff_facing                 else ((2, person.band), (3, person.job_title))
            for column, value in left_values:
                cell = sheet.cell(row=row, column=column, value=value)
                cell.font = SMALL
                cell.alignment = CENTRE if column == 2 else LEFT
                cell.border = BORDER

            for offset, day in enumerate(days):
                cell = sheet.cell(row=row, column=first_col + offset)
                cell.alignment = CENTRE
                cell.border = BORDER
                assignment = scheduler.assignments.get((day, person.staff_id))
                if assignment:
                    cell.value = assignment.shift_code + (
                        "*" if assignment.is_manual else "")
                    cell.fill = shift_fill.get(assignment.shift_code,
                                               PatternFill("solid", fgColor="D9E1F2"))
                    cell.font = shift_font.get(assignment.shift_code, BOLD)
                    continue
                code = scheduler.leave.get((day, person.staff_id))
                if code and not staff_facing:
                    entry = config.leave_types.get(
                        "".join(ch for ch in code.lower() if ch.isalnum()))
                    cell.value = code
                    cell.fill = PatternFill(
                        "solid", fgColor=entry.colour if entry else "FFFF00")
                    cell.font = Font(bold=True, size=9,
                                     color=entry.font_colour if entry else "000000")
                elif scheduler.is_weekend(day):
                    # Absence is left looking like any other non-working day.
                    cell.fill = WEEKEND_FILL
            row += 1

    # Staffing count per shift, per day.
    row += 1
    for shift in config.shifts:
        label = sheet.cell(row=row, column=1, value=f"On {shift.name}")
        label.font = BOLD
        label.fill = TOTAL_FILL
        label.alignment = LEFT
        for column in (2, 3):
            sheet.cell(row=row, column=column).fill = TOTAL_FILL
        for offset, day in enumerate(days):
            cell = sheet.cell(row=row, column=first_col + offset)
            cell.fill = TOTAL_FILL
            cell.border = BORDER
            cell.alignment = CENTRE
            cell.font = Font(bold=True, size=9)
            if shift.applies_on(day, scheduler.rules.weekend_days):
                count = len(scheduler.assigned_to(day, shift))
                requirement = config.requirement_for(shift, day)
                cell.value = count
                if requirement.min_staff and count < requirement.min_staff:
                    cell.fill = CRITICAL_FILL
                    cell.font = Font(bold=True, size=9, color="9C1F2B")
        row += 1

    # Bench allocations underneath, as initials.
    if config.benches:
        row += 1
        sheet.merge_cells(start_row=row, start_column=1,
                          end_row=row, end_column=last_col)
        cell = sheet.cell(row=row, column=1, value="SECTION ALLOCATION")
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = LEFT
        row += 1
        initials = _initials(config.staff)
        for bench in config.benches:
            for column, value in ((1, bench.name), (2, bench.discipline), (3, "")):
                cell = sheet.cell(row=row, column=column, value=value)
                cell.fill = BENCH_FILL
                cell.font = BOLD if column == 1 else SMALL
                cell.alignment = LEFT if column != 2 else CENTRE
                cell.border = BORDER
            for offset, day in enumerate(days):
                cell = sheet.cell(row=row, column=first_col + offset)
                cell.border = BORDER
                cell.alignment = WRAP
                cell.font = Font(size=7)
                allocated = scheduler.bench_staff(day, bench.name)
                cell.value = "/".join(initials[sid] for sid in allocated) or None
                wanted = bench.required_on(day, scheduler.rules.weekend_days)
                if wanted and len(allocated) < wanted:
                    cell.fill = CRITICAL_FILL
            row += 1

    # Legend, with words as well as colours.
    row += 2
    sheet.cell(row=row, column=1, value="LEGEND").font = Font(
        bold=True, size=12, color="1F3864")
    row += 1
    for shift in config.shifts:
        code = sheet.cell(row=row, column=1, value=shift.code)
        code.fill = shift_fill[shift.code]
        code.font = shift_font[shift.code]
        code.alignment = CENTRE
        code.border = BORDER
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        text = sheet.cell(row=row, column=2,
                          value=f"{shift.name} · {shift.times_label} · "
                                f"{shift.hours:g} hours"
                                f"{' · night shift' if shift.is_night else ''}")
        text.font = SMALL
        text.alignment = LEFT
        row += 1
    used_leave = {entry.code for entry in config.leave}
    for entry in (() if staff_facing else config.leave_types.values()):
        if used_leave and entry.code not in used_leave:
            continue
        code = sheet.cell(row=row, column=1, value=entry.code)
        code.fill = PatternFill("solid", fgColor=entry.colour)
        code.font = Font(bold=True, size=9, color=entry.font_colour)
        code.alignment = CENTRE
        code.border = BORDER
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        sheet.cell(row=row, column=2, value=entry.label).font = SMALL
        row += 1
    if not staff_facing:
        sheet.cell(row=row, column=1, value="*").font = BOLD
        sheet.merge_cells(start_row=row, start_column=2, end_row=row,
                          end_column=10)
        sheet.cell(row=row, column=2,
                   value="Manually adjusted by a manager").font = SMALL
        row += 1
    sheet.cell(row=row, column=1, value="—").font = BOLD
    sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    sheet.cell(row=row, column=2,
               value="A blank cell means not scheduled. Shaded columns are "
                     "weekends.").font = SMALL

    sheet.freeze_panes = sheet.cell(row=5, column=first_col).coordinate
    _landscape(sheet)


def _initials(staff) -> dict[str, str]:
    """Short labels for the section allocation rows, kept unique."""
    proposed: dict[str, str] = {}
    for person in staff:
        parts = [part for part in person.name.split() if part]
        if not parts:
            proposed[person.staff_id] = person.staff_id[:3]
        elif len(parts) == 1:
            proposed[person.staff_id] = parts[0][:2].upper()
        else:
            proposed[person.staff_id] = (parts[0][0] + parts[-1][0]).upper()

    by_code: dict[str, list[str]] = defaultdict(list)
    for staff_id, code in proposed.items():
        by_code[code].append(staff_id)
    lookup = {person.staff_id: person for person in staff}
    final: dict[str, str] = {}
    for code, owners in by_code.items():
        if len(owners) == 1:
            final[owners[0]] = code
            continue
        for staff_id in owners:
            parts = lookup[staff_id].name.split()
            surname = parts[-1] if len(parts) > 1 else parts[0]
            extended, index = code + surname[1:2].lower(), 2
            while extended in final.values() and index <= len(surname):
                extended = code + surname[1:index + 1].lower()
                index += 1
            final[staff_id] = extended
    return final


def _staff_sections(workbook, scheduler, analysis) -> None:
    """Which section each person is on, day by day — no coverage judgements.

    Staff need to know where to go. They do not need the manager's assessment of
    whether the section was adequately covered, so the Required / Allocated /
    Status columns of the manager version are left out.
    """
    sheet = workbook.create_sheet("Section Allocations")
    row = _title(sheet, "Section allocations", 5,
                 "Where each person is working. Check with your manager if "
                 "anything looks wrong.")
    row = _headers(sheet, row,
                   ["Date", "Day", "Shift", "Section", "Staff"],
                   [12, 6, 16, 24, 52])
    config = scheduler.config
    for day in scheduler.days:
        for shift in config.shifts:
            if not shift.applies_on(day, scheduler.rules.weekend_days):
                continue
            for bench in scheduler.benches_for(day, shift):
                allocated = scheduler.bench_staff(day, bench.name, shift.code)
                if not allocated:
                    continue
                row = _row(sheet, row, [
                    day, WEEKDAY_SHORT[day.weekday()], shift.name, bench.name,
                    ", ".join(analysis.name_of(sid) for sid in allocated),
                ])
    _landscape(sheet)


def _staff_notes(workbook, scheduler) -> None:
    """A short cover note so a circulated rota explains itself."""
    sheet = workbook.create_sheet("Notes")
    config = scheduler.config
    details = config.details
    row = _title(sheet, "About this rota", 6)
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 70

    for label, value in [
        ("Rota", details.rota_name),
        ("Organisation", details.organisation or "—"),
        ("Department", details.department or "—"),
        ("Site", details.site or "—"),
        ("Period", f"{config.period.start:%d %b %Y} to "
                   f"{config.period.end:%d %b %Y}"),
        ("Issued", f"{date.today():%d %b %Y}"),
    ]:
        sheet.cell(row=row, column=1, value=label).font = BOLD
        sheet.cell(row=row, column=2, value=value).font = SMALL
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Shifts").font = Font(
        bold=True, size=12, color="1F3864")
    row += 1
    for shift in config.shifts:
        code = sheet.cell(row=row, column=1, value=shift.code)
        code.fill = PatternFill("solid", fgColor=shift.colour)
        code.font = Font(bold=True, size=10, color=shift.font_colour)
        code.alignment = CENTRE
        code.border = BORDER
        sheet.cell(row=row, column=2,
                   value=f"{shift.name} · {shift.times_label} · "
                         f"{shift.hours:g} hours").font = SMALL
        row += 1

    row += 1
    for line in [
        "A blank cell means you are not scheduled to work that day.",
        "Please raise any queries about your own rota with your line manager.",
        "This rota may be adjusted; check for a later version before relying on it.",
    ]:
        cell = sheet.cell(row=row, column=1, value=line)
        cell.font = SMALL
        cell.alignment = WRAP_LEFT
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1


def _staff(workbook, scheduler, analysis) -> None:
    sheet = workbook.create_sheet("Staff")
    row = _title(sheet, "Staff included in this roster", 12)
    row = _headers(sheet, row,
                   ["Name", "Staff ID", "Job Title", "Band", "Registered",
                    "Senior", "Coordinator", "Trainee", "Contracted Weekly Hours",
                    "FTE", "Working Pattern", "Working Days", "Nights",
                    "Weekends", "Restrictions"],
                   [24, 10, 20, 7, 11, 8, 12, 9, 13, 6, 16, 20, 8, 10, 30])
    for person in scheduler.config.staff:
        days = person.availability.weekdays.get(1)
        pattern = ("Flexible" if not days
                   else ", ".join(WEEKDAY_SHORT[index] for index in sorted(days)))
        if person.availability.cycle_weeks > 1:
            pattern += f" (week 1 of {person.availability.cycle_weeks})"
        row = _row(sheet, row, [
            person.name, person.staff_id, person.job_title, person.band,
            "Yes" if person.registered else "No",
            "Yes" if person.is_senior else "No",
            "Yes" if person.shift_coordinator else "No",
            "Yes" if person.trainee else "No",
            person.contracted_weekly_hours, person.fte,
            person.working_pattern, pattern,
            "Yes" if person.nights_ok else "No",
            "Yes" if person.weekends_ok else "No",
            person.restrictions,
        ])
    _landscape(sheet)


def _competencies(workbook, scheduler, analysis) -> None:
    sheet = workbook.create_sheet("Competencies")
    config = scheduler.config
    as_of = config.period.start
    row = _title(sheet, "Competency register", 11,
                 "Only Competent, Trainer and Assessor count as coverage. "
                 "In Training and Supervised do not. Expired records never count.")
    row = _headers(sheet, row,
                   ["Name", "Discipline", "Competency", "Status", "Counts as cover",
                    "Achieved", "Review due", "Expires", "Trainer", "Assessor",
                    "Authoriser", "Notes"],
                   [24, 11, 26, 13, 15, 12, 12, 12, 9, 9, 11, 26])
    for record in sorted(config.competencies,
                         key=lambda r: (analysis.name_of(r.staff_id), r.discipline)):
        status = record.effective_status(as_of)
        counts = "Yes" if record.is_independent(as_of) else "No"
        fills = {}
        if status == "Expired":
            fills = {4: CRITICAL_FILL, 5: CRITICAL_FILL}
        elif not record.is_independent(as_of):
            fills = {4: REVIEW_FILL, 5: REVIEW_FILL}
        else:
            fills = {5: PASSED_FILL}
        row = _row(sheet, row, [
            analysis.name_of(record.staff_id), record.discipline, record.name,
            status, counts, record.date_achieved, record.review_date,
            record.expiry_date,
            "Yes" if record.can_train(as_of) else "",
            "Yes" if record.can_assess(as_of) else "",
            "Yes" if record.can_authorise(as_of) else "",
            record.notes,
        ], fills=fills)
    _landscape(sheet)


def _requirements(workbook, scheduler) -> None:
    sheet = workbook.create_sheet("Shift Requirements")
    row = _title(sheet, "What each shift was required to provide", 12)
    row = _headers(sheet, row,
                   ["Shift", "Code", "Times", "Hours", "Applies", "Min Staff",
                    "Min Registered", "Min Senior", "Coordinators", "Trainers",
                    "Max Trainees", "Required Competencies",
                    "Required Authorisers", "Notes"],
                   [16, 7, 14, 8, 12, 10, 13, 10, 12, 9, 11, 24, 22, 24])
    shift_by_code = scheduler.config.shift_by_code()
    for requirement in scheduler.config.requirements:
        shift = shift_by_code.get(requirement.shift_code)
        row = _row(sheet, row, [
            shift.name if shift else requirement.shift_code,
            requirement.shift_code,
            shift.times_label if shift else "",
            round(shift.hours, 2) if shift else "",
            requirement.days, requirement.min_staff,
            requirement.min_registered, requirement.min_senior,
            requirement.min_coordinators, requirement.min_trainers,
            requirement.max_trainees or "—",
            ", ".join(f"{k}:{v}" for k, v
                      in requirement.required_competencies.items()) or "—",
            ", ".join(f"{k}:{v}" for k, v
                      in requirement.required_authorisers.items()) or "—",
            requirement.notes,
        ])
    _landscape(sheet)


def _bench_allocations(workbook, scheduler, analysis) -> None:
    sheet = workbook.create_sheet("Bench Allocations")
    row = _title(sheet, "Section allocation, day by day", 8,
                 "One person is allocated to one section at a time, so coverage "
                 "reflects who is actually there.")
    row = _headers(sheet, row,
                   ["Date", "Day", "Shift", "Section", "Discipline", "Required",
                    "Allocated", "Status", "Staff"],
                   [12, 6, 14, 22, 11, 10, 10, 14, 40])
    config = scheduler.config
    for day in scheduler.days:
        for shift in config.shifts:
            if not shift.applies_on(day, scheduler.rules.weekend_days):
                continue
            for bench in scheduler.benches_for(day, shift):
                wanted = bench.required_on(day, scheduler.rules.weekend_days)
                allocated = scheduler.bench_staff(day, bench.name, shift.code)
                if not wanted and not allocated:
                    continue
                if wanted and len(allocated) >= wanted:
                    status, fill = "Covered", PASSED_FILL
                elif allocated:
                    status, fill = "Below requirement", REVIEW_FILL
                else:
                    status, fill = "Not covered", CRITICAL_FILL
                row = _row(sheet, row, [
                    day, WEEKDAY_SHORT[day.weekday()], shift.name, bench.name,
                    bench.discipline, wanted, len(allocated), status,
                    ", ".join(analysis.name_of(sid) for sid in allocated) or "—",
                ], fills={8: fill})
    _landscape(sheet)


def _issues(workbook, analysis) -> None:
    sheet = workbook.create_sheet("Issues")
    row = _title(sheet, "Issues for review", 8,
                 "Grouped by severity. Critical items should be resolved before "
                 "the rota is published.")
    metrics = analysis.metrics
    sheet.cell(row=row, column=1, value="Roster status").font = BOLD
    status_cell = sheet.cell(row=row, column=2, value=metrics["roster_status"])
    status_cell.font = SEVERITY_FONT[
        CRITICAL if metrics["critical_count"] else
        (REVIEW if metrics["review_count"] else PASSED)]
    status_cell.fill = SEVERITY_FILL[
        CRITICAL if metrics["critical_count"] else
        (REVIEW if metrics["review_count"] else PASSED)]
    row += 2

    row = _headers(sheet, row,
                   ["Severity", "Category", "Issue", "Date", "Shift", "Section",
                    "Staff", "What to check"],
                   [12, 22, 46, 14, 10, 18, 26, 46], height=30)

    order = {CRITICAL: 0, REVIEW: 1, PASSED: 2}
    for issue in sorted(analysis.issues,
                        key=lambda i: (order.get(i.severity, 3), i.category,
                                       i.day or date.min)):
        start_row = row
        row = _row(sheet, row, [
            issue.severity, issue.category, issue.explanation,
            issue.day, issue.shift_code, issue.bench_name,
            ", ".join(analysis.name_of(sid) for sid in issue.staff),
            issue.review_point,
        ], fills={1: SEVERITY_FILL.get(issue.severity, PASSED_FILL)},
           fonts={1: SEVERITY_FONT.get(issue.severity, SMALL)})
        for column in (3, 8):
            sheet.cell(row=start_row, column=column).alignment = WRAP_LEFT
        sheet.row_dimensions[start_row].height = 30
    _landscape(sheet)


def _hours(workbook, analysis) -> None:
    sheet = workbook.create_sheet("Hours Summary")
    row = _title(sheet, "Contracted hours against allocated hours", 12,
                 "Target hours are contracted weekly hours scaled to the length of "
                 "the roster period.")
    row = _headers(sheet, row,
                   ["Name", "Band", "Contracted Weekly", "FTE", "Target Hours",
                    "Worked Hours", "Credited Absence Hours",
                    "Total Accounted Hours", "Variance", "% of Target", "Status",
                    "Shifts", "Nights", "Saturdays", "Sundays", "Full Weekends",
                    "Absence Days"],
                   [24, 7, 14, 6, 12, 13, 16, 16, 11, 11, 16, 8, 8, 10, 9, 13, 12])
    for entry in analysis.hours_rows:
        fill = (PASSED_FILL if entry.status == "Within tolerance" else REVIEW_FILL)
        row = _row(sheet, row, [
            entry.name, entry.band, entry.contracted_weekly_hours, entry.fte,
            entry.target_hours, entry.worked_hours,
            entry.credited_absence_hours, entry.total_accounted_hours,
            entry.variance,
            entry.percent_of_target if entry.percent_of_target is not None else "—",
            entry.status, entry.shifts, entry.nights, entry.saturdays,
            entry.sundays, entry.full_weekends, entry.leave_days,
        ], fills={11: fill})

    totals = analysis.hours_rows
    row = _row(sheet, row, [
        "TOTAL", "", "", "",
        round(sum(r.target_hours for r in totals), 2),
        round(sum(r.worked_hours for r in totals), 2),
        round(sum(r.credited_absence_hours for r in totals), 2),
        round(sum(r.total_accounted_hours for r in totals), 2),
        round(sum(r.variance for r in totals), 2), "", "",
        sum(r.shifts for r in totals), sum(r.nights for r in totals),
        sum(r.saturdays for r in totals), sum(r.sundays for r in totals),
        sum(r.full_weekends for r in totals), sum(r.leave_days for r in totals),
    ], fills={column: TOTAL_FILL for column in range(1, 18)},
       fonts={column: BOLD for column in range(1, 18)})

    row += 1
    note = sheet.cell(row=row, column=1,
                      value="Target hours are contracted weekly hours scaled to "
                            "the roster period. Credited absence hours come from "
                            "each person's own working pattern, so hours lost to "
                            "leave are not made up with extra shifts.")
    note.font = MUTED
    note.alignment = WRAP_LEFT
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    _landscape(sheet)


def _fairness(workbook, scheduler, analysis) -> None:
    sheet = workbook.create_sheet("Fairness Summary")
    row = _title(sheet, "How the unpopular shifts are shared", 10,
                 "Compared only between staff eligible for the same work: night "
                 "counts between those who work nights, weekends between those who "
                 "work weekends. There is no single correct fairness formula, so "
                 "these are judgements to review rather than scores to optimise.")

    sheet.cell(row=row, column=1, value="Weekend fairness").font = BOLD
    sheet.cell(row=row, column=2, value=analysis.weekend_fairness()).font = BOLD
    row += 1
    sheet.cell(row=row, column=1, value="Night fairness").font = BOLD
    sheet.cell(row=row, column=2, value=analysis.night_fairness()).font = BOLD
    row += 2

    row = _headers(sheet, row,
                   ["Name", "Eligible for nights", "Eligible for weekends",
                    "Total Hours", "Nights", "Saturdays", "Sundays",
                    "Full Weekends", "Late shifts", "Early shifts"],
                   [24, 16, 18, 12, 9, 11, 9, 13, 12, 12])
    by_id = {person.staff_id: person for person in scheduler.config.staff}
    for entry in analysis.hours_rows:
        person = by_id[entry.staff_id]
        row = _row(sheet, row, [
            entry.name,
            "Yes" if person.nights_ok else "No",
            "Yes" if person.weekends_ok else "No",
            entry.worked_hours, entry.nights, entry.saturdays, entry.sundays,
            entry.full_weekends, entry.lates, entry.earlies,
        ])
    _landscape(sheet)


def _expiry(workbook, analysis) -> None:
    sheet = workbook.create_sheet("Competency Expiry")
    thresholds = ", ".join(str(day) for day
                           in analysis.rules.expiry_warning_days)
    row = _title(sheet, "Competencies expired or expiring", 7,
                 f"Warning thresholds configured for this laboratory: "
                 f"{thresholds} days.")
    row = _headers(sheet, row,
                   ["Name", "Discipline", "Competency", "Expires", "Days Away",
                    "State", "Action"],
                   [24, 11, 28, 13, 11, 18, 40])
    if not analysis.expiring:
        _row(sheet, row, ["No competencies have expired or expire within the "
                          "configured warning period.", "", "", "", "", "", ""],
             fills={1: PASSED_FILL})
    for item in sorted(analysis.expiring,
                       key=lambda i: (i["days"] if i["days"] is not None else 9999)):
        expired = item["state"] == "Expired"
        row = _row(sheet, row, [
            item["name"], item["discipline"], item["name_of_competency"],
            item["expiry"], item["days"], item["state"],
            "Reassess before this person is counted as cover" if expired
            else "Schedule reassessment",
        ], fills={6: CRITICAL_FILL if expired else REVIEW_FILL})
    _landscape(sheet)
