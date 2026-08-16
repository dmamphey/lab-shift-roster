"""Building the workbook a manager fills in.

Two workbooks come out of here:

* a **blank template** — headings, instructions and dropdown lists, with no
  employee records at all
* an **example laboratory** — a complete fictional department, deliberately built
  so that every kind of warning the tool can raise is triggered by it

Every name, competency and date in the example is invented.  Real staff details
must never ship inside a template: it is distributed with the product, and it
would carry personal data into every copy.

Instructions written into the workbook are for laboratory managers.  Nothing
here mentions Python, command lines or file paths.
"""

from __future__ import annotations

from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import SCHEMA_VERSION
from .models import derive_initials, CREDIT_METHODS, CompetencyStatus, WEEKDAY_SHORT

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="2E5496")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10)
WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

MANAGER_STEPS = [
    "Complete the Staff sheet: one row per member of staff.",
    "Complete the Competencies sheet: one row per person per discipline.",
    "Complete the Shifts sheet: the shifts your laboratory runs, with times. "
    "Set Active to No for any shift you do not run.",
    "Add leave and availability: the Leave sheet, and Week Patterns if anyone "
    "works alternating weeks.",
    "Configure requirements: the Shift Requirements and Benches sheets set the "
    "minimum staffing and competencies each shift needs.",
    "Check the Roster Details and Rules sheets.",
    "Save the workbook.",
    "Upload it to Lab Shift Roster in your browser.",
    "Generate the draft roster.",
    "Review the warnings Lab Shift Roster raises.",
    "Export the finished workbook.",
]

INSTRUCTION_NOTES = [
    ("How to use this workbook", True),
    ("", False),
] + [(f"{index}.  {text}", False)
     for index, text in enumerate(MANAGER_STEPS, start=1)] + [
    ("", False),
    ("Filling in the sheets", True),
    ("Staff — Staff ID must be unique and is what the other sheets refer to. "
     "Contracted Weekly Hours drives how much work each person is given.", False),
    ("Staff — the Mon to Sun columns are the days that person normally works. "
     "Leave them all blank if somebody is fully flexible.", False),
    ("Staff — Senior, Shift Coordinator and Registered are recorded separately "
     "from Band, because they are different things. A Band 7 is not "
     "automatically competent in every discipline.", False),
    ("Competencies — only Competent, Trainer and Assessor count as cover. "
     "In Training and Supervised do not, because those staff need somebody with "
     "them. Expired competencies never count.", False),
    ("Shifts — the five shifts below are a starting point, not a prescription. "
     "Set Active to No for any your laboratory does not run, and add a row for "
     "any it does. An inactive shift is never rostered, and requirements or "
     "benches that mention it are ignored while it is off — so you can switch "
     "one back on later without retyping anything.", False),
    ("Staff — Initials is optional. Leave it blank and the roster works them out "
     "from the name, keeping them unique. Fill it in if your laboratory already "
     "uses its own. These are the initials that appear in the section "
     "allocation rows.", False),
    ("Shift Requirements — Required Competencies is written as discipline and "
     "number, for example: BT:1, HAEM:2, COAG:1", False),
    ("Benches — each bench needs its own person. One member of staff is not "
     "counted as covering two benches at the same time.", False),
    ("Rules — these are your laboratory's own operating rules, including the "
     "rest interval. They are settings for planning, not a determination of "
     "legal or regulatory compliance.", False),
    ("", False),
    ("What Lab Shift Roster produces", True),
    ("A draft roster for you to review, together with a list of staffing, "
     "competency and coverage gaps. Every roster it generates is a draft "
     "requiring managerial review.", False),
]


def _sheet(workbook, title, headers, widths=None, note=None, freeze=True):
    """Create a sheet with a title row, a header row and sensible widths."""
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    row = 1
    if note:
        sheet.cell(row=1, column=1, value=note).font = Font(bold=True, size=10,
                                                            color="7F4F00")
        sheet.cell(row=1, column=1).fill = NOTE_FILL
        sheet.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=max(len(headers), 4))
        row = 2
    for column, heading in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=heading)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = WRAP
        cell.border = BORDER
    sheet.row_dimensions[row].height = 34
    for column, width in enumerate(widths or [], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    if freeze:
        sheet.freeze_panes = sheet.cell(row=row + 1, column=1).coordinate
    sheet.auto_filter.ref = (f"A{row}:"
                             f"{get_column_letter(len(headers))}{row}")
    return sheet, row


def _write_rows(sheet, header_row, rows):
    for offset, values in enumerate(rows):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=header_row + 1 + offset, column=column,
                              value=value)
            cell.border = BORDER
            cell.alignment = LEFT
            if isinstance(value, date):
                cell.number_format = "DD/MM/YYYY"


def _yes_no(sheet, header_row, columns, last_row=400):
    validation = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    sheet.add_data_validation(validation)
    for column in columns:
        letter = get_column_letter(column)
        validation.add(f"{letter}{header_row + 1}:{letter}{last_row}")


def _colour_previews(sheet, header_row, row_count) -> None:
    """Show the configured shift colours, not just their hex codes.

    A manager choosing colours should be able to see them.  The Colour cell is
    filled with the background colour, the Font Colour cell is filled white with
    that font colour applied, and the Preview cell shows the shift name exactly as
    it will appear on the roster.
    """
    # Moved right by one when Active was inserted at column 6.
    colour_column, font_column, preview_column = 8, 9, 10
    for offset in range(row_count):
        row = header_row + 1 + offset
        background = str(sheet.cell(row=row, column=colour_column).value
                         or "FFFFFF").strip().lstrip("#").upper()
        foreground = str(sheet.cell(row=row, column=font_column).value
                         or "000000").strip().lstrip("#").upper()
        name = sheet.cell(row=row, column=2).value or ""

        swatch = sheet.cell(row=row, column=colour_column)
        swatch.fill = PatternFill("solid", fgColor=background)
        swatch.font = Font(bold=True, color=foreground, size=10)
        swatch.alignment = WRAP

        font_cell = sheet.cell(row=row, column=font_column)
        font_cell.font = Font(bold=True, color=foreground, size=10)
        font_cell.alignment = WRAP

        preview = sheet.cell(row=row, column=preview_column,
                             value=str(name).upper())
        preview.fill = PatternFill("solid", fgColor=background)
        preview.font = Font(bold=True, color=foreground, size=11)
        preview.alignment = WRAP
        preview.border = BORDER


def _leave_colour_previews(sheet, header_row, row_count) -> None:
    """The same courtesy for absence codes."""
    for offset in range(row_count):
        row = header_row + 1 + offset
        background = str(sheet.cell(row=row, column=3).value
                         or "FFFFFF").strip().lstrip("#").upper()
        foreground = str(sheet.cell(row=row, column=4).value
                         or "000000").strip().lstrip("#").upper()
        for column in (3, 4):
            cell = sheet.cell(row=row, column=column)
            cell.fill = PatternFill("solid", fgColor=background)
            cell.font = Font(bold=True, color=foreground, size=10)
            cell.alignment = WRAP


def _list_validation(sheet, header_row, column, options, last_row=400):
    formula = '"' + ",".join(options) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    sheet.add_data_validation(validation)
    letter = get_column_letter(column)
    validation.add(f"{letter}{header_row + 1}:{letter}{last_row}")


# --------------------------------------------------------------------------
# column definitions, shared by both workbooks
# --------------------------------------------------------------------------

STAFF_HEADERS = [
    "Staff ID", "Name", "Initials", "Job Title", "Band", "Registered BMS", "Senior",
    "Shift Coordinator", "Trainee", "Contracted Weekly Hours", "FTE",
    "Working Pattern", "Pattern Cycle (weeks)",
    *WEEKDAY_SHORT,
    "Earliest Start", "Latest Finish", "Max Days Per Week",
    "Max Consecutive Days", "Works Nights", "Works Weekends", "Max Nights",
    "Max Weekends", "Max Hours This Period", "Max Weekly Hours",
    "Group", "Restrictions", "Notes",
]
STAFF_WIDTHS = [10, 20, 9, 20, 7, 11, 8, 12, 8, 13, 6, 16, 11,
                5, 5, 5, 5, 5, 5, 5, 11, 11, 11, 12, 10, 11, 10, 11, 13, 13,
                10, 20, 20]

COMPETENCY_HEADERS = [
    "Staff ID", "Name", "Discipline", "Competency", "Status", "Date Achieved",
    "Review Date", "Expiry Date", "Trainer", "Assessor", "Result Authoriser",
    "Notes",
]
COMPETENCY_WIDTHS = [10, 20, 11, 26, 13, 13, 13, 13, 9, 9, 14, 22]

SHIFT_HEADERS = ["Code", "Name", "Start", "End", "Days", "Active", "Night Shift",
                 "Colour", "Font Colour", "Preview"]
SHIFT_WIDTHS = [8, 16, 9, 9, 12, 9, 11, 10, 11, 16]

REQUIREMENT_HEADERS = [
    "Shift Code", "Days", "Min Staff", "Min Registered BMS", "Min Senior",
    "Min Band", "Staff At Min Band", "Min Coordinators", "Min Trainers",
    "Max Trainees", "Required Competencies", "Required Authorisers", "Notes",
]
REQUIREMENT_WIDTHS = [10, 12, 10, 12, 10, 9, 12, 13, 11, 11, 24, 22, 20]

BENCH_HEADERS = ["Bench", "Discipline", "Days", "Min Staff",
                 "Min Staff Weekend", "Shift Codes", "Requires Authoriser",
                 "Rotation Interval (days)"]
BENCH_WIDTHS = [24, 12, 12, 10, 16, 14, 16, 18]

LEAVE_HEADERS = ["Staff ID", "Name", "From", "To", "Type", "Credited Hours",
                 "Reason"]
LEAVE_WIDTHS = [10, 20, 13, 13, 10, 14, 24]

PATTERN_HEADERS = ["Staff ID", "Week", *WEEKDAY_SHORT]
PATTERN_WIDTHS = [10, 8, 5, 5, 5, 5, 5, 5, 5]

LEAVE_TYPE_HEADERS = ["Code", "Label", "Colour", "Font Colour",
                      "Counts Towards Contracted Hours",
                      "Credited Hours Method", "Fixed Hours Per Day"]
LEAVE_TYPE_WIDTHS = [10, 30, 10, 11, 22, 22, 17]

#: The seeded shifts. Every laboratory works a different pattern, so these are a
#: starting point to switch off or edit, not a prescription — which is what the
#: Active column is for.
DEFAULT_SHIFTS = [
    ("E", "Early", "07:00", "15:00", "Weekday", "Y", "N", "FFF2CC", "000000"),
    ("C", "Core", "09:00", "17:30", "Weekday", "Y", "N", "D9E1F2", "000000"),
    ("L", "Late", "13:00", "21:00", "Weekday", "Y", "N", "00B0F0", "000000"),
    ("N", "Night", "21:00", "07:00", "All", "Y", "Y", "7030A0", "FFFFFF"),
    ("W", "Weekend Day", "09:00", "17:30", "Weekend", "Y", "N", "00B050", "FFFFFF"),
]

# "Counts Towards Contracted Hours" is what stops the roster handing somebody
# extra shifts to make up hours they were away for.  Every standard absence
# credits hours by default; change it if your organisation works differently.
DEFAULT_LEAVE_TYPES = [
    ("A/L", "Annual leave", "FFFF00", "000000", "Y", "Working pattern", ""),
    ("S/L", "Sickness absence", "FF9999", "000000", "Y", "Working pattern", ""),
    ("C/L", "Carers or compassionate leave", "FFC000", "000000", "Y",
     "Working pattern", ""),
    ("M/L", "Maternity or paternity leave", "D9D9D9", "000000", "Y",
     "Working pattern", ""),
    ("S/D", "Study or training day", "CC99FF", "000000", "Y",
     "Working pattern", ""),
]

#: Shift Codes says which shifts a section must be staffed during.  Transfusion
#: and haematology run whenever the laboratory is open; coagulation and
#: morphology only during the main day service.
DEFAULT_BENCHES = [
    ("Blood Transfusion", "BT", "All", 1, 1, "C, L, W", "Y", 42),
    ("Haematology", "HAEM", "All", 1, 1, "C, L, W", "N", 42),
    ("Coagulation", "COAG", "Weekday", 1, 0, "C", "N", 56),
    ("Morphology", "MORPH", "Weekday", 1, 0, "C", "N", 56),
]

# Sized so the example workforce can genuinely deliver it, while still leaving
# the deliberate morphology gap that demonstrates a single point of failure.
DEFAULT_REQUIREMENTS = [
    ("E", "Weekday", 2, 1, 0, 0, 0, 0, 0, 1, "", "", ""),
    ("C", "Weekday", 5, 3, 1, 6, 1, 1, 0, 2,
     "BT:1, HAEM:1, COAG:1, MORPH:1", "BT:1", "Main day service"),
    ("L", "Weekday", 2, 1, 1, 0, 0, 0, 0, 1, "BT:1", "BT:1", ""),
    ("N", "All", 1, 1, 1, 0, 0, 0, 0, 0, "BT:1, HAEM:1", "BT:1",
     "Lone working out of hours"),
    # A two-person weekend, with the senior on duty also coordinating. Three was
    # about four per cent more work than this workforce is contracted for, which
    # pushed most people slightly over their target hours.
    ("W", "Weekend", 2, 2, 1, 0, 0, 0, 0, 1, "BT:1, HAEM:1", "BT:1", ""),
]

RULES_ROWS = [
    ("Minimum rest hours between shifts", 11,
     "Your laboratory's own rest rule. Not a legal compliance determination."),
    ("Maximum consecutive days", 6, "Longest unbroken run of working days."),
    ("Maximum consecutive nights", 4, "Longest unbroken run of night shifts."),
    ("Night block length", 3, "Nights are rostered in blocks of this length."),
    ("Recovery days after nights", 2, "Days off after a block of nights."),
    ("Hours tolerance (%)", 10,
     "How far from contracted hours before somebody is flagged for review."),
    ("Competency expiry warnings (days)", "30, 60, 90",
     "Warn when a competency expires within these periods."),
    ("Max simultaneous bench assignments", 1,
     "How many benches one person may cover at the same time. Normally 1."),
    ("Cross cover allowed", "N",
     "Reserved for future use: allowing one person to cover two sections."),
    ("Weekend days", "Saturday, Sunday", "Which days count as the weekend."),
    ("Share nights evenly", "Y",
     "Spread nights across everyone eligible rather than concentrating them. A "
     "night block also books the recovery days after it, so sharing can reduce "
     "day-service cover slightly. Set to N only if that cover matters more."),
    ("Section rotation warning (days)", 56,
     "Warn if somebody has not worked a section for this long."),
    ("Alternative roster number", 42,
     "Change this to generate a different draft from the same information."),
]


def _instructions_sheet(workbook, is_demo: bool,
                        balanced: bool = False) -> None:
    sheet = workbook.active
    sheet.title = "Instructions"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 104

    sheet["A1"] = "Lab Shift Roster"
    sheet["A1"].font = Font(bold=True, size=18, color="1F3864")
    sheet["A2"] = "A free, secure and intelligent workforce planning tool for diagnostic laboratories."
    sheet["A2"].font = Font(size=11, italic=True, color="5B6470")

    row = 4
    if is_demo and balanced:
        banner = ("BALANCED EXAMPLE LABORATORY — a well-staffed fictional "
                  "department, included to show what Lab Shift Roster produces when a "
                  "workforce comfortably covers its requirements. Every person, "
                  "competency and date is fictional.")
        ink, paper = "17603A", "E4F4EA"
    elif is_demo:
        banner = ("CHALLENGING EXAMPLE LABORATORY — this example deliberately "
                  "contains staffing, competency, leave and availability "
                  "constraints so you can see how Lab Shift Roster identifies problems. "
                  "It is not typical. Every person, competency and date is "
                  "fictional.")
        ink, paper = "9C0006", "FBE6E8"
    else:
        banner = ("BLANK WORKBOOK — the Staff, Competencies, Leave and Week "
                  "Patterns sheets are empty and ready for your own information. "
                  "The Shifts, Requirements and Benches sheets contain a common "
                  "starting configuration for you to adjust.")
        ink, paper = "1D6B3F", "E3F4E9"

    # One place applies the styling and advances the row, so no branch can forget
    # to and let the instructions below overwrite the banner.
    cell = sheet.cell(row=row, column=1, value=banner)
    cell.font = Font(bold=True, size=11, color=ink)
    cell.fill = PatternFill("solid", fgColor=paper)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[row].height = 44
    row += 2

    for text, is_heading in INSTRUCTION_NOTES:
        cell = sheet.cell(row=row, column=1, value=text)
        if is_heading:
            cell.font = Font(bold=True, size=12, color="1F3864")
        else:
            cell.font = Font(size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if len(text) > 90:
                sheet.row_dimensions[row].height = 28
        row += 1

    row += 1
    sheet.cell(row=row, column=1,
               value="Your workforce information is processed on your own device, "
                     "in your browser. It is not uploaded to Optymum SS. Please do "
                     "not add sensitive personal details or any patient "
                     "information: this workbook does not need them.")
    sheet.cell(row=row, column=1).font = Font(size=10, italic=True,
                                              color="1D6B3F")
    sheet.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[row].height = 40


def _details_sheet(workbook, period_start: date, period_end: date,
                   is_demo: bool, balanced: bool = False) -> None:
    sheet, header_row = _sheet(workbook, "Roster Details", ["Detail", "Value"],
                               widths=[30, 40], freeze=False)
    rows = [
        ("Rota name", ("Balanced Example Rota" if balanced
                       else "Laboratory Staff Rota")),
        ("Organisation", "Example NHS Trust" if is_demo else ""),
        ("Department", ("Blood Sciences (balanced example)" if balanced
                        else "Blood Sciences") if is_demo else ""),
        ("Site", "Example Hospital" if is_demo else ""),
        ("Prepared by", "A. Manager (example)" if is_demo else ""),
        ("Roster period start", period_start),
        ("Roster period end", period_end),
        ("Workbook version", SCHEMA_VERSION),
    ]
    _write_rows(sheet, header_row, rows)
    for offset in range(len(rows)):
        sheet.cell(row=header_row + 1 + offset, column=1).font = Font(bold=True,
                                                                      size=10)


def _rules_sheet(workbook) -> None:
    sheet, header_row = _sheet(workbook, "Rules",
                               ["Setting", "Value", "What it means"],
                               widths=[36, 18, 68], freeze=False)
    _write_rows(sheet, header_row, RULES_ROWS)
    for offset in range(len(RULES_ROWS)):
        sheet.cell(row=header_row + 1 + offset, column=1).font = Font(bold=True,
                                                                      size=10)
        sheet.cell(row=header_row + 1 + offset, column=3).font = Font(size=9,
                                                                      color="5B6470")
        sheet.cell(row=header_row + 1 + offset, column=3).alignment = \
            Alignment(wrap_text=True, vertical="center")


def _configuration_sheets(workbook, requirements=None) -> None:
    """Shifts, requirements, benches and leave types: the same in both workbooks."""
    sheet, header_row = _sheet(
        workbook, "Shifts", SHIFT_HEADERS, SHIFT_WIDTHS,
        note="Colour and Font Colour are hex codes. The cells show the actual "
             "colour, and Preview shows how the shift will look on the roster.")
    _write_rows(sheet, header_row, DEFAULT_SHIFTS)
    _colour_previews(sheet, header_row, len(DEFAULT_SHIFTS))
    # Active (6) and Night Shift (7). Active is offered on every row down the
    # sheet, not just the seeded ones, so a shift added later can be switched
    # off the same way.
    _yes_no(sheet, header_row, [6, 7], last_row=60)
    _list_validation(sheet, header_row, 5,
                     ["All", "Weekday", "Weekend"], last_row=60)

    sheet, header_row = _sheet(
        workbook, "Shift Requirements", REQUIREMENT_HEADERS, REQUIREMENT_WIDTHS,
        note="Required Competencies and Required Authorisers are written as "
             "discipline and number, for example:  BT:1, HAEM:2, COAG:1")
    _write_rows(sheet, header_row, requirements or DEFAULT_REQUIREMENTS)
    _list_validation(sheet, header_row, 2,
                     ["All", "Weekday", "Weekend", "Monday", "Tuesday",
                      "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
                     last_row=100)

    sheet, header_row = _sheet(
        workbook, "Benches", BENCH_HEADERS, BENCH_WIDTHS,
        note="Each bench needs its own person. One member of staff is never "
             "counted as covering two benches at the same time. Shift Codes says "
             "which shifts the section must be staffed during — leave it blank to "
             "mean every shift except nights.")
    _write_rows(sheet, header_row, DEFAULT_BENCHES)
    _yes_no(sheet, header_row, [7], last_row=60)
    _list_validation(sheet, header_row, 3, ["All", "Weekday", "Weekend"],
                     last_row=60)

    sheet, header_row = _sheet(
        workbook, "Leave Types", LEAVE_TYPE_HEADERS, LEAVE_TYPE_WIDTHS,
        note="'Counts Towards Contracted Hours' stops Lab Shift Roster giving somebody "
             "extra shifts to make up hours they were away for. Credited hours "
             "come from that person's own working pattern unless you set a fixed "
             "figure.")
    _write_rows(sheet, header_row, DEFAULT_LEAVE_TYPES)
    _leave_colour_previews(sheet, header_row, len(DEFAULT_LEAVE_TYPES))
    _yes_no(sheet, header_row, [5], last_row=60)
    _list_validation(sheet, header_row, 6, CREDIT_METHODS, last_row=60)


def _empty_people_sheets(workbook) -> None:
    sheet, header_row = _sheet(
        workbook, "Staff", STAFF_HEADERS, STAFF_WIDTHS,
        note="One row per member of staff. Staff ID must be unique — the other "
             "sheets refer to it. Leave the Mon–Sun columns blank for somebody "
             "who is fully flexible.")
    # Every index moved right by one when Initials was inserted at
    # column 3. Written out rather than offset in a loop so a reader
    # can check them against STAFF_HEADERS by counting.
    _yes_no(sheet, header_row, [6, 7, 8, 9, *range(14, 21), 25, 26])

    sheet, header_row = _sheet(
        workbook, "Competencies", COMPETENCY_HEADERS, COMPETENCY_WIDTHS,
        note="Only Competent, Trainer and Assessor count as coverage. "
             "In Training and Supervised do not. Expired competencies never count.")
    _list_validation(sheet, header_row, 5, CompetencyStatus.ALL)
    _yes_no(sheet, header_row, [9, 10, 11])

    sheet, header_row = _sheet(workbook, "Week Patterns", PATTERN_HEADERS,
                               PATTERN_WIDTHS,
                               note="Only needed for staff who work alternating "
                                    "weeks. One row per person per week.")
    _yes_no(sheet, header_row, list(range(3, 10)))

    sheet, header_row = _sheet(workbook, "Leave", LEAVE_HEADERS, LEAVE_WIDTHS,
                               note="From and To are both included in the "
                                    "absence.")
    _list_validation(sheet, header_row, 5,
                     [code for code, *_ in DEFAULT_LEAVE_TYPES])


def build_blank_template(path) -> None:
    """A workbook with headings, instructions and dropdowns, and no staff."""
    workbook = Workbook()
    today = date.today()
    first = date(today.year + (today.month == 12), (today.month % 12) + 1, 1)
    last = (date(first.year + (first.month == 12), (first.month % 12) + 1, 1)
            - timedelta(days=1))

    _instructions_sheet(workbook, is_demo=False)
    _details_sheet(workbook, first, last, is_demo=False)
    _rules_sheet(workbook)
    _empty_people_sheets(workbook)
    _configuration_sheets(workbook)
    workbook.save(path)


# --------------------------------------------------------------------------
# the example laboratory
# --------------------------------------------------------------------------

def _with_initials(rows):
    """Insert an Initials value after the name in each staff row.

    The demo rows are positional tuples, and hand-editing a dozen of them to add
    a column is exactly how a value ends up one place out.

    The values come from the same function the exporter uses, so the examples
    show the real convention rather than a second approximation of it. A first
    attempt derived them here instead, and immediately gave two of the fictional
    staff the same initials — the copy had the derivation but not the uniqueness
    step.
    """
    codes = derive_initials((row[0], row[1], "") for row in rows)
    return [(row[0], row[1], codes.get(row[0], ""), *row[2:]) for row in rows]


def _demo_staff(period_start: date):
    """A fictional department, built to exercise every warning.

    Deliberate features: part-time staff with fixed non-working days, an
    alternating two-week pattern, staff who cannot work nights or weekends, a
    trainee, a single morphology-competent scientist (a workforce single point of
    failure), an expired competency and one expiring shortly.
    """
    Y, N = "Y", "N"
    # id, name, title, band, reg, senior, coord, trainee, hours, fte, pattern,
    # cycle, Mon..Sun, earliest, latest, maxdays, maxconsec, nights, weekends,
    # maxnights, maxweekends, maxperiodhours, maxweeklyhours, group,
    # restrictions, notes
    return [
        ("S01", "Alex Sample", "Laboratory Manager", "8a", Y, Y, Y, N, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, N, N, "", "", "", "", N, N, "", "", 0,
         "", "Main", "Management time; no nights", ""),
        ("S02", "Jordan Test", "Senior BMS", "7", Y, Y, Y, N, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, Y, Y, "", "", "", "", Y, Y, 0, 0, 0,
         "", "Main", "", ""),
        ("S03", "Casey Example", "Senior BMS", "6", Y, Y, Y, N, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, Y, Y, "", "", "", "", Y, Y, 0, 0, 0,
         "", "Main", "", ""),
        ("S04", "Morgan Demo", "BMS", "6", Y, Y, N, N, 30.0, 0.8,
         "Part time", 1, Y, Y, N, Y, Y, Y, Y, "", "", 4, "", Y, Y, 0, 2, 0,
         "", "Main", "No Wednesdays; maximum 2 weekends", ""),
        ("S05", "Riley Placeholder", "BMS", "6", Y, Y, N, N, 22.5, 0.6,
         "Part time", 1, N, Y, Y, Y, N, N, N, "", "", 3, "", N, N, 0, 0, 0,
         "", "Main", "No Mondays or Fridays; no nights or weekends", ""),
        ("S06", "Jamie Mock", "BMS", "5", Y, N, N, N, 22.5, 0.6,
         "Alternating weeks", 2, "", "", "", "", "", "", "", "", "", "", "",
         Y, Y, 0, 0, 0, "", "Main",
         "Alternating two-week pattern, three days each week", ""),
        ("S07", "Taylor Dummy", "BMS", "5", Y, N, N, N, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, Y, Y, "", "", "", "", Y, Y, 0, 0, 0,
         "", "Main", "", ""),
        ("S08", "Avery Specimen", "BMS", "5", Y, N, N, N, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, Y, Y, "", "", "", "", Y, Y, 0, 0, 0,
         "", "Main", "", ""),
        ("S09", "Quinn Template", "BMS", "5", Y, N, N, N, 18.75, 0.5,
         "Part time", 1, N, N, Y, Y, Y, N, N, "", "16:00", 3, "", N, N, 0, 0, 0,
         "", "Main", "Wednesday to Friday, finishing by 16:00", ""),
        ("S10", "Rowan Draft", "BMS", "5", Y, N, N, N, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, Y, Y, "", "", "", "", Y, Y, 0, 0, 0,
         "", "Main", "", ""),
        ("S11", "Skyler Trial", "Associate Practitioner", "4", N, N, N, N,
         37.5, 1.0, "Full time", 1, Y, Y, Y, Y, Y, Y, Y, "", "", "", "",
         N, Y, 0, 0, 0, "", "Main", "Not registered; no nights", ""),
        ("S12", "Parker Proxy", "Trainee BMS", "4", N, N, N, Y, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, N, N, "", "", "", "", N, N, 0, 0, 0,
         "", "Main", "Trainee; supervised working only", ""),
        ("S13", "Reese Stub", "BMS", "6", Y, Y, Y, N, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, Y, Y, "", "", "", "", Y, Y, 0, 0, 0,
         "", "Main", "", ""),
        ("S14", "Devon Filler", "BMS", "5", Y, N, N, N, 37.5, 1.0,
         "Full time", 1, Y, Y, Y, Y, Y, Y, Y, "", "", "", "", Y, Y, 0, 0, 0,
         "", "Main", "", ""),
        ("S15", "Harper Model", "Support Worker", "3", N, N, N, N, 22.5, 0.6,
         "Part time", 1, N, N, Y, Y, Y, N, N, "", "", 3, "", N, N, 0, 0, 0,
         "", "Support", "Specimen reception only", ""),
    ]


def _demo_patterns():
    """Jamie Mock works Mon–Wed one week, Wed–Fri the next."""
    Y, N = "Y", "N"
    return [
        ("S06", 1, Y, Y, Y, N, N, N, N),
        ("S06", 2, N, N, Y, Y, Y, N, N),
    ]


def _demo_competencies(period_start: date):
    """Fictional competencies, arranged to trigger each competency warning."""
    soon = period_start + timedelta(days=21)          # expires within 30 days
    later = period_start + timedelta(days=200)
    lapsed = period_start - timedelta(days=14)        # already expired
    achieved = period_start - timedelta(days=400)
    C, T, A, S, IT = (CompetencyStatus.COMPETENT, CompetencyStatus.TRAINER,
                      CompetencyStatus.ASSESSOR, CompetencyStatus.SUPERVISED,
                      CompetencyStatus.IN_TRAINING)
    Y, N = "Y", "N"
    # id, name, discipline, competency, status, achieved, review, expiry,
    # trainer, assessor, authoriser, notes
    return [
        ("S01", "Alex Sample", "HAEM", "Full blood count reporting", T,
         achieved, None, later, Y, Y, Y, ""),
        ("S01", "Alex Sample", "BT", "Transfusion authorisation", C,
         achieved, None, later, N, N, Y, ""),
        ("S02", "Jordan Test", "BT", "Crossmatch and electronic issue", T,
         achieved, None, later, Y, Y, Y, ""),
        ("S02", "Jordan Test", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, Y, ""),
        ("S02", "Jordan Test", "COAG", "Routine coagulation", C,
         achieved, None, later, N, N, N, ""),
        ("S03", "Casey Example", "BT", "Crossmatch and electronic issue", C,
         achieved, None, soon, N, N, Y, "Reassessment due"),
        ("S03", "Casey Example", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, N, ""),
        ("S04", "Morgan Demo", "BT", "Crossmatch and electronic issue", C,
         achieved, None, later, N, N, Y, ""),
        ("S04", "Morgan Demo", "COAG", "Routine coagulation", C,
         achieved, None, later, N, N, N, ""),
        ("S05", "Riley Placeholder", "MORPH", "Blood film morphology", C,
         achieved, None, later, Y, N, Y, "Only morphology scientist"),
        ("S05", "Riley Placeholder", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, N, ""),
        ("S06", "Jamie Mock", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, N, ""),
        ("S06", "Jamie Mock", "BT", "Crossmatch and electronic issue", C,
         achieved, None, lapsed, N, N, N, "Lapsed — needs reassessment"),
        ("S07", "Taylor Dummy", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, N, ""),
        ("S07", "Taylor Dummy", "COAG", "Routine coagulation", C,
         achieved, None, later, N, N, N, ""),
        ("S08", "Avery Specimen", "BT", "Crossmatch and electronic issue", C,
         achieved, None, later, N, N, N, ""),
        ("S08", "Avery Specimen", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, N, ""),
        ("S09", "Quinn Template", "COAG", "Routine coagulation", C,
         achieved, None, later, N, N, N, ""),
        ("S10", "Rowan Draft", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, N, ""),
        ("S10", "Rowan Draft", "BT", "Crossmatch and electronic issue", S,
         None, None, None, N, N, N, "Working towards sign-off"),
        ("S11", "Skyler Trial", "HAEM", "Sample processing", C,
         achieved, None, later, N, N, N, ""),
        ("S12", "Parker Proxy", "HAEM", "Full blood count reporting", IT,
         None, None, None, N, N, N, "In training"),
        ("S12", "Parker Proxy", "BT", "Crossmatch and electronic issue", IT,
         None, None, None, N, N, N, "In training"),
        ("S12", "Parker Proxy", "MORPH", "Blood film morphology", IT,
         None, None, None, N, N, N,
         "In training — does not count as morphology cover"),
        ("S13", "Reese Stub", "BT", "Crossmatch and electronic issue", C,
         achieved, None, later, N, N, Y, ""),
        ("S13", "Reese Stub", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, Y, ""),
        ("S14", "Devon Filler", "HAEM", "Full blood count reporting", C,
         achieved, None, later, N, N, N, ""),
        ("S14", "Devon Filler", "COAG", "Routine coagulation", C,
         achieved, None, later, N, N, N, ""),
        ("S15", "Harper Model", "HAEM", "Sample reception", S,
         None, None, None, N, N, N, "Supervised"),
    ]


def _demo_leave(period_start: date, period_end: date):
    return [
        ("S02", "Jordan Test", period_start + timedelta(days=7),
         period_start + timedelta(days=11), "A/L", "", "Booked leave"),
        ("S07", "Taylor Dummy", period_start + timedelta(days=14),
         period_start + timedelta(days=20), "A/L", "", "Booked leave"),
        ("S13", "Reese Stub", period_start + timedelta(days=2),
         period_start + timedelta(days=3), "S/L", "", "Recorded absence"),
        ("S08", "Avery Specimen", period_start + timedelta(days=21),
         period_start + timedelta(days=21), "S/D", "", "Training day"),
    ]


# --------------------------------------------------------------------------
# the balanced demonstration laboratory
# --------------------------------------------------------------------------

def _staff_row(staff_id, name, title, band, *, registered="Y", senior="N",
               coordinator="N", trainee="N", hours=37.5, fte=1.0,
               pattern="Full time", cycle=1, days="", earliest="",
               latest="", max_days="", max_consecutive="", nights="Y",
               weekends="Y", max_nights="", max_weekends="", max_period="",
               max_weekly="", group="Main", restrictions="", notes=""):
    """Build one Staff row without hand-writing thirty-two columns.

    ``days`` is a compact pattern using one letter per day:
    ``M T W R F S U`` for Monday to Sunday, so Thursday is ``R`` and Sunday ``U``.
    An empty string means fully flexible, which is what a full-time member of staff
    normally is.

    Note that these columns are *availability*: somebody whose pattern excludes
    Saturday and Sunday can never be rostered at a weekend, whatever the
    "Works Weekends" column says.
    """
    letters = ["M", "T", "W", "R", "F", "S", "U"]        # R = Thursday, U = Sunday
    marks = ["Y" if letter in days.upper() else "N" for letter in letters] \
        if days else ["", "", "", "", "", "", ""]
    return (staff_id, name, title, band, registered, senior, coordinator,
            trainee, hours, fte, pattern, cycle, *marks, earliest, latest,
            max_days, max_consecutive, nights, weekends, max_nights,
            max_weekends, max_period, max_weekly, group, restrictions, notes)


def _balanced_staff():
    """A fictional department that is genuinely well staffed.

    Deliberately resilient: at least three independently competent staff in every
    discipline, several authorisers, and part-time patterns spread across the week
    rather than clustered at the start.  Contracted hours are sized to the service
    this laboratory actually runs, so most people land close to their target.

    Result authorisation and morphology sit with the senior group, which is how
    many laboratories work.  That concentrates demand on those people, so this
    example carries a deep senior pool rather than pushing authorisation down the
    grades: nine of the fourteen are senior, several of them part-time.
    """
    return [
        _staff_row("B01", "Ada Sample", "Laboratory Manager", "8a",
                   senior="Y", coordinator="Y", nights="N", weekends="N",
                   restrictions="Management time; no nights or weekends"),
        _staff_row("B02", "Bruno Test", "Principal BMS", "7",
                   senior="Y", coordinator="Y"),
        _staff_row("B03", "Cleo Example", "Senior BMS", "6",
                   senior="Y", coordinator="Y"),
        _staff_row("B04", "Dara Demo", "Senior BMS", "6",
                   senior="Y", coordinator="Y"),
        _staff_row("B05", "Emre Placeholder", "Senior BMS", "6",
                   senior="Y", coordinator="Y"),
        _staff_row("B06", "Fen Mock", "Senior BMS", "6", senior="Y"),
        _staff_row("B07", "Gale Dummy", "Senior BMS", "6", senior="Y"),
        _staff_row("B08", "Hero Specimen", "Senior BMS", "6", senior="Y",
                   hours=30.0, fte=0.8, pattern="Part time", days="MTWRS",
                   restrictions="No Fridays; available Saturdays"),
        _staff_row("B09", "Iris Template", "Senior BMS", "6", senior="Y",
                   hours=30.0, fte=0.8, pattern="Part time", days="MTRFU",
                   restrictions="No Wednesdays; available Sundays"),
        _staff_row("B10", "Jules Draft", "BMS", "5", hours=30.0, fte=0.8,
                   pattern="Part time", days="TWRFS", nights="N",
                   restrictions="No Mondays; available Saturdays; no nights"),
        _staff_row("B11", "Kit Trial", "BMS", "5", hours=30.0, fte=0.8,
                   pattern="Part time", days="MWRFU", nights="N",
                   restrictions="No Tuesdays; available Sundays; no nights"),
        _staff_row("B12", "Lior Proxy", "BMS", "5", hours=22.5, fte=0.6,
                   pattern="Part time", days="MTW", weekends="N", nights="N",
                   restrictions="Monday to Wednesday; no nights or weekends"),
        _staff_row("B13", "Mira Stub", "BMS", "5", hours=22.5, fte=0.6,
                   pattern="Part time", days="WRF", nights="N", weekends="N",
                   restrictions="Wednesday to Friday; no nights or weekends"),
        _staff_row("B14", "Noor Filler", "Trainee BMS", "4", hours=22.5, fte=0.6,
                   pattern="Part time", days="MTW",
                   registered="N", trainee="Y", nights="N", weekends="N",
                   restrictions="Trainee; supervised working only"),
    ]


def _balanced_competencies(period_start: date):
    """Every discipline has at least three independently competent staff.

    Transfusion and haematology are widely held, which is what allows out-of-hours
    cover to rotate around the team instead of resting on two or three people.
    """
    later = period_start + timedelta(days=400)
    achieved = period_start - timedelta(days=500)
    C, T, A, IT = (CompetencyStatus.COMPETENT, CompetencyStatus.TRAINER,
                   CompetencyStatus.ASSESSOR, CompetencyStatus.IN_TRAINING)
    Y, N = "Y", "N"

    # staff, discipline, competency, status, trainer, assessor, authoriser
    plan = [
        ("B01", "HAEM", "Full blood count reporting", T, Y, Y, Y),
        ("B01", "MORPH", "Blood film morphology", T, Y, Y, Y),
        ("B02", "BT", "Crossmatch and electronic issue", T, Y, Y, Y),
        ("B02", "HAEM", "Full blood count reporting", C, N, N, Y),
        ("B02", "MORPH", "Blood film morphology", C, N, N, Y),
        ("B03", "BT", "Crossmatch and electronic issue", C, N, N, Y),
        ("B03", "HAEM", "Full blood count reporting", C, N, N, Y),
        ("B03", "COAG", "Routine coagulation", C, N, N, N),
        ("B04", "BT", "Crossmatch and electronic issue", C, N, N, Y),
        ("B04", "HAEM", "Full blood count reporting", C, N, N, Y),
        ("B04", "COAG", "Routine coagulation", A, Y, Y, N),
        ("B04", "MORPH", "Blood film morphology", C, N, N, Y),
        ("B05", "BT", "Crossmatch and electronic issue", C, N, N, Y),
        ("B05", "HAEM", "Full blood count reporting", C, N, N, Y),
        ("B05", "MORPH", "Blood film morphology", C, N, N, N),
        ("B06", "BT", "Crossmatch and electronic issue", C, N, N, Y),
        ("B06", "HAEM", "Full blood count reporting", C, N, N, Y),
        ("B06", "COAG", "Routine coagulation", C, N, N, N),
        ("B07", "BT", "Crossmatch and electronic issue", C, N, N, Y),
        ("B07", "HAEM", "Full blood count reporting", C, N, N, N),
        ("B07", "COAG", "Routine coagulation", C, N, N, N),
        ("B08", "BT", "Crossmatch and electronic issue", C, N, N, Y),
        ("B08", "HAEM", "Full blood count reporting", C, N, N, N),
        ("B08", "COAG", "Routine coagulation", C, N, N, N),
        ("B08", "MORPH", "Blood film morphology", C, N, N, Y),
        ("B09", "BT", "Crossmatch and electronic issue", C, N, N, Y),
        ("B09", "HAEM", "Full blood count reporting", C, N, N, N),
        ("B09", "MORPH", "Blood film morphology", C, N, N, N),
        ("B07", "MORPH", "Blood film morphology", C, N, N, N),
        ("B13", "COAG", "Routine coagulation", C, N, N, N),
        ("B10", "BT", "Crossmatch and electronic issue", C, N, N, N),
        ("B10", "HAEM", "Full blood count reporting", C, N, N, N),
        ("B10", "COAG", "Routine coagulation", C, N, N, N),
        ("B11", "BT", "Crossmatch and electronic issue", C, N, N, N),
        ("B11", "HAEM", "Full blood count reporting", C, N, N, N),
        ("B12", "HAEM", "Full blood count reporting", C, N, N, N),
        ("B12", "COAG", "Routine coagulation", C, N, N, N),
        ("B13", "BT", "Crossmatch and electronic issue", C, N, N, N),
        ("B13", "HAEM", "Full blood count reporting", C, N, N, N),
        ("B14", "HAEM", "Sample processing", IT, N, N, N),
    ]
    names = {person[0]: person[1] for person in _balanced_staff()}
    return [(staff_id, names.get(staff_id, ""), discipline, competency, status,
             achieved, None, later, trainer, assessor, authoriser, "")
            for staff_id, discipline, competency, status,
            trainer, assessor, authoriser in plan]


def _balanced_leave(period_start: date):
    """A little absence, so credited hours are demonstrated without causing gaps."""
    return [
        ("B07", "Gale Dummy", period_start + timedelta(days=7),
         period_start + timedelta(days=11), "A/L", "", "Booked leave"),
        ("B10", "Jules Draft", period_start + timedelta(days=17),
         period_start + timedelta(days=18), "A/L", "", "Booked leave"),
    ]


#: Requirements the balanced workforce can comfortably deliver.
BALANCED_REQUIREMENTS = [
    ("E", "Weekday", 1, 1, 0, 0, 0, 0, 0, 1, "", "", ""),
    # Five on the main day service, not four: four sections needing four distinct
    # competent people would leave no slack at all, so a single absence would put
    # the shift out of compliance.
    ("C", "Weekday", 5, 3, 1, 0, 0, 1, 0, 2,
     "BT:1, HAEM:1, COAG:1, MORPH:1", "BT:1", "Main day service"),
    ("L", "Weekday", 2, 1, 1, 0, 0, 0, 0, 1, "BT:1, HAEM:1", "BT:1", ""),
    # The out-of-hours lone worker must be registered and competent. Requiring
    # "senior" as well would restrict nights to a handful of people, tie them up
    # in night blocks and recovery days, and leave the day service short of the
    # very people it needs.
    ("N", "All", 1, 1, 0, 0, 0, 0, 0, 0, "BT:1, HAEM:1", "BT:1",
     "Lone working out of hours"),
    # A two-person weekend, with the senior on duty also coordinating. Three was
    # about four per cent more work than this workforce is contracted for, which
    # pushed most people slightly over their target hours.
    ("W", "Weekend", 2, 2, 1, 0, 0, 0, 0, 1, "BT:1, HAEM:1", "BT:1", ""),
]


def build_balanced_workbook(path) -> None:
    """A fictional laboratory that produces a strong draft with few problems.

    The challenging example is deliberately hard, which is useful for showing what
    Lab Shift Roster detects but misleading as a first impression: nobody should conclude
    that the tool normally reports dozens of critical errors.
    """
    _build_example(path, flavour="balanced")


def build_demo_workbook(path) -> None:
    """A complete fictional laboratory that triggers every kind of warning."""
    _build_example(path, flavour="challenging")


def _build_example(path, flavour: str) -> None:
    """Write one of the two demonstration laboratories.

    ``challenging`` is the stress test: thin morphology cover, an expired
    competency, part-time constraints and leave, so every kind of warning appears.
    ``balanced`` is a well-staffed department that should come out with few or no
    critical problems.
    """
    balanced = flavour == "balanced"
    workbook = Workbook()
    today = date.today()
    first = date(today.year + (today.month == 12), (today.month % 12) + 1, 1)
    last = (date(first.year + (first.month == 12), (first.month % 12) + 1, 1)
            - timedelta(days=1))

    _instructions_sheet(workbook, is_demo=True, balanced=balanced)
    _details_sheet(workbook, first, last, is_demo=True, balanced=balanced)
    _rules_sheet(workbook)

    label = ("BALANCED EXAMPLE" if balanced else "CHALLENGING EXAMPLE")
    sheet, header_row = _sheet(
        workbook, "Staff", STAFF_HEADERS, STAFF_WIDTHS,
        note=f"{label} — every person below is fictional. Replace with your own "
             f"staff before using this for real planning.")
    _write_rows(sheet, header_row, _with_initials(
        _balanced_staff() if balanced else _demo_staff(first)))
    # Every index moved right by one when Initials was inserted at
    # column 3. Written out rather than offset in a loop so a reader
    # can check them against STAFF_HEADERS by counting.
    _yes_no(sheet, header_row, [6, 7, 8, 9, *range(14, 21), 25, 26])

    sheet, header_row = _sheet(
        workbook, "Competencies", COMPETENCY_HEADERS, COMPETENCY_WIDTHS,
        note=f"{label} — fictional competencies. Only Competent, Trainer and "
             f"Assessor count as coverage.")
    _write_rows(sheet, header_row,
                _balanced_competencies(first) if balanced
                else _demo_competencies(first))
    _list_validation(sheet, header_row, 5, CompetencyStatus.ALL)
    _yes_no(sheet, header_row, [9, 10, 11])

    sheet, header_row = _sheet(
        workbook, "Week Patterns", PATTERN_HEADERS, PATTERN_WIDTHS,
        note=f"{label} — only needed for staff on alternating weeks.")
    if not balanced:
        _write_rows(sheet, header_row, _demo_patterns())
    _yes_no(sheet, header_row, list(range(3, 10)))

    sheet, header_row = _sheet(workbook, "Leave", LEAVE_HEADERS, LEAVE_WIDTHS,
                               note=f"{label} — fictional absence records.")
    _write_rows(sheet, header_row,
                _balanced_leave(first) if balanced else _demo_leave(first, last))
    _list_validation(sheet, header_row, 5,
                     [code for code, *_ in DEFAULT_LEAVE_TYPES])

    _configuration_sheets(
        workbook,
        requirements=BALANCED_REQUIREMENTS if balanced else DEFAULT_REQUIREMENTS)
    workbook.save(path)
