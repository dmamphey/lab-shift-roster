"""Reading a manager's workbook, and saying clearly what is wrong with it.

Two principles here:

* **Report everything at once.**  A manager correcting a workbook should not have
  to re-upload five times to discover five problems, so validation gathers all
  of them and hands them back together.
* **Explain in the manager's language.**  An error says which sheet, which row
  and what to do, not which Python key was missing.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from . import SCHEMA_VERSION
from .models import (
    CREDIT_FIXED, CREDIT_FROM_PATTERN, CREDIT_METHODS,
    Availability, Bench, Competency, CompetencyStatus, Config, LeaveEntry,
    LeaveType, Period, RosterDetails, Rules, ShiftRequirement, ShiftType, Staff,
    WEEKDAY_NAMES, WEEKDAY_SHORT,
)
from .timeutils import TimeError, parse_time

ERROR = "error"
WARNING = "warning"

REQUIRED_SHEETS = ["Roster Details", "Rules", "Staff", "Competencies",
                   "Shifts", "Shift Requirements", "Benches", "Leave"]

#: Sheets that identify a workbook produced by the first version of the tool.
LEGACY_MARKERS = {"settings"}


@dataclass
class Problem:
    severity: str
    sheet: str
    message: str
    row: int | None = None

    @property
    def location(self) -> str:
        if self.row:
            return f"{self.sheet}, row {self.row}"
        return self.sheet


class WorkbookError(Exception):
    """The workbook could not be read at all."""

    def __init__(self, message: str, problems: list[Problem] | None = None):
        super().__init__(message)
        self.problems = problems or []


def norm(text) -> str:
    return re.sub(r"[^a-z0-9]", "", str(text or "").lower())


def _to_float(value, default=0.0):
    if value in (None, ""):
        return default
    try:
        return float(str(value).strip().rstrip("%"))
    except (TypeError, ValueError):
        return default


def _to_int(value, default=0):
    return int(_to_float(value, default))


def _is_yes(value, default=False) -> bool:
    text = str(value).strip().lower()
    if text in ("y", "yes", "true", "1", "t", "x"):
        return True
    if text in ("n", "no", "false", "0", "f"):
        return False
    return default


def _to_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y",
                "%d %b %Y", "%d %B %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_counts(value) -> dict[str, int]:
    """Read 'BT:1, HAEM:2' into {'BT': 1, 'HAEM': 2}.

    A bare discipline with no number is taken to mean one.
    """
    result: dict[str, int] = {}
    for part in re.split(r"[,;|]", str(value or "")):
        part = part.strip()
        if not part:
            continue
        if ":" in part or "=" in part:
            key, count = re.split(r"[:=]", part, maxsplit=1)
            result[key.strip().upper()] = max(0, _to_int(count, 1))
        else:
            result[part.upper()] = 1
    return {k: v for k, v in result.items() if v}


class Reader:
    """Reads one workbook into a :class:`~labroster.models.Config`."""

    def __init__(self, source):
        self.problems: list[Problem] = []
        #: Shifts the workbook defines but the laboratory has switched off.
        #: Kept, rather than discarded, so that requirements and benches can
        #: tell "switched off" from "misspelt" — the first is deliberate and the
        #: second is a mistake worth reporting.
        self.inactive_shifts: list[ShiftType] = []
        try:
            self.workbook = load_workbook(source, data_only=True)
        except (zipfile.BadZipFile, InvalidFileException) as error:
            raise WorkbookError(
                "That file could not be opened as an Excel workbook. It may be a "
                "different type of file, an older .xls file, or damaged. Download "
                "a fresh template and fill that in."
            ) from error

    # -- helpers ---------------------------------------------------------

    def add(self, severity, sheet, message, row=None) -> None:
        self.problems.append(Problem(severity, sheet, message, row))

    def sheet(self, *names):
        for candidate in names:
            for title in self.workbook.sheetnames:
                if norm(title) == norm(candidate):
                    return self.workbook[title]
        return None

    def table(self, sheet, required_columns, sheet_label) -> list[dict]:
        """Find the header row and return rows as dictionaries.

        Headers are matched loosely, so capitalisation, spacing and column order
        do not matter.  A missing required column is reported by name.
        """
        if sheet is None:
            return []
        wanted = {norm(name): name for name in required_columns}
        for scan_row in range(1, 12):
            found = {}
            for cell in sheet[scan_row]:
                if cell.value not in (None, ""):
                    found[norm(cell.value)] = cell.column
            if all(key in found for key in wanted):
                return self._rows_from(sheet, scan_row, found)

        # Report exactly which columns are missing, using the manager's names.
        best_row, best_found = 1, {}
        for scan_row in range(1, 12):
            found = {norm(c.value): c.column for c in sheet[scan_row]
                     if c.value not in (None, "")}
            if len(found) > len(best_found):
                best_row, best_found = scan_row, found
        missing = [name for key, name in wanted.items() if key not in best_found]
        self.add(ERROR, sheet_label,
                 f"The {sheet_label} sheet is missing the "
                 f"{', '.join(repr(m) for m in missing)} column"
                 f"{'s' if len(missing) > 1 else ''}. Download the latest template, "
                 f"or add the column heading before generating the roster.",
                 row=best_row)
        return self._rows_from(sheet, best_row, best_found)

    @staticmethod
    def _rows_from(sheet, header_row, columns) -> list[dict]:
        rows = []
        for excel_row in range(header_row + 1, sheet.max_row + 1):
            record = {key: sheet.cell(row=excel_row, column=col).value
                      for key, col in columns.items()}
            if any(value not in (None, "") for value in record.values()):
                record["_row"] = excel_row
                rows.append(record)
        return rows

    # -- sections --------------------------------------------------------

    def read_details(self) -> tuple[RosterDetails, Period, int]:
        sheet = self.sheet("Roster Details", "Details")
        values: dict[str, object] = {}
        if sheet is not None:
            for row in sheet.iter_rows(min_row=1, max_row=40, values_only=True):
                if row and row[0] not in (None, "") and len(row) > 1:
                    values[norm(row[0])] = row[1]

        def get(key, default=None):
            value = values.get(norm(key))
            return default if value in (None, "") else value

        details = RosterDetails(
            rota_name=str(get("Rota name", "Laboratory Staff Rota")),
            organisation=str(get("Organisation", "") or ""),
            department=str(get("Department", "") or ""),
            site=str(get("Site", "") or ""),
            prepared_by=str(get("Prepared by", "") or ""),
        )
        version = _to_int(get("Workbook version", SCHEMA_VERSION), SCHEMA_VERSION)

        start = _to_date(get("Roster period start"))
        end = _to_date(get("Roster period end"))
        if start is None or end is None:
            self.add(ERROR, "Roster Details",
                     "The roster period start and end dates are needed. Enter them "
                     "on the Roster Details sheet as dd/mm/yyyy.")
            today = date.today()
            start = start or today
            end = end or today
        elif end < start:
            self.add(ERROR, "Roster Details",
                     f"The roster period ends ({end:%d/%m/%Y}) before it starts "
                     f"({start:%d/%m/%Y}). Correct the dates on the Roster Details "
                     f"sheet.")
            end = start
        elif (end - start).days > 400:
            self.add(WARNING, "Roster Details",
                     "The roster period is longer than a year, which will produce a "
                     "very wide spreadsheet. Consider generating one month or "
                     "quarter at a time.")

        return details, Period(start=start, end=end), version

    def read_rules(self) -> Rules:
        sheet = self.sheet("Rules", "Settings")
        values: dict[str, object] = {}
        if sheet is not None:
            for row in sheet.iter_rows(min_row=1, max_row=60, values_only=True):
                if row and row[0] not in (None, "") and len(row) > 1:
                    values[norm(row[0])] = row[1]

        def get(key, default=None):
            value = values.get(norm(key))
            return default if value in (None, "") else value

        weekend = set()
        for token in re.split(r"[,;|]", str(get("Weekend days", "Saturday, Sunday"))):
            token = norm(token)
            if not token:
                continue
            for index, name in enumerate(WEEKDAY_NAMES):
                if norm(name).startswith(token) and len(token) >= 3:
                    weekend.add(index)
        if not weekend:
            weekend = {5, 6}

        thresholds = [abs(_to_int(part)) for part
                      in re.split(r"[,;|]",
                                  str(get("Competency expiry warnings (days)",
                                          "30, 60, 90")))
                      if _to_int(part)]

        if values.get(norm("Senior band threshold")) not in (None, ""):
            self.add(WARNING, "Rules",
                     "'Senior band threshold' is no longer used. Seniority is "
                     "recorded per person in the Senior column on the Staff sheet, "
                     "and shifts needing a particular grade use Min Band on the "
                     "Requirements sheet. You can delete this row.")

        rules = Rules(
            minimum_rest_hours=_to_float(
                get("Minimum rest hours between shifts", 11), 11.0),
            max_consecutive_days=_to_int(get("Maximum consecutive days", 6), 6),
            max_consecutive_nights=_to_int(get("Maximum consecutive nights", 4), 4),
            night_block_length=max(1, _to_int(get("Night block length", 3), 3)),
            recovery_days_after_nights=max(
                0, _to_int(get("Recovery days after nights", 2), 2)),
            hours_tolerance_percent=_to_float(get("Hours tolerance (%)", 10), 10.0),
            expiry_warning_days=sorted(thresholds) or [30, 60, 90],
            max_simultaneous_bench_assignments=max(
                1, _to_int(get("Max simultaneous bench assignments", 1), 1)),
            cross_cover_allowed=_is_yes(get("Cross cover allowed", "N")),
            weekend_days=weekend,
            seed=_to_int(get("Alternative roster number",
                             get("Random seed", 42)), 42),
            rotation_warning_days=_to_int(
                get("Section rotation warning (days)", 56), 56),
            share_nights_evenly=_is_yes(get("Share nights evenly", "Y"), True),
        )
        if rules.minimum_rest_hours < 0:
            self.add(ERROR, "Rules",
                     "Minimum rest hours cannot be negative.")
            rules.minimum_rest_hours = 0.0
        return rules

    def read_staff(self, period: Period) -> list[Staff]:
        sheet = self.sheet("Staff")
        if sheet is None:
            self.add(ERROR, "Staff", "The workbook has no Staff sheet.")
            return []

        rows = self.table(sheet, ["Staff ID", "Name", "Contracted Weekly Hours"],
                          "Staff")
        staff: list[Staff] = []
        seen_ids: dict[str, int] = {}

        for record in rows:
            row_number = record.get("_row")
            staff_id = str(record.get("staffid") or "").strip()
            name = str(record.get("name") or "").strip()

            if not staff_id and not name:
                continue
            if not staff_id:
                self.add(ERROR, "Staff",
                         f"'{name}' has no Staff ID. Every member of staff needs a "
                         f"unique ID, because the other sheets refer to it.",
                         row=row_number)
                continue
            if not name:
                self.add(ERROR, "Staff",
                         f"Staff ID '{staff_id}' has no name.", row=row_number)
                continue
            if staff_id in seen_ids:
                self.add(ERROR, "Staff",
                         f"Staff ID '{staff_id}' is used twice, on rows "
                         f"{seen_ids[staff_id]} and {row_number}. IDs must be "
                         f"unique.", row=row_number)
                continue
            seen_ids[staff_id] = row_number

            weekly = _to_float(record.get("contractedweeklyhours"), 0.0)
            fte = _to_float(record.get("fte"), 0.0)
            if weekly < 0:
                self.add(ERROR, "Staff",
                         f"{name} has negative contracted weekly hours "
                         f"({weekly}). Enter the hours they are contracted to work.",
                         row=row_number)
                weekly = 0.0
            if not weekly and not fte:
                self.add(WARNING, "Staff",
                         f"{name} has neither contracted weekly hours nor an FTE, "
                         f"so their hours cannot be balanced. A full-time week will "
                         f"be assumed.", row=row_number)
            if fte and not 0 < fte <= 1.5:
                self.add(WARNING, "Staff",
                         f"{name} has an FTE of {fte}, which looks unlikely. FTE is "
                         f"normally between 0 and 1.", row=row_number)

            # Working pattern: the Mon–Sun columns.
            cycle = max(1, _to_int(record.get("patterncycleweeks"), 1))
            weekdays: set[int] = set()
            any_marked = False
            for index, short in enumerate(WEEKDAY_SHORT):
                value = record.get(norm(short))
                if value not in (None, ""):
                    any_marked = True
                    if _is_yes(value):
                        weekdays.add(index)

            earliest = latest = None
            for key, label in (("earlieststart", "Earliest Start"),
                               ("latestfinish", "Latest Finish")):
                raw = record.get(key)
                if raw in (None, ""):
                    continue
                try:
                    parsed = parse_time(raw, label)
                except TimeError as error:
                    self.add(ERROR, "Staff",
                             f"{name}: {error}", row=row_number)
                    continue
                if key == "earlieststart":
                    earliest = parsed
                else:
                    latest = parsed
            if earliest and latest and latest <= earliest:
                self.add(ERROR, "Staff",
                         f"{name} has a latest finish ({latest:%H:%M}) at or before "
                         f"their earliest start ({earliest:%H:%M}), so no shift "
                         f"could ever fit. Correct one of the two times.",
                         row=row_number)
                latest = None

            availability = Availability(
                cycle_weeks=cycle,
                weekdays={1: weekdays} if any_marked else {},
                earliest_start=earliest, latest_finish=latest,
                max_days_per_week=max(0, _to_int(record.get("maxdaysperweek"), 0)),
            )
            if any_marked and not weekdays:
                self.add(ERROR, "Staff",
                         f"{name} is marked as not working on any day of the week, "
                         f"so they can never be rostered. Mark at least one day, or "
                         f"clear all seven to make them fully flexible.",
                         row=row_number)

            person = Staff(
                staff_id=staff_id, name=name,
                # Optional: blank means "work it out from the name".
                initials=str(record.get("initials") or "").strip(),
                job_title=str(record.get("jobtitle") or "").strip(),
                band=str(record.get("band") or "").strip(),
                registered=_is_yes(record.get("registeredbms"), True),
                is_senior=_is_yes(record.get("senior"), False),
                shift_coordinator=_is_yes(record.get("shiftcoordinator"), False),
                trainee=_is_yes(record.get("trainee"), False),
                contracted_weekly_hours=weekly,
                fte=fte or 1.0,
                working_pattern=str(record.get("workingpattern") or "").strip(),
                max_period_hours=max(0.0, _to_float(
                    record.get("maxhoursthisperiod"), 0.0)),
                max_weekly_hours=max(0.0, _to_float(
                    record.get("maxweeklyhours"), 0.0)),
                availability=availability,
                nights_ok=_is_yes(record.get("worksnights"), True),
                weekends_ok=_is_yes(record.get("worksweekends"), True),
                max_nights=max(0, _to_int(record.get("maxnights"), 0)),
                max_weekends=max(0, _to_int(record.get("maxweekends"), 0)),
                max_consecutive_days=max(
                    0, _to_int(record.get("maxconsecutivedays"), 0)),
                restrictions=str(record.get("restrictions") or "").strip(),
                group=str(record.get("group") or "Main").strip() or "Main",
                notes=str(record.get("notes") or "").strip(),
            )
            staff.append(person)

        if not staff:
            self.add(ERROR, "Staff",
                     "The Staff sheet has no staff on it. Add at least one member "
                     "of staff, or download the example laboratory to see how it "
                     "should look.")
        self._read_week_patterns(staff)
        return staff

    def _read_week_patterns(self, staff: list[Staff]) -> None:
        sheet = self.sheet("Week Patterns", "WeekPatterns")
        if sheet is None:
            return
        rows = self.table(sheet, ["Staff ID", "Week"], "Week Patterns")
        by_id = {person.staff_id: person for person in staff}
        for record in rows:
            staff_id = str(record.get("staffid") or "").strip()
            if not staff_id:
                continue
            person = by_id.get(staff_id)
            if person is None:
                self.add(WARNING, "Week Patterns",
                         f"Staff ID '{staff_id}' is not on the Staff sheet, so this "
                         f"working pattern has been ignored.",
                         row=record.get("_row"))
                continue
            week = _to_int(record.get("week"), 1)
            if week < 1:
                self.add(ERROR, "Week Patterns",
                         f"Week must be 1 or higher for {person.name}.",
                         row=record.get("_row"))
                continue
            days = {index for index, short in enumerate(WEEKDAY_SHORT)
                    if _is_yes(record.get(norm(short)))}
            person.availability.weekdays[week] = days
            person.availability.cycle_weeks = max(
                person.availability.cycle_weeks, week)

    def read_competencies(self, staff: list[Staff],
                          period: Period) -> list[Competency]:
        sheet = self.sheet("Competencies", "Competency")
        if sheet is None:
            self.add(ERROR, "Competencies",
                     "The workbook has no Competencies sheet. Lab Shift Roster needs it to "
                     "know who can work in each laboratory section.")
            return []

        rows = self.table(sheet, ["Staff ID", "Discipline", "Status"],
                          "Competencies")
        known = {person.staff_id for person in staff}
        records: list[Competency] = []

        for record in rows:
            row_number = record.get("_row")
            staff_id = str(record.get("staffid") or "").strip()
            if not staff_id:
                continue
            if staff_id not in known:
                self.add(ERROR, "Competencies",
                         f"Staff ID '{staff_id}' is not on the Staff sheet. Add the "
                         f"member of staff, or correct the ID.", row=row_number)
                continue

            discipline = str(record.get("discipline") or "").strip().upper()
            if not discipline:
                self.add(ERROR, "Competencies",
                         f"A competency for '{staff_id}' has no discipline. Enter "
                         f"the laboratory section it applies to, for example BT.",
                         row=row_number)
                continue

            status = CompetencyStatus.normalise(record.get("status"))
            if status is None:
                self.add(ERROR, "Competencies",
                         f"'{record.get('status')}' is not a competency status. Use "
                         f"one of: {', '.join(CompetencyStatus.ALL)}.",
                         row=row_number)
                continue

            expiry = _to_date(record.get("expirydate"))
            achieved = _to_date(record.get("dateachieved"))
            if achieved and expiry and expiry < achieved:
                self.add(WARNING, "Competencies",
                         f"An expiry date of {expiry:%d/%m/%Y} is before the date "
                         f"achieved ({achieved:%d/%m/%Y}). Check the dates.",
                         row=row_number)

            records.append(Competency(
                staff_id=staff_id, discipline=discipline,
                name=str(record.get("competency") or "").strip(),
                status=status, date_achieved=achieved,
                review_date=_to_date(record.get("reviewdate")),
                expiry_date=expiry,
                trainer=_is_yes(record.get("trainer")),
                assessor=_is_yes(record.get("assessor")),
                authoriser=_is_yes(record.get("resultauthoriser")),
                notes=str(record.get("notes") or "").strip(),
            ))
        return records

    def read_shifts(self) -> list[ShiftType]:
        sheet = self.sheet("Shifts", "Shift")
        if sheet is None:
            self.add(ERROR, "Shifts", "The workbook has no Shifts sheet.")
            return []

        rows = self.table(sheet, ["Code", "Name", "Start", "End"], "Shifts")
        shifts: list[ShiftType] = []
        seen: dict[str, int] = {}

        for record in rows:
            row_number = record.get("_row")
            code = str(record.get("code") or "").strip()
            if not code:
                continue
            if code in seen:
                self.add(ERROR, "Shifts",
                         f"Shift code '{code}' appears twice, on rows "
                         f"{seen[code]} and {row_number}. Each shift needs its own "
                         f"code.", row=row_number)
                continue
            seen[code] = row_number

            try:
                start = parse_time(record.get("start"), f"'{code}' start time")
                end = parse_time(record.get("end"), f"'{code}' finish time")
            except TimeError as error:
                self.add(ERROR, "Shifts", str(error), row=row_number)
                continue
            if start == end:
                self.add(ERROR, "Shifts",
                         f"Shift '{code}' starts and finishes at "
                         f"{start:%H:%M}, so it has no length. Correct the times.",
                         row=row_number)
                continue

            shifts.append(ShiftType(
                code=code,
                name=str(record.get("name") or code).strip(),
                start=start, end=end,
                days=str(record.get("days") or "All").strip() or "All",
                # Absent column, or a blank cell, means the shift runs. A
                # workbook written before this column existed must keep working
                # exactly as it did, and "I left it empty" has to mean the same
                # as "yes" or every existing roster would silently empty itself.
                active=_is_yes(record.get("active"), default=True),
                is_night=_is_yes(record.get("nightshift")),
                colour=str(record.get("colour") or "D9E1F2").strip().lstrip("#").upper(),
                font_colour=str(record.get("fontcolour")
                                or "000000").strip().lstrip("#").upper(),
            ))

        # Split rather than filter, so the sheets that refer to shift codes can
        # tell a switched-off shift from a misspelt one.
        self.inactive_shifts = [shift for shift in shifts if not shift.active]
        active = [shift for shift in shifts if shift.active]

        if not shifts:
            self.add(ERROR, "Shifts",
                     "No shifts are defined, so no roster can be produced. Add the "
                     "shifts your laboratory runs, with their start and finish "
                     "times.")
        elif not active:
            # A different problem from an empty sheet, and worth its own words:
            # the shifts are all there, somebody has just turned every one off.
            self.add(ERROR, "Shifts",
                     "Every shift on the Shifts sheet is marked as not active, so "
                     "there is nothing to roster. Set Active to Yes for at least "
                     "one shift.")
        return active

    def read_requirements(self, shifts: list[ShiftType]) -> list[ShiftRequirement]:
        sheet = self.sheet("Shift Requirements", "ShiftRequirements",
                           "Requirements")
        if sheet is None:
            self.add(ERROR, "Shift Requirements",
                     "The workbook has no Shift Requirements sheet, so Lab Shift Roster "
                     "does not know how many staff each shift needs.")
            return []

        rows = self.table(sheet, ["Shift Code", "Min Staff"], "Shift Requirements")
        # `shifts` has already had the inactive ones filtered out, so a
        # code that is switched off is neither known nor a mistake — it is
        # its own third state, and both sheets below treat it as such.
        codes = {shift.code for shift in shifts}
        inactive_codes = {shift.code for shift in self.inactive_shifts}
        requirements: list[ShiftRequirement] = []

        for record in rows:
            row_number = record.get("_row")
            code = str(record.get("shiftcode") or "").strip()
            if not code:
                continue
            if code in inactive_codes:
                # Not an error, and deliberately not a warning either. Leaving
                # the requirement in place is the correct way to switch a shift
                # off for a while: it is still there, correct, and waiting for
                # the shift to come back. Complaining about it every run would
                # train people to ignore the warnings that matter.
                continue
            if code not in codes:
                self.add(ERROR, "Shift Requirements",
                         f"Shift code '{code}' is not on the Shifts sheet. Add the "
                         f"shift, or correct the code.", row=row_number)
                continue
            minimum = _to_int(record.get("minstaff"), 0)
            if minimum < 0:
                self.add(ERROR, "Shift Requirements",
                         f"Minimum staff for '{code}' cannot be negative.",
                         row=row_number)
                minimum = 0

            requirements.append(ShiftRequirement(
                shift_code=code,
                days=str(record.get("days") or "All").strip() or "All",
                min_staff=minimum,
                min_registered=max(0, _to_int(record.get("minregisteredbms"), 0)),
                min_senior=max(0, _to_int(record.get("minsenior"), 0)),
                min_band=_to_float(record.get("minband"), 0.0),
                min_at_band=max(0, _to_int(record.get("staffatminband"), 0)),
                min_coordinators=max(0, _to_int(record.get("mincoordinators"), 0)),
                min_trainers=max(0, _to_int(record.get("mintrainers"), 0)),
                max_trainees=max(0, _to_int(record.get("maxtrainees"), 0)),
                required_competencies=_parse_counts(
                    record.get("requiredcompetencies")),
                required_authorisers=_parse_counts(
                    record.get("requiredauthorisers")),
                notes=str(record.get("notes") or "").strip(),
            ))

        covered = {requirement.shift_code for requirement in requirements}
        for shift in shifts:
            if shift.code not in covered:
                self.add(WARNING, "Shift Requirements",
                         f"Shift '{shift.name}' has no entry on the Shift "
                         f"Requirements sheet, so no staff will be rostered onto it. "
                         f"Add a row saying how many people it needs.")
        return requirements

    def read_benches(self, shifts: list[ShiftType],
                     competencies: list[Competency]) -> list[Bench]:
        sheet = self.sheet("Benches", "Bench", "Sections")
        if sheet is None:
            return []
        rows = self.table(sheet, ["Bench", "Discipline"], "Benches")
        # `shifts` has already had the inactive ones filtered out, so a
        # code that is switched off is neither known nor a mistake — it is
        # its own third state, and both sheets below treat it as such.
        codes = {shift.code for shift in shifts}
        inactive_codes = {shift.code for shift in self.inactive_shifts}
        held = {record.discipline.upper() for record in competencies}
        benches: list[Bench] = []

        for record in rows:
            row_number = record.get("_row")
            name = str(record.get("bench") or "").strip()
            if not name:
                continue
            discipline = str(record.get("discipline") or "").strip().upper()
            if not discipline:
                self.add(ERROR, "Benches",
                         f"Bench '{name}' has no discipline, so Lab Shift Roster cannot "
                         f"tell who is competent to cover it.", row=row_number)
                continue
            if discipline not in held:
                self.add(WARNING, "Benches",
                         f"Nobody in the workforce holds a '{discipline}' "
                         f"competency, so {name} can never be covered. Check the "
                         f"discipline code matches the Competencies sheet.",
                         row=row_number)

            shift_codes = [part.strip() for part
                           in re.split(r"[,;|]", str(record.get("shiftcodes") or ""))
                           if part.strip()]
            # A reference to a shift that has been switched off is dropped
            # silently, for the same reason as on the requirements sheet.
            unknown = [code for code in shift_codes
                       if code not in codes and code not in inactive_codes]
            if unknown:
                self.add(WARNING, "Benches",
                         f"Bench '{name}' refers to shift code(s) "
                         f"{', '.join(unknown)} that are not on the Shifts sheet.",
                         row=row_number)
            shift_codes = [code for code in shift_codes if code in codes]

            weekend_min = record.get("minstaffweekend")
            benches.append(Bench(
                name=name, discipline=discipline,
                days=str(record.get("days") or "All").strip() or "All",
                min_staff=max(0, _to_int(record.get("minstaff"), 1)),
                min_weekend=(None if weekend_min in (None, "")
                             else max(0, _to_int(weekend_min, 1))),
                shift_codes=shift_codes,
                requires_authoriser=_is_yes(record.get("requiresauthoriser")),
                target_rotation_days=max(
                    0, _to_int(record.get("rotationintervaldays"), 0)),
            ))
        return benches

    def read_leave_types(self) -> dict[str, LeaveType]:
        sheet = self.sheet("Leave Types", "LeaveTypes")
        types: dict[str, LeaveType] = {}
        if sheet is not None:
            for record in self.table(sheet, ["Code", "Label"], "Leave Types"):
                code = str(record.get("code") or "").strip()
                if not code:
                    continue
                method = str(record.get("creditedhoursmethod")
                             or CREDIT_FROM_PATTERN).strip()
                matched = next((option for option in CREDIT_METHODS
                                if norm(option) == norm(method)), None)
                if matched is None:
                    self.add(WARNING, "Leave Types",
                             f"'{method}' is not a way of working out credited "
                             f"hours for '{code}'. Use one of: "
                             f"{', '.join(CREDIT_METHODS)}. The working pattern "
                             f"has been used.", row=record.get("_row"))
                    matched = CREDIT_FROM_PATTERN
                types[norm(code)] = LeaveType(
                    code=code,
                    label=str(record.get("label") or code).strip(),
                    colour=str(record.get("colour")
                               or "FFFF00").strip().lstrip("#").upper(),
                    font_colour=str(record.get("fontcolour")
                                    or "000000").strip().lstrip("#").upper(),
                    credits_hours=_is_yes(
                        record.get("countstowardscontractedhours"), True),
                    credited_method=matched,
                    fixed_daily_hours=max(0.0, _to_float(
                        record.get("fixedhoursperday"), 0.0)))

        for code, label in [("A/L", "Annual leave"), ("S/L", "Sickness absence"),
                            ("C/L", "Carers or compassionate leave"),
                            ("M/L", "Maternity or paternity leave"),
                            ("S/D", "Study or training day")]:
            types.setdefault(norm(code), LeaveType(code, label))
        return types

    def read_leave(self, staff: list[Staff], period: Period,
                   leave_types: dict[str, LeaveType]) -> list[LeaveEntry]:
        sheet = self.sheet("Leave", "Absence")
        if sheet is None:
            return []
        rows = self.table(sheet, ["Staff ID", "From", "To"], "Leave")
        known = {person.staff_id: person for person in staff}
        entries: list[LeaveEntry] = []

        for record in rows:
            row_number = record.get("_row")
            staff_id = str(record.get("staffid") or "").strip()
            if not staff_id:
                continue
            person = known.get(staff_id)
            if person is None:
                self.add(ERROR, "Leave",
                         f"Staff ID '{staff_id}' is not on the Staff sheet.",
                         row=row_number)
                continue

            start = _to_date(record.get("from"))
            end = _to_date(record.get("to")) or start
            if start is None:
                self.add(ERROR, "Leave",
                         f"The absence for {person.name} has no start date. Enter it "
                         f"as dd/mm/yyyy.", row=row_number)
                continue
            if end < start:
                self.add(ERROR, "Leave",
                         f"{person.name}'s absence ends ({end:%d/%m/%Y}) before it "
                         f"starts ({start:%d/%m/%Y}).", row=row_number)
                continue
            if end < period.start or start > period.end:
                self.add(WARNING, "Leave",
                         f"{person.name}'s absence from {start:%d/%m/%Y} to "
                         f"{end:%d/%m/%Y} is entirely outside the roster period, so "
                         f"it has no effect on this roster.", row=row_number)

            code = str(record.get("type") or "A/L").strip() or "A/L"
            if norm(code) not in leave_types:
                leave_types[norm(code)] = LeaveType(code, code)
                self.add(WARNING, "Leave",
                         f"'{code}' is not on the Leave Types sheet. It has been "
                         f"used as-is; add it to control how it is shown.",
                         row=row_number)

            explicit = record.get("creditedhours")
            entries.append(LeaveEntry(
                staff_id=staff_id, start=start, end=end,
                code=leave_types[norm(code)].code,
                reason=str(record.get("reason") or "").strip(),
                credited_hours=(None if explicit in (None, "")
                                else max(0.0, _to_float(explicit, 0.0)))))
        return entries

    # -- orchestration ---------------------------------------------------

    def check_format(self) -> None:
        """Recognise a workbook from the previous version rather than crashing."""
        titles = {norm(title) for title in self.workbook.sheetnames}
        has_new = norm("Roster Details") in titles and norm("Competencies") in titles
        looks_legacy = bool(LEGACY_MARKERS & titles) and not has_new

        if looks_legacy:
            raise WorkbookError(
                "This workbook was made with an earlier version of Lab Shift Roster. It "
                "does not contain the competency, contracted hours or shift "
                "requirement information that this version needs, and those cannot "
                "safely be guessed. Download the latest blank template and copy "
                "your staff into it.",
                problems=[Problem(ERROR, "Workbook",
                                  "Earlier workbook format detected: no "
                                  "'Roster Details' or 'Competencies' sheet.")])

        missing = [name for name in REQUIRED_SHEETS
                   if norm(name) not in titles]
        if missing:
            raise WorkbookError(
                "This workbook is missing sheet"
                f"{'s' if len(missing) > 1 else ''} that Lab Shift Roster needs: "
                f"{', '.join(missing)}. Download the latest template, or add the "
                f"missing sheet"
                f"{'s' if len(missing) > 1 else ''}.",
                problems=[Problem(ERROR, "Workbook",
                                  f"Missing sheet: {name}") for name in missing])

    def read(self) -> Config:
        self.check_format()
        details, period, version = self.read_details()
        rules = self.read_rules()
        staff = self.read_staff(period)
        competencies = self.read_competencies(staff, period)
        shifts = self.read_shifts()
        requirements = self.read_requirements(shifts)
        benches = self.read_benches(shifts, competencies)
        leave_types = self.read_leave_types()
        leave = self.read_leave(staff, period, leave_types)

        if version < SCHEMA_VERSION:
            self.add(WARNING, "Roster Details",
                     f"This workbook says it is version {version}; the current "
                     f"version is {SCHEMA_VERSION}. It has been read successfully, "
                     f"but a fresh template may contain new options.")

        return Config(details=details, period=period, rules=rules, staff=staff,
                      competencies=competencies, shifts=shifts,
                      requirements=requirements, benches=benches, leave=leave,
                      leave_types=leave_types, schema_version=version)

    @property
    def errors(self) -> list[Problem]:
        return [problem for problem in self.problems
                if problem.severity == ERROR]

    @property
    def warnings(self) -> list[Problem]:
        return [problem for problem in self.problems
                if problem.severity == WARNING]


def read_workbook(source) -> tuple[Config, list[Problem]]:
    """Read and validate a workbook, returning the config and every problem found.

    Raises :class:`WorkbookError` only when nothing useful can be read at all.
    Otherwise every problem is collected so a manager can fix them in one pass.
    """
    reader = Reader(source)
    config = reader.read()
    return config, reader.problems
