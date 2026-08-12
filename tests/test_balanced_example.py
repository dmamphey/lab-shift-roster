"""The balanced demonstration laboratory, and what it is there to prove.

A first-time user should not conclude that LabRoster normally reports dozens of
critical problems.  The challenging example is deliberately hard; this one is a
well-staffed department, and these tests keep it that way.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from labroster import api
from labroster.workbook import ERROR, read_workbook


@pytest.fixture(scope="module")
def balanced_bytes():
    return api.balanced_workbook_bytes()


@pytest.fixture(scope="module")
def balanced(balanced_bytes):
    result = api.generate(balanced_bytes)
    assert result["ok"], result.get("fatal")
    return result


@pytest.fixture(scope="module")
def challenging():
    return api.generate(api.demo_workbook_bytes())


# --- it reads cleanly -------------------------------------------------------

def test_the_balanced_workbook_has_no_validation_errors(balanced_bytes):
    config, problems = read_workbook(io.BytesIO(balanced_bytes))
    assert [p for p in problems if p.severity == ERROR] == []
    assert len(config.staff) == 14


def test_it_is_labelled_as_the_balanced_example(balanced_bytes):
    sheet = load_workbook(io.BytesIO(balanced_bytes))["Instructions"]
    text = " ".join(str(cell.value) for row in sheet.iter_rows()
                    for cell in row if cell.value is not None).lower()
    assert "balanced" in text
    assert "fictional" in text


def test_the_challenging_example_says_it_is_not_typical():
    sheet = load_workbook(io.BytesIO(api.demo_workbook_bytes()))["Instructions"]
    text = " ".join(str(cell.value) for row in sheet.iter_rows()
                    for cell in row if cell.value is not None).lower()
    assert "deliberately" in text
    assert "not typical" in text


# --- it demonstrates success ------------------------------------------------

def test_every_staffing_position_is_filled(balanced):
    assert balanced["dashboard"]["staffing_slot_coverage_percent"] == 100.0


def test_nearly_every_shift_meets_all_configured_requirements(balanced):
    assert balanced["dashboard"]["shifts_meeting_all_requirements_percent"] >= 95.0


def test_there_are_few_critical_problems(balanced):
    """The point of this fixture: a strong draft, not a wall of red."""
    assert balanced["dashboard"]["critical_count"] <= 5


def test_it_produces_far_fewer_criticals_than_the_challenging_example(
        balanced, challenging):
    assert balanced["dashboard"]["critical_count"] < \
        challenging["dashboard"]["critical_count"] / 3


def test_there_are_no_rest_rule_conflicts(balanced):
    assert balanced["dashboard"]["rest_conflicts"] == 0


def test_there_are_no_single_points_of_failure(balanced):
    assert balanced["dashboard"]["single_points_of_failure"] == 0
    for item in balanced["resilience"]:
        assert item["competent"] >= 3, \
            f"{item['discipline']} has only {item['competent']} competent staff"


def test_nobody_is_rostered_while_unavailable(balanced):
    breaches = [issue for issue in balanced["issues"]
                if issue["category"] == "Availability"
                and issue["severity"] == "CRITICAL"]
    assert breaches == []


def test_senior_cover_is_met_throughout(balanced):
    gaps = [issue for issue in balanced["issues"]
            if issue["category"] == "Senior cover"
            and issue["severity"] == "CRITICAL"]
    assert gaps == []


def test_contracted_hours_land_close_to_target(balanced):
    off_target = balanced["dashboard"]["staff_outside_target_hours"]
    assert off_target <= 4, f"{off_target} staff are outside their target hours"


def test_no_competency_has_expired(balanced):
    assert balanced["dashboard"]["competencies_expired"] == 0


def test_sections_are_covered_by_distinct_people(balanced):
    """The exclusivity guarantee must hold here too."""
    doubled = [issue for issue in balanced["issues"]
               if "more than one bench" in issue["title"].lower()]
    assert doubled == []


def test_absence_is_present_so_credited_hours_are_demonstrated(balanced):
    credited = [row for row in balanced["hours"] if row["credited"] > 0]
    assert credited, "the balanced example should include some absence"
    for row in credited:
        assert row["accounted"] == pytest.approx(row["worked"] + row["credited"])


# --- the night shift semantics this dataset exposed -------------------------

def test_a_single_handed_night_can_satisfy_two_competencies(balanced_bytes):
    """One person competent in both BT and HAEM covers "BT:1, HAEM:1".

    Sections need distinct people because they are physical stations. A shift's
    own competency list is not a set of stations, so requiring two distinct people
    for a one-person night shift could never be satisfied.
    """
    config, _ = read_workbook(io.BytesIO(balanced_bytes))
    night = next(shift for shift in config.shifts if shift.is_night)
    requirement = config.requirement_for(night, config.period.start)
    assert requirement.min_staff == 1
    assert len(requirement.required_competencies) >= 2

    result = api.generate(balanced_bytes)
    night_gaps = [issue for issue in result["issues"]
                  if issue["shift"] == night.code
                  and issue["severity"] == "CRITICAL"]
    assert night_gaps == [], \
        f"the single-handed night should be satisfiable: {night_gaps[:2]}"
