"""Switching a shift off, and showing staff initials on the roster.

Two changes driven by the same complaint: the workbook assumed every laboratory
works the pattern the template happens to seed, and the roster used initials in
the section rows without ever saying whose they were.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from labroster import api
from labroster.workbook import read_workbook


def _sheet_cell(sheet, header: str, row: int) -> str:
    for candidates in sheet.iter_rows():
        values = [str(item.value).strip() if item.value is not None else ""
                  for item in candidates]
        if header in values:
            return f"{candidates[values.index(header)].column_letter}{row}"
    raise AssertionError(f"no column headed {header!r} on {sheet.title!r}")


def _roster_sheet(result, audience: str = "manager"):
    """The exported Roster sheet for an already-generated result."""
    data = api.export_workbook(result["_scheduler"], result["_analysis"], audience)
    return load_workbook(io.BytesIO(data))["Roster"]


def _grid_bounds(sheet):
    """The first and last rows of the staff grid.

    The sheet continues past the staff into the section allocation block and the
    legend, and those rows reuse the same columns for other things. Scraping a
    column to `max_row` therefore collects bench disciplines and legend text as
    though they were people, which is how a test can fail for a reason that has
    nothing to do with what it is checking.
    """
    first = 5
    last = first - 1
    for row in range(first, sheet.max_row + 1):
        label = sheet.cell(row=row, column=1).value
        text = str(label).strip() if label else ""
        if text.upper() in ("SECTION ALLOCATION", "KEY", "LEGEND"):
            break
        if text:
            last = row
    return first, last


def _staff_initials(sheet):
    header = [str(cell.value).strip() if cell.value else "" for cell in sheet[3]]
    column = header.index("Initials") + 1
    first, last = _grid_bounds(sheet)
    out = []
    for row in range(first, last + 1):
        # Group banners span the width and leave the other columns empty.
        if sheet.cell(row=row, column=column).value:
            out.append(str(sheet.cell(row=row, column=column).value).strip())
    return out


def _section_initials(sheet):
    """Every initial used in the section allocation rows."""
    first_day_column = 5
    started = False
    codes = set()
    for row in range(1, sheet.max_row + 1):
        label = sheet.cell(row=row, column=1).value
        if str(label).strip().upper() == "SECTION ALLOCATION" if label else False:
            started = True
            continue
        if not started:
            continue
        for column in range(first_day_column, sheet.max_column + 1):
            value = sheet.cell(row=row, column=column).value
            if not isinstance(value, str):
                continue
            for part in value.split("/"):
                if part.strip():
                    codes.add(part.strip())
    return codes


def _modified(source: bytes, change) -> bytes:
    workbook = load_workbook(io.BytesIO(source))
    change(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _shift_rows(sheet):
    """Every shift row as (row number, code), skipping the header."""
    header = None
    out = []
    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [str(item).strip() if item is not None else "" for item in row]
        if header is None:
            if "Code" in values:
                header = index
            continue
        if values and values[0]:
            out.append((index, values[0]))
    return out


# --------------------------------------------------------------------------
# switching a shift off
# --------------------------------------------------------------------------

def test_the_template_offers_an_active_column():
    workbook = load_workbook(io.BytesIO(api.blank_template_bytes()))
    sheet = workbook["Shifts"]
    headers = [str(cell.value).strip() for cell in
               next(row for row in sheet.iter_rows()
                    if any(str(c.value).strip() == "Code" for c in row))]
    assert "Active" in headers
    # Seeded shifts ship switched on, so a workbook nobody edits behaves exactly
    # as it did before the column existed.
    column = _sheet_cell(sheet, "Active", 0)[0]
    for row, _code in _shift_rows(sheet):
        assert str(sheet[f"{column}{row}"].value).strip().upper() == "Y"


def test_a_missing_active_column_means_every_shift_runs():
    """Workbooks written before this column existed must be unaffected."""
    def drop_the_column(workbook):
        sheet = workbook["Shifts"]
        letter = _sheet_cell(sheet, "Active", 0)[0]
        sheet.delete_cols(ord(letter) - ord("A") + 1)

    source = _modified(api.demo_workbook_bytes(), drop_the_column)
    config, _problems = read_workbook(io.BytesIO(source))
    assert len(config.shifts) >= 4
    assert all(shift.active for shift in config.shifts)


def test_switching_a_shift_off_removes_it_from_the_roster():
    """The reported case: a laboratory that does not run an early shift."""
    def no_early(workbook):
        sheet = workbook["Shifts"]
        column = _sheet_cell(sheet, "Active", 0)[0]
        for row, code in _shift_rows(sheet):
            if code == "E":
                sheet[f"{column}{row}"] = "N"

    result = api.generate(_modified(api.demo_workbook_bytes(), no_early))
    assert result["ok"], result.get("problems")

    codes = set()
    for row in _roster_sheet(result).iter_rows(values_only=True):
        for value in row:
            if isinstance(value, str) and value.strip().rstrip("*") == "E":
                codes.add("E")
    assert "E" not in codes, "an inactive shift was still rostered"


def test_requirements_for_an_inactive_shift_are_not_an_error():
    """The thing that makes deactivation usable.

    The demo workbook has requirements and benches referring to the early shift.
    If switching the shift off turned those into errors, the manager would have
    to edit three sheets to remove one shift — which is the problem, not the fix.
    """
    def no_early(workbook):
        sheet = workbook["Shifts"]
        column = _sheet_cell(sheet, "Active", 0)[0]
        for row, code in _shift_rows(sheet):
            if code == "E":
                sheet[f"{column}{row}"] = "N"

    result = api.generate(_modified(api.demo_workbook_bytes(), no_early))
    assert result["ok"]
    text = " ".join(str(problem) for problem in result.get("problems", []))
    assert "'E' is not on the Shifts sheet" not in text
    assert "not on the Shifts sheet" not in text


def test_a_genuinely_unknown_shift_code_is_still_an_error():
    """Switched off and misspelt are different, and only one is a mistake."""
    def typo(workbook):
        sheet = workbook["Shift Requirements"]
        letter = _sheet_cell(sheet, "Shift Code", 0)[0]
        sheet[f"{letter}3"] = "ZZZ"

    result = api.generate(_modified(api.demo_workbook_bytes(), typo))
    text = " ".join(str(problem) for problem in result.get("problems", []))
    assert "ZZZ" in text


def test_switching_every_shift_off_is_reported_clearly():
    def all_off(workbook):
        sheet = workbook["Shifts"]
        column = _sheet_cell(sheet, "Active", 0)[0]
        for row, _code in _shift_rows(sheet):
            sheet[f"{column}{row}"] = "N"

    result = api.generate(_modified(api.demo_workbook_bytes(), all_off))
    assert not result["ok"]
    text = " ".join(str(problem) for problem in result.get("problems", []))
    assert "not active" in text.lower()


# --------------------------------------------------------------------------
# initials
# --------------------------------------------------------------------------

def test_the_roster_names_the_initials_it_uses():
    """The reported confusion: codes in the section rows with nothing to match."""
    result = api.generate(api.demo_workbook_bytes())
    assert result["ok"]
    sheet = _roster_sheet(result)

    header = [str(cell.value).strip() if cell.value else ""
              for cell in sheet[3]]
    assert "Initials" in header, header[:6]

    assert _staff_initials(sheet), "the Initials column is empty"


def test_supplied_initials_are_used_exactly_as_written():
    def set_initials(workbook):
        sheet = workbook["Staff"]
        sheet[_sheet_cell(sheet, "Initials", 3)] = "ZQX"

    result = api.generate(_modified(api.demo_workbook_bytes(), set_initials))
    assert result["ok"]
    sheet = _roster_sheet(result)
    values = {str(cell.value).strip()
              for row in sheet.iter_rows() for cell in row if cell.value}
    assert "ZQX" in values, "a laboratory's own initials were overridden"


def test_derived_initials_stay_unique():
    """Two people who would collide must not end up with the same label."""
    def clashing_names(workbook):
        sheet = workbook["Staff"]
        sheet[_sheet_cell(sheet, "Name", 3)] = "Sam Trent"
        sheet[_sheet_cell(sheet, "Name", 4)] = "Sara Thompson"

    result = api.generate(_modified(api.demo_workbook_bytes(), clashing_names))
    assert result["ok"]
    codes = _staff_initials(_roster_sheet(result))
    assert len(codes) == len(set(codes)), f"duplicate initials: {codes}"


def test_the_initials_column_matches_the_section_rows():
    """The column exists to explain the codes, so it must be the same strings."""
    result = api.generate(api.demo_workbook_bytes())
    sheet = _roster_sheet(result)
    grid_codes = set(_staff_initials(sheet))
    section_codes = _section_initials(sheet)

    unexplained = section_codes - grid_codes
    assert not unexplained, (
        f"the section rows use initials that appear nowhere in the grid: "
        f"{sorted(unexplained)}")
