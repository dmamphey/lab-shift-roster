"""Workbook reading, validation messages, and the two templates."""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook, load_workbook

from labroster import api
from labroster.workbook import ERROR, WARNING, WorkbookError, read_workbook


def cell(sheet, header: str, row: int) -> str:
    """The cell for a named column, found by its heading.

    Tests used to address these by letter, which meant that inserting a column
    into the Staff sheet silently moved every assertion one place to the left —
    two tests started writing to the wrong field and passed for the wrong reason
    until they happened to fail. Looking the heading up costs nothing and cannot
    drift.
    """
    for candidates in sheet.iter_rows():
        values = [str(item.value).strip() if item.value is not None else ""
                  for item in candidates]
        if header in values:
            return f"{candidates[values.index(header)].column_letter}{row}"
    raise AssertionError(f"no column headed {header!r} on sheet {sheet.title!r}")



# --------------------------------------------------------------------------
# the two templates
# --------------------------------------------------------------------------

@pytest.fixture(scope="module")
def blank_bytes():
    return api.blank_template_bytes()


@pytest.fixture(scope="module")
def demo_bytes():
    return api.demo_workbook_bytes()


@pytest.fixture(scope="module")
def demo_result(demo_bytes):
    return api.generate(demo_bytes)


def test_both_templates_open_as_workbooks(blank_bytes, demo_bytes):
    for data in (blank_bytes, demo_bytes):
        assert data[:2] == b"PK"
        load_workbook(io.BytesIO(data))


def test_the_blank_template_has_no_staff(blank_bytes):
    sheet = load_workbook(io.BytesIO(blank_bytes))["Staff"]
    values = [cell.value for row in sheet.iter_rows(min_row=3)
              for cell in row if cell.value not in (None, "")]
    assert values == [], "the blank template must contain no employee records"


def test_the_blank_template_has_no_competencies_or_leave(blank_bytes):
    workbook = load_workbook(io.BytesIO(blank_bytes))
    for name in ("Competencies", "Leave", "Week Patterns"):
        values = [cell.value for row in workbook[name].iter_rows(min_row=3)
                  for cell in row if cell.value not in (None, "")]
        assert values == [], f"{name} should be empty in the blank template"


def test_the_blank_template_still_has_headings_and_instructions(blank_bytes):
    workbook = load_workbook(io.BytesIO(blank_bytes))
    assert "Instructions" in workbook.sheetnames
    headings = [cell.value for cell in workbook["Staff"][2]]
    assert "Staff ID" in headings
    assert "Contracted Weekly Hours" in headings


def test_no_workbook_mentions_command_lines_or_python():
    """Manager-facing instructions must not contain developer instructions."""
    banned = ["python", "lab_roster.py", "pip install", "--input", "--out",
              "command", "pyodide", "openpyxl", "webassembly"]
    for data in (api.blank_template_bytes(), api.demo_workbook_bytes()):
        workbook = load_workbook(io.BytesIO(data))
        text = " ".join(
            str(cell.value).lower()
            for name in workbook.sheetnames
            for row in workbook[name].iter_rows()
            for cell in row if cell.value is not None)
        for word in banned:
            assert word not in text, f"'{word}' appears in a user-facing workbook"


def test_the_default_rota_name_is_organisation_neutral(blank_bytes):
    sheet = load_workbook(io.BytesIO(blank_bytes))["Roster Details"]
    values = [str(cell.value) for row in sheet.iter_rows() for cell in row
              if cell.value is not None]
    assert "Laboratory Staff Rota" in values
    joined = " ".join(values)
    assert "GEH" not in joined and "BMS Rota" not in joined


def test_the_demo_workbook_is_labelled_as_fictional(demo_bytes):
    sheet = load_workbook(io.BytesIO(demo_bytes))["Instructions"]
    text = " ".join(str(cell.value) for row in sheet.iter_rows()
                    for cell in row if cell.value is not None).lower()
    assert "fictional" in text


# --------------------------------------------------------------------------
# reading the demo end to end
# --------------------------------------------------------------------------

def test_the_demo_workbook_generates_a_roster(demo_result):
    assert demo_result["ok"], demo_result.get("fatal")
    assert demo_result["dashboard"]["total_assignments"] > 0
    assert api.export_workbook(demo_result["_scheduler"],
                               demo_result["_analysis"])[:2] == b"PK"


def test_the_demo_reads_without_blocking_errors(demo_bytes):
    config, problems = read_workbook(io.BytesIO(demo_bytes))
    assert [p for p in problems if p.severity == ERROR] == []
    assert len(config.staff) == 15
    assert config.competencies
    assert config.requirements


def test_the_demo_exercises_every_kind_of_warning(demo_result):
    """The example laboratory should demonstrate each category of issue."""
    categories = {issue["category"] for issue in demo_result["issues"]}
    for expected in ["Bench coverage", "Competency", "Contracted hours",
                     "Workforce resilience", "Shift resilience", "Rest",
                     "Availability"]:
        assert expected in categories, f"demo data does not trigger {expected}"


def test_the_demo_contains_all_three_severities(demo_result):
    severities = {issue["severity"] for issue in demo_result["issues"]}
    assert severities == {"CRITICAL", "REVIEW", "PASSED"}


def test_the_demo_has_a_deliberate_single_point_of_failure(demo_result):
    critical = [item for item in demo_result["resilience"]
                if item["severity"] == "CRITICAL"]
    assert critical, "the example should show at least one single point of failure"
    assert critical[0]["competent"] == 1


def test_the_demo_shows_an_expired_and_an_expiring_competency(demo_result):
    states = {item["state"] for item in demo_result["expiring"]}
    assert "Expired" in states
    assert any(state.startswith("Within") for state in states)


def test_the_demo_includes_part_time_staff_with_hours_below_full_time(demo_bytes):
    config, _ = read_workbook(io.BytesIO(demo_bytes))
    weekly = [person.contracted_weekly_hours for person in config.staff]
    assert min(weekly) < 37.5, "the example should include part-time staff"
    assert any(person.availability.cycle_weeks > 1 for person in config.staff), \
        "the example should include an alternating week pattern"


def test_nobody_in_the_demo_is_rostered_while_on_leave(demo_result):
    breaches = [issue for issue in demo_result["issues"]
                if issue["category"] == "Availability"
                and issue["severity"] == "CRITICAL"]
    assert breaches == []


def test_the_exported_workbook_has_the_expected_sheets(demo_result):
    workbook = load_workbook(io.BytesIO(api.export_workbook(
        demo_result["_scheduler"], demo_result["_analysis"], "manager")))
    for name in ["Instructions", "Roster", "Staff", "Competencies",
                 "Shift Requirements", "Bench Allocations", "Issues",
                 "Hours Summary", "Fairness Summary", "Competency Expiry"]:
        assert name in workbook.sheetnames, f"{name} missing from the export"


def test_the_exported_roster_is_marked_as_a_draft(demo_result):
    workbook = load_workbook(io.BytesIO(api.export_workbook(
        demo_result["_scheduler"], demo_result["_analysis"], "manager")))
    text = " ".join(str(cell.value) for row in workbook["Roster"].iter_rows()
                    for cell in row if cell.value is not None).lower()
    assert "draft" in text


def test_the_same_alternative_number_reproduces_the_same_roster(demo_bytes):
    first = api.generate(demo_bytes, alternative=5)
    second = api.generate(demo_bytes, alternative=5)
    assert first["dashboard"] == second["dashboard"]


def test_a_different_alternative_number_changes_the_roster(demo_bytes):
    first = api.generate(demo_bytes, alternative=1)
    second = api.generate(demo_bytes, alternative=99)
    rosters = []
    for result in (first, second):
        rosters.append([(row["name"], sorted(
            (iso, cell["code"]) for iso, cell in row["cells"].items()))
            for row in result["roster"]["rows"]])
    assert rosters[0] != rosters[1]


# --------------------------------------------------------------------------
# validation messages
# --------------------------------------------------------------------------

def _modified_demo(mutate):
    data = api.demo_workbook_bytes()
    workbook = load_workbook(io.BytesIO(data))
    mutate(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _problems(data):
    result = api.generate(data)
    return result, [p["message"] for p in result.get("problems", [])]


def test_a_non_spreadsheet_is_explained_not_crashed():
    result = api.generate(b"this is not a spreadsheet at all")
    assert not result["ok"]
    assert "Excel workbook" in result["fatal"]


def test_an_older_workbook_format_is_recognised():
    """A v1 workbook must be explained, not silently given invented defaults."""
    workbook = Workbook()
    workbook.active.title = "Settings"
    workbook.create_sheet("Staff")
    workbook.create_sheet("Shifts")
    workbook.create_sheet("Leave")
    buffer = io.BytesIO()
    workbook.save(buffer)
    result = api.generate(buffer.getvalue())
    assert not result["ok"]
    assert "earlier version" in result["fatal"].lower()


def test_a_missing_sheet_is_named():
    data = _modified_demo(lambda wb: wb.remove(wb["Competencies"]))
    result = api.generate(data)
    assert not result["ok"]
    assert "Competencies" in result["fatal"]


def test_a_missing_column_names_the_column_not_a_python_key():
    def drop_hours(workbook):
        sheet = workbook["Staff"]
        for cell in sheet[2]:
            if cell.value == "Contracted Weekly Hours":
                cell.value = None
    result, messages = _problems(_modified_demo(drop_hours))
    assert not result["ok"]
    joined = " ".join(messages)
    assert "Contracted Weekly Hours" in joined
    assert "KeyError" not in joined


def test_duplicate_staff_ids_are_reported_with_both_rows():
    def duplicate(workbook):
        staff = workbook["Staff"]
        staff[cell(staff, "Staff ID", 4)] = staff[cell(staff, "Staff ID", 3)].value
    result, messages = _problems(_modified_demo(duplicate))
    assert not result["ok"]
    assert any("used twice" in message for message in messages)


def test_negative_contracted_hours_are_reported():
    def negative(workbook):
        staff = workbook["Staff"]
        staff[cell(staff, "Contracted Weekly Hours", 3)] = -10
    result, messages = _problems(_modified_demo(negative))
    assert not result["ok"]
    assert any("negative" in message.lower() for message in messages)


def test_an_invalid_competency_status_is_reported_with_the_valid_options():
    def bad_status(workbook):
        workbook["Competencies"]["E3"] = "Sort of competent"
    result, messages = _problems(_modified_demo(bad_status))
    assert not result["ok"]
    assert any("not a competency status" in message for message in messages)
    assert any("Competent" in message for message in messages)


def test_an_unknown_staff_id_on_the_competency_sheet_is_reported():
    def unknown(workbook):
        workbook["Competencies"]["A3"] = "ZZZ"
    result, messages = _problems(_modified_demo(unknown))
    assert not result["ok"]
    assert any("ZZZ" in message for message in messages)


def test_impossible_availability_is_reported():
    def no_days(workbook):
        for column in "MNOPQRS":                      # the Mon–Sun columns
            workbook["Staff"][f"{column}3"] = "N"
    result, messages = _problems(_modified_demo(no_days))
    assert not result["ok"]
    assert any("never be rostered" in message for message in messages)


def test_a_finish_time_before_a_start_time_is_reported():
    def impossible(workbook):
        staff = workbook["Staff"]
        staff[cell(staff, "Earliest Start", 3)] = "18:00"
        staff[cell(staff, "Latest Finish", 3)] = "09:00"
    result, messages = _problems(_modified_demo(impossible))
    assert not result["ok"]
    assert any("earliest start" in message.lower() for message in messages)


def test_a_zero_length_shift_is_reported():
    def zero_length(workbook):                    # data starts on row 3
        shifts = workbook["Shifts"]
        shifts[cell(shifts, "End", 3)] = shifts[cell(shifts, "Start", 3)].value
    result, messages = _problems(_modified_demo(zero_length))
    assert not result["ok"]
    assert any("no length" in message or "starts and finishes" in message
               for message in messages)


def test_a_duplicate_shift_code_is_reported():
    def duplicate(workbook):                      # data starts on row 3
        shifts = workbook["Shifts"]
        shifts[cell(shifts, "Code", 4)] = shifts[cell(shifts, "Code", 3)].value
    result, messages = _problems(_modified_demo(duplicate))
    assert not result["ok"]
    assert any("appears twice" in message for message in messages)


def test_leave_dates_the_wrong_way_round_are_reported():
    def reversed_dates(workbook):
        sheet = workbook["Leave"]                     # data starts on row 3
        sheet["C3"], sheet["D3"] = sheet["D3"].value, sheet["C3"].value
    result, messages = _problems(_modified_demo(reversed_dates))
    assert not result["ok"]
    assert any("before it starts" in message for message in messages)


def test_a_roster_period_the_wrong_way_round_is_reported():
    def reversed_period(workbook):
        sheet = workbook["Roster Details"]
        sheet["B7"], sheet["B8"] = sheet["B8"].value, sheet["B7"].value
    result, messages = _problems(_modified_demo(reversed_period))
    assert not result["ok"]


def test_every_problem_is_reported_at_once_not_one_at_a_time():
    """A manager should be able to fix everything in a single pass."""
    def several(workbook):
        staff = workbook["Staff"]
        staff[cell(staff, "Contracted Weekly Hours", 3)] = -5
        staff[cell(staff, "Staff ID", 4)] = staff[cell(staff, "Staff ID", 3)].value
        workbook["Competencies"]["E3"] = "Nonsense"       # bad status
    result, messages = _problems(_modified_demo(several))
    assert not result["ok"]
    assert len(messages) >= 3, f"expected several problems, got {messages}"


def test_an_unknown_discipline_on_the_benches_sheet_is_a_warning_not_an_error():
    def unknown_discipline(workbook):
        workbook["Benches"]["B3"] = "NOSUCH"          # first data row
    data = _modified_demo(unknown_discipline)
    result = api.generate(data)
    assert result["ok"], "an uncoverable bench should warn, not block generation"
    assert any("NOSUCH" in p["message"] for p in result["problems"])


# --------------------------------------------------------------------------
# the two export audiences
# --------------------------------------------------------------------------

def _exports(demo_result):
    scheduler = demo_result["_scheduler"]
    analysis = demo_result["_analysis"]
    return (api.export_workbook(scheduler, analysis, "staff"),
            api.export_workbook(scheduler, analysis, "manager"))


def _all_text(data):
    workbook = load_workbook(io.BytesIO(data))
    return " ".join(str(cell.value) for name in workbook.sheetnames
                    for row in workbook[name].iter_rows()
                    for cell in row if cell.value is not None)


def test_the_staff_rota_contains_only_what_staff_need(demo_result):
    staff, _ = _exports(demo_result)
    workbook = load_workbook(io.BytesIO(staff))
    assert workbook.sheetnames == ["Roster", "Section Allocations", "Notes"]


def test_the_manager_report_keeps_the_full_analysis(demo_result):
    _, manager = _exports(demo_result)
    workbook = load_workbook(io.BytesIO(manager))
    for name in ["Instructions", "Roster", "Staff", "Competencies",
                 "Shift Requirements", "Bench Allocations", "Issues",
                 "Hours Summary", "Fairness Summary", "Competency Expiry"]:
        assert name in workbook.sheetnames


def test_the_staff_rota_does_not_disclose_management_information(demo_result):
    """Data minimisation: a circulated rota must not leak these."""
    staff, _ = _exports(demo_result)
    text = _all_text(staff).lower()
    for forbidden in ["competent", "in training", "expired", "variance",
                      "fairness", "single point", "resilience", "restrictions",
                      "target hours", "authoriser", "trainee"]:
        assert forbidden not in text, \
            f"the staff rota discloses '{forbidden}'"


def _cell_values(data):
    """Every cell value, so a check can match whole values rather than substrings."""
    workbook = load_workbook(io.BytesIO(data))
    return [str(cell.value).strip() for name in workbook.sheetnames
            for row in workbook[name].iter_rows()
            for cell in row if cell.value is not None]


@pytest.mark.parametrize("data_source", ["challenging", "balanced"])
def test_the_staff_rota_does_not_reveal_why_somebody_is_absent(data_source):
    """Showing S/L on a departmental rota would disclose sickness.

    Matched against whole cell values, not as substrings: the section allocation
    row holds initials, and a pair such as Ada Sample and Lior Proxy renders as
    "AS/LP", which contains "S/L" without disclosing anything. Substring matching
    here would fail on that and pass for the wrong reasons elsewhere.
    """
    workbook_bytes = (api.demo_workbook_bytes() if data_source == "challenging"
                      else api.balanced_workbook_bytes())
    result = api.generate(workbook_bytes)
    staff = api.export_workbook(result["_scheduler"], result["_analysis"], "staff")

    values = set(_cell_values(staff))
    for code in ["S/L", "A/L", "M/L", "C/L", "S/D"]:
        assert code not in values, \
            f"absence code '{code}' appears as a cell value in the staff rota"

    # And the manager export, for contrast, does record them.
    manager = api.export_workbook(result["_scheduler"], result["_analysis"],
                                  "manager")
    assert "A/L" in set(_cell_values(manager))


def test_the_manager_report_does_still_show_absence_codes(demo_result):
    _, manager = _exports(demo_result)
    assert "A/L" in _all_text(manager)


def test_the_staff_rota_still_shows_who_is_working(demo_result):
    staff, _ = _exports(demo_result)
    workbook = load_workbook(io.BytesIO(staff))
    sheet = workbook["Roster"]
    codes = {str(cell.value) for row in sheet.iter_rows(min_row=5)
             for cell in row if cell.value in ("C", "E", "L", "N", "W")}
    assert codes, "the staff rota should still show shift assignments"


def test_the_staff_rota_carries_the_organisation_and_period(demo_result):
    staff, _ = _exports(demo_result)
    text = _all_text(staff)
    assert "Example NHS Trust" in text
    assert "Blood Sciences" in text
    assert "Period" in text


def test_the_staff_rota_is_not_labelled_a_draft_for_review(demo_result):
    """The manager reviews the draft; staff receive the agreed rota."""
    staff, manager = _exports(demo_result)
    assert "DRAFT for review" not in _all_text(staff)
    assert "DRAFT for review" in _all_text(manager)


def test_both_exports_come_from_one_generated_draft(demo_result):
    """Exporting twice must not reschedule, or manual changes would be lost."""
    scheduler = demo_result["_scheduler"]
    before = dict(scheduler.assignments)
    api.export_workbook(scheduler, demo_result["_analysis"], "staff")
    api.export_workbook(scheduler, demo_result["_analysis"], "manager")
    assert dict(scheduler.assignments) == before


def test_both_exports_open_as_valid_workbooks(demo_result):
    for data in _exports(demo_result):
        assert data[:2] == b"PK"
        load_workbook(io.BytesIO(data))


def test_the_staff_rota_identifies_what_produced_it(demo_result):
    """A circulated rota should say where it came from, without saying more."""
    staff, _ = _exports(demo_result)
    text = _all_text(staff)
    assert "Lab Shift Roster" in text
    assert "free, secure and intelligent" in text


@pytest.mark.parametrize("builder", ["blank_template_bytes",
                                     "balanced_workbook_bytes",
                                     "demo_workbook_bytes"])
def test_every_workbook_carries_the_product_name_and_tagline(builder):
    data = getattr(api, builder)()
    text = _all_text(data)
    assert "Lab Shift Roster" in text, f"{builder} does not name the product"
    assert "free, secure and intelligent" in text, \
        f"{builder} does not carry the tagline"
    assert "LabRoster" not in text.replace("Lab Shift Roster", ""), \
        f"{builder} still uses the old name"
